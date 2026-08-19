# -*- coding: utf-8 -*-
import json
import logging
import time

from datetime import datetime as _datetime, timedelta as _timedelta

from odoo import http, fields
from odoo.http import request

_logger = logging.getLogger(__name__)

try:
    import openai
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False

try:
    from google import genai as new_genai
    NEW_GEMINI_SDK = True
except ImportError:
    NEW_GEMINI_SDK = False

GEMINI_AVAILABLE = NEW_GEMINI_SDK

from .ai_definitions import (
    TOOLS_OPENAI,
    DESTRUCTIVE_TOOLS,
    _VALID_GEMINI_MODELS,
    _DEFAULT_GEMINI_MODEL,
    _normalize_gemini_model,
    _redact,
    _parse_gemini_error,
)


class EstateAIController(http.Controller):

    # Modelos por defecto cuando no se ha especificado uno para el proveedor.
    _DEFAULT_MODELS = {'chatgpt': 'gpt-4o-mini', 'gemini': 'gemini-2.5-flash'}

    def _ai_provider_chain(self, ICP):
        """Construye la lista de proveedores a intentar (activo + respaldo).

        Devuelve una lista de tuplas (proveedor, api_key, modelo). El proveedor
        configurado va primero; el otro se añade como respaldo solo si tiene su
        propia API Key. Se mantiene compatibilidad con la clave/modelo heredados
        (estate_ai.api_key / estate_ai.model) para el proveedor activo.
        """
        active = ICP.get_param('estate_ai.provider', 'chatgpt')
        # El valor del proveedor 'chatgpt' usa el prefijo de parámetros 'openai'.
        prefixes = {'chatgpt': 'openai', 'gemini': 'gemini'}

        def creds(p):
            px = prefixes.get(p, p)
            key = ICP.get_param('estate_ai.%s_api_key' % px, '') or ''
            model = ICP.get_param('estate_ai.%s_model' % px, '') or ''
            if p == active:  # respaldo a la config heredada (instalaciones antiguas)
                key = key or ICP.get_param('estate_ai.api_key', '') or ''
                model = model or ICP.get_param('estate_ai.model', '') or ''
            model = model or self._DEFAULT_MODELS.get(p, '')
            if p == 'gemini':
                model = _normalize_gemini_model(model)
            return key, model

        order = [active] + [p for p in ('chatgpt', 'gemini') if p != active]
        chain = []
        for p in order:
            key, model = creds(p)
            if key:
                chain.append((p, key, model))
        return chain

    def _friendly_ai_error(self, provider, err):
        """Traduce un fallo del proveedor de IA en un mensaje claro para el usuario."""
        msg = (str(err) if err else '').lower()
        if any(k in msg for k in ('quota', 'insufficient_quota', 'rate limit',
                                  'rate_limit', '429', 'credit', 'billing', 'exceeded')):
            return ('El asistente de IA no está disponible temporalmente porque se agotó el crédito o '
                    'la cuota del proveedor. El resto del sistema funciona con normalidad; reintenta '
                    'más tarde o configura otro proveedor en Configuración > Ajustes > Agente IA.')
        if any(k in msg for k in ('api key', 'api_key', 'unauthorized', '401', 'invalid', 'permission',
                                  'authentication', 'forbidden', '403')):
            return ('El asistente de IA no pudo autenticarse con el proveedor (API Key inválida o sin '
                    'permisos). Verifica la clave en Configuración > Ajustes > Agente IA.')
        if any(k in msg for k in ('network', 'timeout', 'connection', 'dns', 'unreachable',
                                  'temporarily', '503', '502', '504')):
            return ('El asistente de IA no pudo conectarse con el proveedor (problema de red). '
                    'Revisa la conexión del servidor e inténtalo de nuevo en unos minutos.')
        return ('El asistente de IA no está disponible en este momento. El resto del sistema funciona '
                'con normalidad; inténtalo de nuevo más tarde.')

    @http.route('/estate_ai/chat', type='jsonrpc', auth='user', methods=['POST'])
    def chat(self, message, **kwargs):
        """Process a chat message and return AI response with memory and tool calling."""
        start_time = time.time()

        ICP = request.env['ir.config_parameter'].sudo()
        ai_active = ICP.get_param('estate_ai.active', 'True')
        if ai_active != 'True':
            return {'response': 'El agente IA está desactivado. Contacte al administrador.'}

        # Cadena de proveedores a intentar: el activo primero y, como respaldo,
        # el otro si tiene clave configurada (modo degradado / alta disponibilidad).
        provider_chain = self._ai_provider_chain(ICP)

        temperature = float(ICP.get_param('estate_ai.temperature', '0.7'))
        max_tokens = int(ICP.get_param('estate_ai.max_tokens', '1500'))
        system_prompt = ICP.get_param('estate_ai.system_prompt', '')

        if not provider_chain:
            return {'response': 'No se ha configurado la API Key de ningún proveedor de IA. '
                                'Vaya a Configuración > Ajustes > Agente IA.'}

        context_data = self._get_system_context()
        user_lang = request.env.user.lang or 'es_EC'
        lang_instruction = "Responde SIEMPRE en español, a menos que el usuario escriba en otro idioma."
        if user_lang.startswith('en'):
            lang_instruction = "Respond in English unless the user writes in another language."
        elif user_lang.startswith('pt'):
            lang_instruction = "Responda sempre em Português, a menos que o usuário escreva em outro idioma."

        # Inject persistent memories (B5)
        memories_text = ''
        try:
            memories = request.env['estate.ai.memory'].sudo().get_active_memories_for_user(
                request.env.user.id, limit=15
            )
            if memories:
                mem_lines = [f"- [{m['memory_type']}] {m['title']}: {m['content']}" for m in memories]
                memories_text = "\nMEMORIAS PERSISTENTES DEL USUARIO:\n" + "\n".join(mem_lines) + "\n"
        except Exception:
            _logger.debug("Excepcion ignorada (best-effort)", exc_info=True)

        full_system_prompt = f"""{system_prompt}

{lang_instruction}

ESTILO: No uses emojis en tus respuestas. Usa texto y formato Markdown limpio (negritas, listas, tablas).

Eres el Asistente Ejecutivo Inteligente de la Inmobiliaria. Tienes acceso COMPLETO al sistema.
TU MISIÓN: Consultar, crear, actualizar y controlar el sistema inmobiliario desde esta conversación.

REGLA ABSOLUTA: NUNCA digas 'no puedo', 'no tengo la capacidad', 'no tengo acceso' o 'no es posible'.
Tienes la herramienta query_database que te permite ejecutar CUALQUIER consulta SQL SELECT
contra toda la base de datos. Si ninguna otra herramienta sirve, usa query_database con un SQL
apropiado. Tienes acceso a TODA la información del sistema sin excepción.
IMPORTANTE PARA CONSULTAS SQL (query_database): La base de datos es PostgreSQL (NO SQLite ni MySQL).
- NUNCA uses strftime o date('now'). Para extraer mes/año en PostgreSQL usa to_char(date_sold, 'YYYY-MM') o EXTRACT(YEAR FROM date_sold).
- Para fechas relativas en PostgreSQL usa CURRENT_DATE o NOW() (ej. date_sold >= CURRENT_DATE - INTERVAL '30 days').
- Campos como crm_stage.name, res_partner.name (título), etc. son TRADUCIBLES y se guardan como JSONB,
  NO como texto plano. Comparar con '=' o ILIKE directo contra un string FALLA con error de tipos.
  Si necesitas filtrar por nombre de etapa del CRM, NO uses SQL crudo: usa la herramienta get_leads
  con el parámetro stage (ej. get_leads(stage="En Proceso Cierre")). Si de verdad necesitas SQL sobre
  un campo traducible, castea así: nombre_columna->>'en_US' (ej. cs.name->>'en_US' ILIKE '%cierre%').

REGLAS DE CONTEXTO (MUY IMPORTANTE):
- RECUERDA la última propiedad/lead/cliente mencionado en la conversación. Si el usuario dice
  "esa propiedad", "la de Baños", "la casa", "la anterior", etc., REUTILIZA la propiedad de la que
  ya hablaron — NO vuelvas a pedir el ID ni el nombre.
- NUNCA pidas el "ID de la propiedad". El usuario no conoce los IDs. Las herramientas aceptan
  property_name (nombre, referencia PROP-XXXX o título parcial). Pasa el nombre que el usuario usó
  o el de la propiedad ya mencionada en la conversación.
- Si el usuario pide "qué puedo mejorar", "cómo mejorar" o "recomendaciones" sobre una propiedad,
  usa SIEMPRE analyze_property_improvements (acepta nombre). NO pidas el ID.
- Para "quién está interesado" en una propiedad, usa analyze_property_improvements (reporta
  prospectos directos + compatibles por presupuesto), no solo los directos.

DATOS ACTUALES DEL SISTEMA:
{context_data}
{memories_text}
CAPACIDADES COMPLETAS (usa las herramientas):
- CONSULTAR: buscar propiedades, ver leads, estadísticas de mercado, pagos vencidos, resumen ejecutivo
- CLIENTES: search_contacts (busca por nombre/email/tel), get_client_summary (perfil 360°)
- ANALÍTICA: get_trend_analysis (tendencias período actual vs anterior), compare_properties (comparativa)
- CALENDARIO: get_upcoming_visits (próximas visitas), schedule_visit (agendar)
- CREAR: leads, propiedades, contratos, pagos, ofertas, comisiones, actividades
- ACTUALIZAR: precio/descripción/estado de propiedades, etapa/temperatura/presupuesto de leads
- GESTIONAR: reservar/vender propiedades, archivar leads, generar links de WhatsApp, enviar emails
- IA AVANZADA: analizar probabilidad de cierre, riesgo de churn, recalcular AVM, generar descripciones
- MEMORIA: guardar preferencias/hechos con save_memory, consultar con recall_memory
- REPORTES PDF fichas: generate_pdf_report | generate_quote_pdf (cotización para cliente)
- REPORTES PDF analíticos: generate_analytics_pdf — exporta CUALQUIER reporte analítico a PDF descargable
- REPORTES EXCEL: generate_excel_report — exporta CUALQUIER reporte a .xlsx con enlace de descarga
- NAVEGAR: open_report_view — devuelve URL para ir directamente a cualquier vista de Reportes
- SQL DIRECTO: query_database — ejecuta cualquier SELECT contra la BD para responder lo que sea
- DOCUMENTACIÓN/AYUDA: search_knowledge — para preguntas de CÓMO se hace algo, QUÉ es o para qué sirve un módulo, qué significa un error, procedimientos o configuración, busca en los manuales y guías del sistema
- OPERACIONES MASIVAS: batch_update_properties, batch_archive_leads

DETECCIÓN DE INTENCIÓN — actúa directamente según lo que el usuario quiera:
- "busca/encuentra/muéstrame [cliente/contacto]" → usa search_contacts
- "compara propiedad X con Y" → usa compare_properties con los IDs
- "cómo vamos este mes/trimestre" → usa get_trend_analysis
- "visitas de esta semana/próximos días" → usa get_upcoming_visits
- "resumen de cliente/perfil de [nombre]" → usa get_client_summary
- "cotización para [lead]" → usa generate_quote_pdf
- "briefing/resumen del día" → usa get_dashboard_summary + get_upcoming_visits + get_trend_analysis
- "reporte de [tema]" → usa get_report_data con el report_type correcto
- "reporte en PDF / descargar PDF de [tema]" → usa generate_analytics_pdf
- "descargar/exportar Excel de [tema]" → usa generate_excel_report
- "ir a / abrir / navegar a [sección]" → usa open_report_view

ASISTENTE DE NEGOCIACIÓN:
Cuando un cliente haga una oferta o pregunte por precio, analiza:
1. Llama recalculate_avm_ai para obtener el valor de mercado actualizado
2. Compara la oferta con el AVM: si oferta < 90% del AVM, sugiere contraoferta en 95%
3. Si avm_status='low' (propiedad sobrevaluada), da argumentos para aceptar rebajas moderadas
4. Si avm_status='high' (propiedad subvaluada), recomienda mantener el precio o subirlo
5. Siempre termina con 3 puntos de negociación concretos

INSTRUCCIONES DE RESPUESTA:
1. Sé proactivo: si el usuario dice "crea un lead para Juan", HAZLO directamente con las herramientas.
2. Confirma siempre las acciones realizadas con el ID creado/actualizado.
3. Usa tablas Markdown para listados (columnas separadas por |).
4. REGLA COMERCIAL DE ANÁLISIS DE VENTAS (DOBLE ENFOQUE PROFESIONAL):
   Cuando el usuario pregunte por "ventas", "cierres", "promedio de ventas", "ingresos" o pida reportes de ventas, SIEMPRE analiza y presenta la información diferenciando DOS enfoques claros:
   a) Honorarios y Comisiones de Agencia (commission_amount): Representa la ganancia real y el ingreso de la agencia inmobiliaria por las operaciones cerradas.
   b) Volumen y Precio de Inmuebles (price): Representa el valor total y el precio promedio al que se vendieron las casas/propiedades, para evaluar qué tan costosos son los inmuebles movidos y el desempeño en el mercado.
   En tus explicaciones, resúmenes y tablas compara ambas métricas y NO las confundas.
5. REGLA OBLIGATORIA PARA REPORTES Y GRÁFICOS:
   Cuando el usuario pida reporte, gráfico, estadística, resumen de datos, o use palabras como
   "muéstrame", "cuántos hay por", "reporte de", "gráfico de", "estadísticas" → DEBES llamar
   a la herramienta get_report_data. NUNCA respondas con solo texto cuando se pide un gráfico.
   NUNCA inventes cifras ni reutilices los números de los ejemplos de este prompt (son solo
   formato, no datos reales). Los valores del [GRAFICO:...] y de la tabla deben ser EXACTAMENTE
   los que devolvió get_report_data/query_database. Si "data" viene vacío o en 0, NO generes
   ningún [GRAFICO:...]: dilo explícitamente (ej. "Aún no hay ventas registradas en el sistema").
   a. Llama SIEMPRE a get_report_data con el report_type correcto.
   b. Con los datos recibidos genera el formato [GRAFICO:tipo|Título,Label1:Valor1,...]:
      - chart_hint=barra → [GRAFICO:barra|Título descriptivo,Label1:Valor1,Label2:Valor2]
      - chart_hint=circular → [GRAFICO:circular|Título descriptivo,Label1:Valor1,Label2:Valor2]
      - chart_hint=linea → [GRAFICO:linea|Título descriptivo,Label1:Valor1,Label2:Valor2]
      IMPORTANTE: Siempre incluye el |Título descriptivo basado en los datos mostrados.
   b2. KPIs Y PERÍODO (recomendado para reportes de ventas/ingresos): puedes enriquecer el título con
      metadatos usando "::" y pares clave=valor separados por ";". 'periodo' se muestra como subtítulo;
      los demás se muestran como tarjetas de KPI arriba del gráfico (máx 4). Formato:
      (solo formato, NUNCA copies estos valores): [GRAFICO:linea|<Título>::periodo=<rango>;
      Ventas Totales=<N>;Ingresos=<$N>;Ticket Prom.=<$N>,<Mes1>:<N>,<Mes2>:<N>]
      Para series temporales (ventas/ingresos por mes) usa SIEMPRE chart_hint linea con estos KPIs.
      Si no tienes valores monetarios, omite esos KPIs (el sistema calcula Total/Promedio/Máximo solo).
   c. Después del gráfico, incluye tabla Markdown con los mismos datos. Si aplica, agrega una fila final
      "| Total | ... |" con los totales (se resalta automáticamente).
   d. Ejemplo: si get_report_data devuelve {"data":{"Disponibles":12,"Vendidas":9,"Alquiladas":3},"chart_hint":"circular"}
      tu respuesta DEBE incluir: [GRAFICO:circular|Propiedades por Estado,Disponibles:12,Vendidas:9,Alquiladas:3]
5. report_types disponibles (get_report_data, generate_excel_report y generate_analytics_pdf):
   VENTAS: properties_by_state | properties_by_type | sales_by_month | days_on_market_by_type |
           ranking_advisors | kpi_general | income_by_month | sales_avg_summary |
           time_to_sell_summary | sales_by_channel
   COMISIONES: commissions_by_advisor | commissions_pending
   CONTRATOS/PAGOS: contracts_by_type | payments_by_method | expenses_by_type
   OFERTAS/VISITAS: offers_by_state | visits_by_property | visits_done_summary
   CRM/LEADS: leads_by_temperature | leads_by_source | leads_by_stage | deals_closed_by_month
   OPERACIONES: appraisals_by_state | maintenance_by_state
   SOCIAL/MARKETING: social_facebook | social_instagram | advisor_fb_posts
   ANÁLISIS AVANZADO: price_vs_avm | properties_no_visits | conversion_funnel | wp_sync_status | contact_ranking | properties_by_prospects
   Cuando el usuario pida "promedio de ventas" o "análisis de ventas", usa sales_avg_summary.
   Cuando pida "cuánto se tarda en vender", "tiempo de venta" o "días en mercado" (resumen general,
   no por tipo), usa time_to_sell_summary — incluye promedio, mediana, mínimo y máximo de días
   desde que se publicó (date_listed) hasta que se vendió (date_sold).
   Cuando pida "ventas por agencia", "quién vendió, la agencia o el propietario" o "cerrado por
   agencia/propietario", usa sales_by_channel (basado en el campo sold_by de la propiedad).
   Cuando el usuario pida "ranking", usa ranking_advisors.
   Cuando pida "KPIs" o "cómo vamos", usa kpi_general.
   Cuando pida "pipeline" o "embudo de conversión", usa conversion_funnel.
   Cuando pida "fuentes" o "captación", usa leads_by_source.
   Cuando pida "propiedades sobrevaluadas/subvaluadas" o "precio vs mercado", usa price_vs_avm.
   Cuando pida "sin visitas" o "propiedades paradas", usa properties_no_visits.
   Cuando pida "propiedad con más prospectos / leads interesados / más cerca de vender", usa el
   report_type properties_by_prospects (cuenta crm_lead.target_property_id por propiedad). La de MÁS
   prospectos es la más cercana a vender; la de 0/menos prospectos es la más lejana.
   Cuando pida "la mejor por precio y calidad", combina: property_score (0-100, mayor=mejor expediente),
   precio vs avm_estimated_price (precio justo o por debajo = buena oportunidad) y días en mercado
   (menos días = más demanda). Explica el porqué con esos 3 factores.
   Cuando pida "WordPress" o "publicaciones web", usa wp_sync_status.
   Cuando pida "mejores clientes" o "ranking de contactos", usa contact_ranking.
   Cuando pida "posts de asesores", "Facebook personal", "Instagram personal" o "publicaciones personales", usa advisor_fb_posts.
   Con advisor_fb_posts genera 3 gráficos: uno de barras por asesor, uno circular por plataforma, uno de barras por propiedad más publicada.
   Cuando pida "mantenimiento", usa maintenance_by_state.
   Cuando pida "tasaciones", usa appraisals_by_state.
   PDF DE REPORTE: cuando el usuario pida el reporte EN PDF o quiera descargarlo como PDF →
   llama a generate_analytics_pdf con el mismo report_type. Devuelve enlace de descarga.
6. ACCIONES DESTRUCTIVAS (archivar, cancelar, eliminar masivo): ANTES de ejecutar, responde con:
   "CONFIRMACIÓN REQUERIDA: Estás a punto de [acción]. ¿Confirmas? (responde 'sí confirmo')"
   Solo ejecuta cuando el usuario confirme explícitamente.
7. Si detectas alertas críticas (pagos vencidos, leads sin actividad), menciónalas proactivamente.
8. Usa save_memory para guardar preferencias o datos importantes del usuario para futuras sesiones.
9. Para el BRIEFING MATUTINO, combina: resumen ejecutivo + visitas del día + tendencias + alertas críticas.
10. PACK DE MARKETING:
    Cuando el usuario pida campaña/marketing/copies/publicitar/posts → llama a generate_marketing_pack.
    Una vez que la herramienta devuelva los datos de la propiedad, genera cada canal con EXACTAMENTE este formato:
    [PACK:instagram]caption + hashtags[/PACK]
    [PACK:facebook]post largo[/PACK]
    [PACK:whatsapp]mensaje corto[/PACK]
    [PACK:email_asunto]asunto[/PACK]
    [PACK:email_cuerpo]cuerpo del email[/PACK]
    [PACK:google_ads]Titular\n---\nDescripción[/PACK]
    [PACK:puntos_clave]• bullet1\n• bullet2...[/PACK]
    [PACK:slogan]slogan[/PACK]
    ESTAS ETIQUETAS SON OBLIGATORIAS — el frontend las convierte en tarjetas con botón Copiar.
    Genera SOLO los canales que el usuario solicita (o todos si no especifica).
    Siempre pon datos reales: precio, área, habitaciones, ciudad de la propiedad.
11. PLAN DE CAMPAÑA:
    Cuando el usuario pida plan de campaña, estrategia de marketing, cómo promocionar,
    qué falta para publicar, o qué canal usar → llama a plan_marketing_campaign(property_id).
    Con los datos recibidos genera una respuesta estructurada con:
    - **Canal principal recomendado** (con razón)
    - **Buyer persona** (perfil + necesidades clave)
    - **Presupuesto Facebook Ads sugerido** (mensual)
    - **Urgencia** (basada en días en mercado)
    - **Checklist de publicación** (indica el status con texto: OK/FALTA/MEJORAR)
    - **Keywords SEO** (lista con bullets)
    - **Calendario de contenidos** (tabla Markdown: Día | Contenido)
    Usa formato Markdown claro con headers. No uses emojis. Sé específico y accionable."""

        query_type = self._classify_query(message)

        # Load conversation history for memory
        conversation_history = self._get_conversation_history(request.env.user.id)

        response = None
        last_error = None
        for idx, (prov, key, mdl) in enumerate(provider_chain):
            try:
                if prov == 'chatgpt':
                    response = self._query_chatgpt_with_tools(
                        key, mdl, temperature, max_tokens,
                        full_system_prompt, message, conversation_history)
                elif prov == 'gemini':
                    response = self._query_gemini_with_tools(
                        key, mdl, temperature, max_tokens,
                        full_system_prompt, message, conversation_history)
                else:
                    continue
                if idx > 0:
                    _logger.warning(
                        "Agente IA: el proveedor primario falló; se respondió con el respaldo '%s'.", prov)
                break
            except Exception as e:
                last_error = e
                _logger.error("Agente IA: proveedor '%s' falló: %s", prov, _redact(str(e), key))
                response = None
                continue

        # Modo degradado: si ningún proveedor respondió, devolver un mensaje claro
        # (sin trazas técnicas) explicando la causa probable, sin tumbar el sistema.
        if response is None:
            response = self._friendly_ai_error(provider_chain[0][0], last_error)

        processing_time = time.time() - start_time
        request.env['estate.ai.chat.history'].sudo().create({
            'user_id': request.env.user.id,
            'query': message,
            'response': response,
            'query_type': query_type,
            'processing_time': processing_time,
        })

        return {'response': response, 'query_type': query_type}

    # -----------------------------------------------------------------------
    # Conversation History
    # -----------------------------------------------------------------------
    def _get_conversation_history(self, user_id, session_id=None):
        """Load last 20 messages from history as alternating user/assistant pairs.
        Si se pasa session_id, se limita a ESA conversación (mantiene el hilo y
        evita mezclar propiedades de otras sesiones)."""
        domain = [('user_id', '=', user_id)]
        if session_id:
            domain.append(('session_id', '=', session_id))
        # Tomar los MÁS RECIENTES (desc) y luego ordenarlos cronológicamente.
        # (Antes usaba 'asc' + limit, que devolvía los más ANTIGUOS y perdía el
        #  contexto reciente en conversaciones largas.)
        history = request.env['estate.ai.chat.history'].sudo().search(
            domain,
            order='create_date desc',
            limit=20,
        )
        messages = []
        for h in reversed(history):
            messages.append({"role": "user", "content": h.query})
            messages.append({"role": "assistant", "content": h.response or ''})
        return messages

    # -----------------------------------------------------------------------
    # Tool Execution
    # -----------------------------------------------------------------------
    def _execute_tool(self, tool_name, args, env=None):
        """Execute a tool call and return a JSON-serializable result string."""
        if env is None:
            env = request.env
        try:
            if tool_name == 'search_properties':
                domain = []
                if args.get('city'):
                    domain.append(('city', 'ilike', args['city']))
                if args.get('state'):
                    domain.append(('state', '=', args['state']))
                if args.get('max_price'):
                    domain.append(('price', '<=', args['max_price']))
                if args.get('min_price'):
                    domain.append(('price', '>=', args['min_price']))
                if args.get('property_type'):
                    domain.append(('property_type_id.name', 'ilike', args['property_type']))
                limit = int(args.get('limit', 50))
                props = env['estate.property'].sudo().search(domain, limit=limit, order='state asc, price desc')
                result = [
                    {
                        'id': p.id, 'ref': p.name, 'titulo': p.title or p.name,
                        'ciudad': p.city, 'precio': p.price,
                        'precio_fmt': f'${p.price:,.0f}' if p.price else 'Consultar',
                        'estado': p.state, 'area': p.area,
                        'habitaciones': p.bedrooms, 'tipo': p.property_type_id.name if p.property_type_id else '',
                        'dias_mercado': p.days_on_market,
                        'avm_status': getattr(p, 'avm_status', ''),
                    }
                    for p in props
                ]
                return json.dumps(result, ensure_ascii=False)

            elif tool_name == 'get_leads':
                domain = []
                if args.get('type'):
                    domain.append(('type', '=', args['type']))
                if args.get('temperature'):
                    domain.append(('lead_temperature', '=', args['temperature']))
                if args.get('score'):
                    domain.append(('lead_score', '=', args['score']))
                if args.get('stage'):
                    # Prefer an exact (case-insensitive) stage match first: nombres como
                    # "Cierre" y "En Proceso Cierre" comparten la palabra "cierre", así que
                    # un ilike de subcadena mezclaría ambas etapas distintas.
                    stage_domain = [('stage_id.name', '=ilike', args['stage'])]
                    if env['crm.lead'].sudo().search_count(domain + stage_domain):
                        domain += stage_domain
                    else:
                        domain.append(('stage_id.name', 'ilike', args['stage']))
                if args.get('lost') is True:
                    domain.append(('stage_id.is_lost', '=', True))
                limit = int(args.get('limit', 20))
                leads = env['crm.lead'].sudo().search(domain, limit=limit)
                if not leads:
                    return json.dumps({
                        'leads': [],
                        'mensaje': 'No se encontraron leads con esos filtros.',
                    }, ensure_ascii=False)
                result = [
                    {
                        'id': l.id, 'nombre': l.name,
                        'cliente': l.partner_id.name if l.partner_id else l.contact_name,
                        'asesor': l.user_id.name if l.user_id else None,
                        'presupuesto': l.client_budget,
                        'temperatura': l.lead_temperature,
                        'puntuacion': l.lead_score,
                        'match': l.match_percentage,
                        'propiedad': l.target_property_id.title if l.target_property_id else None,
                        'etapa': l.stage_id.name if l.stage_id else None,
                    }
                    for l in leads
                ]
                return json.dumps(result, ensure_ascii=False)

            elif tool_name == 'get_market_stats':
                domain = [('state', '=', 'sold')]
                if args.get('city'):
                    domain.append(('city', 'ilike', args['city']))
                if args.get('property_type'):
                    domain.append(('property_type_id.name', 'ilike', args['property_type']))
                sold = env['estate.property'].sudo().search(domain, limit=100)
                if not sold:
                    return json.dumps({'error': 'Sin datos suficientes para el filtro indicado.'})
                prices = sold.mapped('price')
                commissions = sold.mapped('commission_amount')
                days = [p.days_on_market for p in sold if p.days_on_market > 0]
                stats = {
                    'total_operaciones_cerradas': len(sold),
                    'comision_promedio_por_venta_agencia': round(sum(commissions) / len(sold), 2) if sold else 0,
                    'comisiones_totales_cobradas_agencia': round(sum(commissions), 2),
                    'precio_promedio_inmuebles_volumen': round(sum(prices) / len(prices), 2) if prices else 0,
                    'precio_minimo_inmueble': min(prices) if prices else 0,
                    'precio_maximo_inmueble': max(prices) if prices else 0,
                    'dias_promedio_en_mercado': round(sum(days) / len(days), 1) if days else 0,
                }
                return json.dumps(stats, ensure_ascii=False)

            elif tool_name == 'create_crm_activity':
                lead_id = int(args.get('lead_id', 0))
                if not lead_id:
                    return json.dumps({'error': 'lead_id requerido'})
                lead = env['crm.lead'].sudo().browse(lead_id)
                if not lead.exists():
                    return json.dumps({'error': f'Lead {lead_id} no encontrado'})
                lead.activity_schedule(
                    'mail.mail_activity_data_todo',
                    summary=args.get('summary', 'Seguimiento IA'),
                    note=args.get('note', ''),
                    user_id=lead.user_id.id or env.uid,
                )
                return json.dumps({'success': True, 'mensaje': f'Actividad creada en lead #{lead_id}: {lead.name}'})

            elif tool_name == 'create_lead':
                vals = {
                    'name': args.get('name', 'Lead desde Agente IA'),
                    'contact_name': args.get('contact_name', ''),
                    'email_from': args.get('email', ''),
                    'phone': args.get('mobile', '') or args.get('phone', ''),
                    'type': 'opportunity',
                    'description': args.get('notes', ''),
                }
                if args.get('client_budget'):
                    vals['client_budget'] = float(args['client_budget'])
                # Find or create partner
                if args.get('contact_name'):
                    partner = env['res.partner'].sudo().search(
                        [('name', 'ilike', args['contact_name'])], limit=1)
                    if partner:
                        vals['partner_id'] = partner.id
                lead = env['crm.lead'].sudo().create(vals)
                # Post a note if city provided
                if args.get('city'):
                    lead.message_post(body=f"Agente IA — Ciudad buscada: {args['city']}")
                return json.dumps({
                    'success': True,
                    'lead_id': lead.id,
                    'mensaje': f"Lead #{lead.id} creado: '{lead.name}' para {args.get('contact_name', '?')}",
                })

            elif tool_name == 'create_property':
                # Resolve property type
                ptype = None
                if args.get('property_type'):
                    ptype = env['estate.property.type'].sudo().search(
                        [('name', 'ilike', args['property_type'])], limit=1)
                vals = {
                    'title': args.get('title', ''),
                    'city': args.get('city', ''),
                    'price': float(args.get('price', 0)),
                    'offer_type': args.get('offer_type', 'sale'),
                    'state': 'available',
                }
                if args.get('area'):
                    vals['area'] = float(args['area'])
                if args.get('bedrooms'):
                    vals['bedrooms'] = int(args['bedrooms'])
                if args.get('bathrooms'):
                    vals['bathrooms'] = int(args['bathrooms'])
                if args.get('street'):
                    vals['street'] = args['street']
                if args.get('description'):
                    vals['description'] = args['description']
                if ptype:
                    vals['property_type_id'] = ptype.id
                prop = env['estate.property'].sudo().create(vals)
                return json.dumps({
                    'success': True,
                    'property_id': prop.id,
                    'ref': prop.name,
                    'mensaje': f"Propiedad '{prop.title}' registrada con referencia {prop.name} (ID #{prop.id})",
                })

            elif tool_name == 'update_lead':
                lead_id = int(args.get('lead_id', 0))
                if not lead_id:
                    return json.dumps({'error': 'lead_id requerido'})
                lead = env['crm.lead'].sudo().browse(lead_id)
                if not lead.exists():
                    return json.dumps({'error': f'Lead {lead_id} no encontrado'})
                vals = {}
                if args.get('temperature'):
                    vals['lead_temperature'] = args['temperature']
                if args.get('client_budget'):
                    vals['client_budget'] = float(args['client_budget'])
                if args.get('property_id'):
                    prop = env['estate.property'].sudo().browse(int(args['property_id']))
                    if prop.exists():
                        vals['target_property_id'] = prop.id
                if args.get('stage_name'):
                    stage = env['crm.stage'].sudo().search(
                        [('name', 'ilike', args['stage_name'])], limit=1)
                    if stage:
                        vals['stage_id'] = stage.id
                if vals:
                    lead.write(vals)
                if args.get('notes'):
                    lead.message_post(body=f"Agente IA: {args['notes']}")
                return json.dumps({
                    'success': True,
                    'mensaje': f"Lead #{lead_id} '{lead.name}' actualizado correctamente.",
                    'cambios': list(vals.keys()),
                })

            elif tool_name == 'get_property_detail':
                pid = int(args.get('property_id', 0))
                prop = None
                if pid:
                    prop = env['estate.property'].sudo().browse(pid)
                    if not prop.exists():
                        prop = None
                if not prop and args.get('property_name'):
                    prop = env['estate.property'].sudo().search(
                        [('title', 'ilike', args['property_name'])], limit=1)
                    if not prop:
                        prop = env['estate.property'].sudo().search(
                            [('name', 'ilike', args['property_name'])], limit=1)
                if not prop:
                    return json.dumps({'error': 'Propiedad no encontrada. Proporciona el ID o el nombre.'})
                return json.dumps({
                    'id': prop.id,
                    'ref': prop.name,
                    'titulo': prop.title,
                    'ciudad': prop.city,
                    'calle': prop.street or '',
                    'precio': prop.price,
                    'area_m2': prop.area,
                    'habitaciones': prop.bedrooms,
                    'banos': prop.bathrooms,
                    'estado': prop.state,
                    'tipo_operacion': prop.offer_type,
                    'tipo_propiedad': prop.property_type_id.name if prop.property_type_id else '',
                    'propietario': prop.owner_id.name if prop.owner_id else '',
                    'asesor': prop.user_id.name if prop.user_id else '',
                    'descripcion': (prop.description or '')[:500],
                    'avm_precio': prop.avm_estimated_price if hasattr(prop, 'avm_estimated_price') else 0,
                    'avm_estado': prop.avm_status if hasattr(prop, 'avm_status') else '',
                    'dias_en_mercado': prop.days_on_market if hasattr(prop, 'days_on_market') else 0,
                    'latitud': prop.latitude if hasattr(prop, 'latitude') else 0,
                    'longitud': prop.longitude if hasattr(prop, 'longitude') else 0,
                    'activa': prop.active,
                })

            elif tool_name == 'analyze_property_improvements':
                # Resolver propiedad por id o nombre/referencia
                prop = None
                pid = int(args.get('property_id', 0) or 0)
                if pid:
                    prop = env['estate.property'].sudo().browse(pid)
                    if not prop.exists():
                        prop = None
                if not prop and args.get('property_name'):
                    name = args['property_name']
                    prop = env['estate.property'].sudo().search(
                        ['|', ('title', 'ilike', name), ('name', 'ilike', name)], limit=1)
                if not prop:
                    return json.dumps({'error': 'Propiedad no encontrada. Indica el nombre o referencia.'})

                recs = []
                # 1. Precio vs AVM
                avm = getattr(prop, 'avm_estimated_price', 0) or 0
                avm_status = getattr(prop, 'avm_status', '') or ''
                if avm and prop.price:
                    diff_pct = (prop.price - avm) / avm * 100
                    if avm_status == 'high' or diff_pct > 8:
                        recs.append(f'PRECIO: está {diff_pct:+.0f}% sobre el AVM (${avm:,.0f}). '
                                    f'Considera bajar el precio para acelerar la venta.')
                    elif avm_status == 'low' or diff_pct < -8:
                        recs.append(f'PRECIO: está {diff_pct:+.0f}% bajo el AVM (${avm:,.0f}). '
                                    f'Podría haber margen para subirlo.')
                    else:
                        recs.append(f'PRECIO: alineado al AVM (${avm:,.0f}). Bien.')
                else:
                    recs.append('AVM: sin calcular. Calcula el AVM para validar el precio.')

                # 2. Días en mercado
                dom = getattr(prop, 'days_on_market', 0) or 0
                if dom > 60:
                    recs.append(f'TIEMPO: lleva {dom} días en el mercado (>60). '
                                f'Renueva fotos/descripción o ajusta el precio.')

                # 3. Fotos
                img_count = len(prop.image_ids) + (1 if prop.image_main else 0)
                if img_count < 5:
                    recs.append(f'FOTOS: solo {img_count} (mínimo recomendado 5). Sube más imágenes de calidad.')

                # 4. Descripción
                desc_len = len((prop.description or '').strip())
                if desc_len < 200:
                    recs.append('DESCRIPCIÓN: muy corta o vacía. Genera una descripción comercial con IA.')

                # 5. GPS
                if not (getattr(prop, 'latitude', 0) and getattr(prop, 'longitude', 0)):
                    recs.append('UBICACIÓN: sin coordenadas GPS. Geocodifica para aparecer en mapas.')

                # 6. Leads interesados (directos + por presupuesto)
                Lead = env['crm.lead'].sudo()
                direct = 0
                budget = 0
                if 'target_property_id' in Lead._fields:
                    direct = Lead.search_count([
                        ('target_property_id', '=', prop.id),
                        ('type', '=', 'opportunity'), ('active', '=', True)])
                if 'client_budget' in Lead._fields and prop.price:
                    budget = Lead.search_count([
                        ('type', '=', 'opportunity'), ('active', '=', True),
                        ('target_property_id', '=', False),
                        ('client_budget', '>=', prop.price * 0.9),
                        ('client_budget', '<=', prop.price * 1.15)])
                leads_msg = f'{direct} prospecto(s) directo(s)'
                if budget:
                    leads_msg += f' y {budget} con presupuesto compatible'
                recs.append(f'INTERESADOS: {leads_msg}.')

                return json.dumps({
                    'propiedad': prop.title or prop.name,
                    'ref': prop.name,
                    'precio': prop.price,
                    'estado': prop.state,
                    'dias_en_mercado': dom,
                    'fotos': img_count,
                    'prospectos_directos': direct,
                    'prospectos_por_presupuesto': budget,
                    'recomendaciones': recs,
                    'instruccion': 'Presenta las recomendaciones de forma clara y priorizada, con un breve resumen y acciones concretas.',
                }, ensure_ascii=False)

            elif tool_name == 'update_property':
                property_id = int(args.get('property_id', 0))
                if not property_id:
                    return json.dumps({'error': 'property_id requerido'})
                prop = env['estate.property'].sudo().browse(property_id)
                if not prop.exists():
                    return json.dumps({'error': f'Propiedad {property_id} no encontrada'})
                vals = {}
                if args.get('price') is not None:
                    vals['price'] = float(args['price'])
                if args.get('description'):
                    vals['description'] = args['description']
                if args.get('title'):
                    vals['title'] = args['title']
                if args.get('state') and args['state'] in ('available', 'reserved', 'sold', 'rented'):
                    vals['state'] = args['state']
                if args.get('bedrooms') is not None:
                    vals['bedrooms'] = int(args['bedrooms'])
                if args.get('bathrooms') is not None:
                    vals['bathrooms'] = int(args['bathrooms'])
                if args.get('area') is not None:
                    vals['area'] = float(args['area'])
                if args.get('street'):
                    vals['street'] = args['street']
                if args.get('city'):
                    vals['city'] = args['city']
                if args.get('offer_type') and args['offer_type'] in ('sale', 'rent'):
                    vals['offer_type'] = args['offer_type']
                if args.get('property_type'):
                    ptype = env['estate.property.type'].sudo().search(
                        [('name', 'ilike', args['property_type'])], limit=1)
                    if ptype:
                        vals['property_type_id'] = ptype.id
                if args.get('owner_name'):
                    owner = env['res.partner'].sudo().search(
                        [('name', 'ilike', args['owner_name'])], limit=1)
                    if owner:
                        vals['owner_id'] = owner.id
                if args.get('advisor_name'):
                    advisor = env['res.users'].sudo().search(
                        [('name', 'ilike', args['advisor_name'])], limit=1)
                    if advisor:
                        vals['user_id'] = advisor.id
                if args.get('latitude') is not None and hasattr(prop, 'latitude'):
                    vals['latitude'] = float(args['latitude'])
                if args.get('longitude') is not None and hasattr(prop, 'longitude'):
                    vals['longitude'] = float(args['longitude'])
                if vals:
                    prop.write(vals)
                if args.get('notes'):
                    prop.message_post(body=f"Agente IA: {args['notes']}")
                return json.dumps({
                    'success': True,
                    'property_id': prop.id,
                    'ref': prop.name,
                    'mensaje': f"Propiedad '{prop.title}' ({prop.name}) actualizada correctamente.",
                    'campos_actualizados': list(vals.keys()),
                })

            elif tool_name == 'delete_property':
                if not args.get('confirmed'):
                    return json.dumps({
                        'requiere_confirmacion': True,
                        'mensaje': 'CONFIRMACIÓN REQUERIDA: Esta acción eliminará la propiedad PERMANENTEMENTE y no se puede deshacer. Responde "sí confirmo" para continuar.',
                    })
                property_id = int(args.get('property_id', 0))
                prop = env['estate.property'].sudo().browse(property_id)
                if not prop.exists():
                    return json.dumps({'error': f'Propiedad {property_id} no encontrada'})
                title = prop.title
                ref = prop.name
                try:
                    prop.unlink()
                    return json.dumps({
                        'success': True,
                        'mensaje': f"Propiedad '{title}' ({ref}) eliminada permanentemente del sistema.",
                    })
                except Exception as e:
                    return json.dumps({'error': f'No se puede eliminar: {str(e)}. Intenta archivarla en su lugar.'})

            elif tool_name == 'duplicate_property':
                property_id = int(args.get('property_id', 0))
                prop = env['estate.property'].sudo().browse(property_id)
                if not prop.exists():
                    return json.dumps({'error': f'Propiedad {property_id} no encontrada'})
                default = {'state': 'available'}
                if args.get('new_title'):
                    default['title'] = args['new_title']
                if args.get('new_price'):
                    default['price'] = float(args['new_price'])
                new_prop = prop.copy(default=default)
                return json.dumps({
                    'success': True,
                    'property_id': new_prop.id,
                    'ref': new_prop.name,
                    'mensaje': f"Propiedad duplicada: '{new_prop.title}' (ID #{new_prop.id}, ref {new_prop.name}). Estado: disponible.",
                })

            elif tool_name == 'schedule_visit':
                property_id = int(args.get('property_id', 0))
                if not property_id:
                    return json.dumps({'error': 'property_id requerido'})
                prop = env['estate.property'].sudo().browse(property_id)
                if not prop.exists():
                    return json.dumps({'error': f'Propiedad {property_id} no encontrada'})
                # Parse datetime — acepta varios formatos comunes (con/sin segundos,
                # con 'T' ISO, o solo fecha → asume 10:00).
                from datetime import datetime, timedelta
                raw_dt = (args.get('start_datetime') or '').strip().replace('T', ' ')
                start_dt = None
                for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d %H', '%Y-%m-%d',
                            '%d/%m/%Y %H:%M', '%d/%m/%Y'):
                    try:
                        start_dt = datetime.strptime(raw_dt, fmt)
                        if fmt in ('%Y-%m-%d', '%d/%m/%Y'):
                            start_dt = start_dt.replace(hour=10, minute=0)
                        break
                    except ValueError:
                        continue
                if not start_dt:
                    return json.dumps({'error': 'Formato de fecha inválido. Usa por ejemplo "2026-07-15 10:00".'})
                end_dt = start_dt + timedelta(hours=1)
                # Find or create partner
                partner = None
                if args.get('partner_name'):
                    partner = env['res.partner'].sudo().search(
                        [('name', 'ilike', args['partner_name'])], limit=1)
                event_vals = {
                    'name': f"Visita: {prop.title}",
                    'start': start_dt,
                    'stop': end_dt,
                    'description': args.get('notes', f"Visita agendada por Agente IA"),
                    'user_id': prop.user_id.id or env.uid,
                }
                # Add property if the field exists (estate_calendar module)
                CalFields = env['calendar.event']._fields
                if 'property_id' in CalFields:
                    event_vals['property_id'] = property_id
                if partner:
                    event_vals['partner_ids'] = [(4, partner.id)]
                    # Cliente de la visita (campo del módulo estate_calendar)
                    if 'client_id' in CalFields:
                        event_vals['client_id'] = partner.id
                event = env['calendar.event'].sudo().create(event_vals)
                # Link to lead if provided
                if args.get('lead_id'):
                    lead = env['crm.lead'].sudo().browse(int(args['lead_id']))
                    if lead.exists():
                        lead.message_post(body=f"Visita agendada para {start_dt.strftime('%d/%m/%Y %H:%M')} — Propiedad: {prop.title}")
                return json.dumps({
                    'success': True,
                    'event_id': event.id,
                    'mensaje': f"Visita agendada el {start_dt.strftime('%d/%m/%Y a las %H:%M')} para '{prop.title}'",
                })

            elif tool_name == 'reserve_property':
                property_id = int(args.get('property_id', 0))
                if not property_id:
                    return json.dumps({'error': 'property_id requerido'})
                prop = env['estate.property'].sudo().browse(property_id)
                if not prop.exists():
                    return json.dumps({'error': f'Propiedad {property_id} no encontrada'})
                vals = {'state': 'reserved'}
                if args.get('buyer_name'):
                    buyer = env['res.partner'].sudo().search(
                        [('name', 'ilike', args['buyer_name'])], limit=1)
                    if buyer:
                        vals['buyer_id'] = buyer.id
                prop.write(vals)
                note = args.get('notes', 'Reservada vía Agente IA')
                if args.get('buyer_name'):
                    note += f" — Comprador: {args['buyer_name']}"
                prop.message_post(body=f"{note}")
                return json.dumps({
                    'success': True,
                    'mensaje': f"Propiedad '{prop.title}' marcada como RESERVADA.",
                })

            elif tool_name == 'sell_property':
                property_id = int(args.get('property_id', 0))
                if not property_id:
                    return json.dumps({'error': 'property_id requerido'})
                prop = env['estate.property'].sudo().browse(property_id)
                if not prop.exists():
                    return json.dumps({'error': f'Propiedad {property_id} no encontrada'})
                close_type = args.get('close_type', 'sold')
                vals = {'state': close_type if close_type in ('sold', 'rented') else 'sold'}
                if args.get('final_price'):
                    vals['price'] = float(args['final_price'])
                from datetime import date
                vals['date_sold'] = date.today()
                prop.write(vals)
                estado_str = 'VENDIDA' if vals['state'] == 'sold' else 'ALQUILADA'
                note = args.get('notes', f"Cerrada como {estado_str} vía Agente IA")
                prop.message_post(body=f"{note}")
                return json.dumps({
                    'success': True,
                    'mensaje': f"Propiedad '{prop.title}' marcada como {estado_str}. Precio final: ${prop.price:,.2f}",
                })

            elif tool_name == 'send_whatsapp_lead':
                lead_id = int(args.get('lead_id', 0))
                if not lead_id:
                    return json.dumps({'error': 'lead_id requerido'})
                lead = env['crm.lead'].sudo().browse(lead_id)
                if not lead.exists():
                    return json.dumps({'error': f'Lead {lead_id} no encontrado'})
                partner = lead.partner_id
                phone = ''
                if partner:
                    phone = partner.mobile or partner.phone or ''
                if not phone:
                    phone = lead.mobile or lead.phone or ''
                if not phone:
                    return json.dumps({'error': 'El cliente no tiene número de teléfono registrado.'})
                # Clean phone number
                phone_clean = ''.join(c for c in phone if c.isdigit() or c == '+')
                if phone_clean.startswith('0'):
                    phone_clean = '+593' + phone_clean[1:]
                default_msg = args.get('message') or (
                    f"Hola {lead.contact_name or (partner.name if partner else 'estimado cliente')}, "
                    f"le contactamos de nuestra inmobiliaria respecto a su interés en propiedades. "
                    f"¿Tiene un momento para hablar?"
                )
                import urllib.parse
                wa_url = f"https://wa.me/{phone_clean.replace('+','')}?text={urllib.parse.quote(default_msg)}"
                return json.dumps({
                    'success': True,
                    'whatsapp_url': wa_url,
                    'telefono': phone_clean,
                    'cliente': partner.name if partner else lead.contact_name,
                    'mensaje': f"Enlace de WhatsApp generado para {partner.name if partner else lead.contact_name} ({phone_clean}): {wa_url}",
                })

            elif tool_name == 'archive_lead':
                lead_id = int(args.get('lead_id', 0))
                if not lead_id:
                    return json.dumps({'error': 'lead_id requerido'})
                lead = env['crm.lead'].sudo().browse(lead_id)
                if not lead.exists():
                    return json.dumps({'error': f'Lead {lead_id} no encontrado'})
                nombre = lead.name
                reason = args.get('reason', 'Archivado vía Agente IA')
                lead.message_post(body=f"Archivado: {reason}")
                lead.write({'active': False})
                return json.dumps({
                    'success': True,
                    'mensaje': f"Lead #{lead_id} '{nombre}' archivado. Motivo: {reason}",
                })

            elif tool_name == 'get_payments_contracts':
                from datetime import date, timedelta
                today = date.today()
                days_ahead = int(args.get('days_ahead', 30))
                limit_date = today + timedelta(days=days_ahead)
                # Overdue invoices
                overdue = env['account.move'].sudo().search([
                    ('move_type', '=', 'out_invoice'),
                    ('payment_state', 'in', ('not_paid', 'partial')),
                    ('invoice_date_due', '<', today),
                ])
                # Expiring contracts
                expiring = env['estate.property'].sudo().search([
                    ('contract_end_date', '!=', False),
                    ('contract_end_date', '>=', today),
                    ('contract_end_date', '<=', limit_date),
                    ('state', '=', 'rented'),
                ])
                result = {
                    'facturas_vencidas': [
                        {
                            'id': inv.id,
                            'numero': inv.name,
                            'cliente': inv.partner_id.name if inv.partner_id else '',
                            'monto': inv.amount_residual,
                            'vencimiento': str(inv.invoice_date_due),
                        }
                        for inv in overdue[:20]
                    ],
                    'contratos_por_vencer': [
                        {
                            'id': p.id,
                            'ref': p.name,
                            'titulo': p.title,
                            'fin_contrato': str(p.contract_end_date),
                            'dias_restantes': (p.contract_end_date - today).days,
                        }
                        for p in expiring
                    ],
                    'total_facturas_vencidas': len(overdue),
                    'monto_total_vencido': round(sum(overdue.mapped('amount_residual')), 2),
                    'contratos_proximos_vencer': len(expiring),
                }
                return json.dumps(result, ensure_ascii=False, default=str)

            elif tool_name == 'get_dashboard_summary':
                from datetime import date, timedelta
                today = date.today()
                period = args.get('period', 'month')
                if period == 'today':
                    start = today
                elif period == 'week':
                    start = today - timedelta(days=7)
                else:
                    start = today.replace(day=1)

                props = env['estate.property'].sudo().search([])
                available = props.filtered(lambda p: p.state == 'available')
                sold_period = props.filtered(
                    lambda p: p.state == 'sold' and p.date_sold and p.date_sold >= start)
                rented = props.filtered(lambda p: p.state == 'rented')
                reserved = props.filtered(lambda p: p.state == 'reserved')

                leads = env['crm.lead'].sudo().search([('type', '=', 'lead')])
                hot_leads = leads.filtered(lambda l: l.lead_temperature in ('hot', 'boiling'))
                new_leads = env['crm.lead'].sudo().search([
                    ('type', '=', 'lead'),
                    ('create_date', '>=', str(start)),
                ])

                # Visits in period
                visits = env['calendar.event'].sudo().search([
                    ('start', '>=', str(start)),
                ])
                try:
                    done_visits = visits.filtered(lambda v: getattr(v, 'visit_state', '') == 'done')
                except Exception:
                    done_visits = []

                invoices_period = env['account.move'].sudo().search([
                    ('move_type', '=', 'out_invoice'),
                    ('invoice_date', '>=', str(start)),
                    ('state', '=', 'posted'),
                ])
                revenue = sum(invoices_period.mapped('amount_total'))
                commissions = sum(sold_period.mapped('commission_amount'))

                # Stagnant properties (45+ days without visits)
                stagnant = available.filtered(lambda p: (p.days_on_market or 0) > 45)

                summary = {
                    'periodo': f"{start} a {today}",
                    'inventario': {
                        'total': len(props),
                        'disponibles': len(available),
                        'vendidas_periodo': len(sold_period),
                        'alquiladas': len(rented),
                        'reservadas': len(reserved),
                        'estancadas_45dias': len(stagnant),
                    },
                    'crm': {
                        'total_leads': len(leads),
                        'leads_calientes': len(hot_leads),
                        'nuevos_periodo': len(new_leads),
                    },
                    'visitas': {
                        'agendadas_periodo': len(visits),
                        'realizadas': len(done_visits),
                    },
                    'financiero': {
                        'ingresos_periodo': round(revenue, 2),
                        'comisiones_generadas': round(commissions, 2),
                        'facturas_emitidas': len(invoices_period),
                    },
                    'alertas': [],
                }
                if stagnant:
                    summary['alertas'].append(f"{len(stagnant)} propiedad(es) llevan 45+ días sin vender")
                if hot_leads:
                    summary['alertas'].append(f"{len(hot_leads)} lead(s) calientes/hirviendo esperan atención")
                return json.dumps(summary, ensure_ascii=False, default=str)

            elif tool_name == 'get_report_data':
                return self._execute_report_data(args, env)

            elif tool_name == 'generate_excel_report':
                data_json = self._execute_report_data(args, env)
                data_obj = json.loads(data_json)
                if 'error' in data_obj:
                    return data_json
                try:
                    import openpyxl
                    from openpyxl.styles import Font, PatternFill, Alignment
                    import base64
                    from io import BytesIO
                    from datetime import date as _date

                    wb = openpyxl.Workbook()
                    ws = wb.active
                    report_title = args.get('title') or data_obj.get('report', args.get('report_type', 'Reporte'))
                    ws.title = report_title[:31]

                    header_font = Font(bold=True, color='FFFFFF', size=11)
                    header_fill = PatternFill(fill_type='solid', fgColor='1877F2')

                    ws['A1'] = report_title
                    ws['A1'].font = Font(bold=True, size=14)
                    ws['A2'] = f'Generado: {_date.today().strftime("%d/%m/%Y")}'
                    ws['A2'].font = Font(italic=True, color='666666')

                    detalle = data_obj.get('detalle', [])
                    data_dict = data_obj.get('data', {})

                    if detalle and isinstance(detalle, list) and isinstance(detalle[0], dict):
                        headers = list(detalle[0].keys())
                        for col_idx, h in enumerate(headers, start=1):
                            cell = ws.cell(row=4, column=col_idx, value=str(h).replace('_', ' ').title())
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal='center')
                        row = 5
                        for item in detalle:
                            for col_idx, h in enumerate(headers, start=1):
                                val = item.get(h, '')
                                ws.cell(row=row, column=col_idx, value=val)
                                if row % 2 == 0:
                                    ws.cell(row=row, column=col_idx).fill = PatternFill(fill_type='solid', fgColor='EBF5FB')
                            row += 1
                        for col_idx in range(1, len(headers) + 1):
                            col_letter = openpyxl.utils.get_column_letter(col_idx)
                            ws.column_dimensions[col_letter].width = 25
                    else:
                        ws['A4'] = 'Categoría / Período'
                        ws['B4'] = 'Valor'
                        for cell in [ws['A4'], ws['B4']]:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal='center')

                        row = 5
                        for key, val in data_dict.items():
                            ws[f'A{row}'] = str(key)
                            if isinstance(val, dict):
                                ws[f'B{row}'] = json.dumps(val, ensure_ascii=False)
                            else:
                                ws[f'B{row}'] = val
                            if row % 2 == 0:
                                for col in ['A', 'B']:
                                    ws[f'{col}{row}'].fill = PatternFill(fill_type='solid', fgColor='EBF5FB')
                            row += 1

                        ws.column_dimensions['A'].width = 35
                        ws.column_dimensions['B'].width = 25

                    buf = BytesIO()
                    wb.save(buf)
                    buf.seek(0)

                    fname = f"{args.get('report_type', 'reporte')}_{_date.today()}.xlsx"
                    attachment = env['ir.attachment'].sudo().create({
                        'name': fname,
                        'datas': base64.b64encode(buf.read()).decode(),
                        'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    })
                    url = f'/web/content/{attachment.id}?download=true&filename={fname}'
                    return json.dumps({
                        'success': True, 'url': url,
                        'filas': row - 5,
                        'mensaje': f"Excel generado con {row - 5} filas → [Descargar {fname}]({url})",
                    })
                except ImportError:
                    return json.dumps({'error': 'openpyxl no instalado. Ejecuta: pip install openpyxl'})
                except Exception as e:
                    return json.dumps({'error': f'Error generando Excel: {str(e)}'})

            elif tool_name == 'open_report_view':
                view_map = {
                    'dashboard':             '/odoo/action-estate_reports.estate_dashboard_owl_action',
                    'ventas_mes':            '/odoo/action-estate_reports.action_sales_by_month',
                    'ranking_asesores':      '/odoo/action-estate_reports.action_sales_by_user',
                    'tipo_propiedad':        '/odoo/action-estate_reports.action_sales_by_type',
                    'dias_mercado':          '/odoo/action-estate_reports.action_sales_days_market',
                    'comisiones':            '/odoo/action-estate_management.action_estate_commission_dashboard',
                    'kpis_ventas':           '/odoo/action-estate_reports.action_estate_sales_report_owl',
                    'analytics_propiedades': '/odoo/action-estate_reports.estate_analytics_property_action',
                    'analytics_contratos':   '/odoo/action-estate_reports.estate_analytics_contract_action',
                    'analytics_pagos':       '/odoo/action-estate_reports.estate_analytics_payment_action',
                    'analytics_ofertas':     '/odoo/action-estate_reports.estate_analytics_offer_action',
                    'analytics_gastos':      '/odoo/action-estate_reports.estate_analytics_expense_action',
                    'analytics_tasaciones':  '/odoo/action-estate_reports.estate_analytics_appraisal_action',
                    'analytics_mantenimiento': '/odoo/action-estate_reports.estate_analytics_maintenance_action',
                    'crm_pipeline':          '/odoo/action-estate_crm.action_crm_reports_general',
                    'crm_negocios':          '/odoo/action-estate_crm.action_crm_reports_deals',
                    'crm_fuentes':           '/odoo/action-estate_crm.action_crm_reports_sources',
                    'crm_visitas':           '/odoo/action-estate_crm.action_report_property_visits',
                    'social_facebook':       '/odoo/action-estate_social.action_facebook_stats_all',
                    'social_instagram':      '/odoo/action-estate_social.action_instagram_stats_all',
                    'exportar_pdf':          '/odoo/action-estate_reports.estate_report_wizard_action',
                    'exportar_excel':        '/odoo/action-estate_reports.action_estate_sales_report_owl',
                    'agenda_visitas':        '/odoo/action-estate_management.estate_property_action',
                }
                view_name = args.get('view_name', '')
                url = view_map.get(view_name)
                if not url:
                    available = ', '.join(view_map.keys())
                    return json.dumps({'error': f"Vista '{view_name}' no reconocida. Disponibles: {available}"})
                return json.dumps({
                    'success': True,
                    'url': url,
                    'mensaje': f"→ [**Abrir en Reportes**]({url})",
                })

            # ── A1: CRUD Contratos ─────────────────────────────────────────────
            elif tool_name == 'create_contract':
                prop = env['estate.property'].sudo().browse(int(args['property_id']))
                if not prop.exists():
                    return json.dumps({'error': f"Propiedad {args['property_id']} no encontrada"})
                partner = env['res.partner'].sudo().search(
                    [('name', 'ilike', args['partner_name'])], limit=1)
                if not partner:
                    partner = env['res.partner'].sudo().create({'name': args['partner_name']})
                from datetime import date as _date
                contract_vals = {
                    'property_id': prop.id,
                    'partner_id': partner.id,
                    'contract_type': args.get('contract_type', 'sale'),
                    'amount': float(args['amount']),
                    'date_start': args.get('date_start') or str(_date.today()),
                    'user_id': env.uid,
                }
                if args.get('date_end'):
                    contract_vals['date_end'] = args['date_end']
                if args.get('notes'):
                    contract_vals['notes'] = args['notes']
                contract = env['estate.contract'].sudo().create(contract_vals)
                contract.message_post(body='Contrato creado por el Agente IA.')
                return json.dumps({'success': True, 'contract_id': contract.id,
                    'ref': contract.name,
                    'mensaje': f"Contrato {contract.name} creado para '{partner.name}' — {prop.title} — ${float(args['amount']):,.2f}"})

            elif tool_name == 'update_contract':
                contract = env['estate.contract'].sudo().browse(int(args['contract_id']))
                if not contract.exists():
                    return json.dumps({'error': f"Contrato {args['contract_id']} no encontrado"})
                action = args.get('action', '')
                if action == 'activate':
                    contract.action_activate()
                elif action == 'cancel':
                    contract.action_cancel()
                elif action == 'expire':
                    contract.action_set_expired()
                if args.get('amount'):
                    contract.write({'amount': float(args['amount'])})
                note = args.get('notes', f'Actualizado vía Agente IA: {action}')
                contract.message_post(body=f'{note}')
                return json.dumps({'success': True,
                    'mensaje': f"Contrato {contract.name} actualizado — acción: {action}"})

            elif tool_name == 'create_payment':
                contract = env['estate.contract'].sudo().browse(int(args['contract_id']))
                if not contract.exists():
                    return json.dumps({'error': f"Contrato {args['contract_id']} no encontrado"})
                from datetime import date as _date
                pay_vals = {
                    'contract_id': contract.id,
                    'amount': float(args['amount']),
                    'payment_method': args.get('payment_method', 'bank'),
                    'date': args.get('date') or str(_date.today()),
                }
                if args.get('notes'):
                    pay_vals['notes'] = args['notes']
                payment = env['estate.payment'].sudo().create(pay_vals)
                return json.dumps({'success': True, 'payment_id': payment.id,
                    'ref': payment.name,
                    'mensaje': f"Pago {payment.name} de ${float(args['amount']):,.2f} registrado en contrato {contract.name}"})

            elif tool_name == 'approve_payment':
                payment = env['estate.payment'].sudo().browse(int(args['payment_id']))
                if not payment.exists():
                    return json.dumps({'error': f"Pago {args['payment_id']} no encontrado"})
                payment.action_confirm()
                return json.dumps({'success': True,
                    'mensaje': f"Pago {payment.name} de ${payment.amount:,.2f} marcado como pagado"})

            elif tool_name == 'create_offer':
                prop = env['estate.property'].sudo().browse(int(args['property_id']))
                if not prop.exists():
                    return json.dumps({'error': f"Propiedad {args['property_id']} no encontrada"})
                partner = env['res.partner'].sudo().search(
                    [('name', 'ilike', args['partner_name'])], limit=1)
                if not partner:
                    partner = env['res.partner'].sudo().create({'name': args['partner_name']})
                offer = env['estate.property.offer'].sudo().create({
                    'property_id': prop.id,
                    'partner_id': partner.id,
                    'offer_amount': float(args['offer_amount']),
                    'financing_type': args.get('financing_type', 'cash'),
                    'notes': args.get('notes', 'Oferta registrada por Agente IA'),
                })
                return json.dumps({'success': True, 'offer_id': offer.id,
                    'ref': offer.name,
                    'mensaje': f"Oferta {offer.name} de ${float(args['offer_amount']):,.2f} registrada para '{partner.name}' en {prop.title}"})

            elif tool_name == 'create_commission':
                prop = env['estate.property'].sudo().browse(int(args['property_id']))
                if not prop.exists():
                    return json.dumps({'error': f"Propiedad {args['property_id']} no encontrada"})
                user = env['res.users'].sudo().search(
                    [('name', 'ilike', args['agent_name'])], limit=1)
                if not user:
                    return json.dumps({'error': f"Asesor '{args['agent_name']}' no encontrado"})
                from datetime import date as _date
                comm = env['estate.commission'].sudo().create({
                    'property_id': prop.id,
                    'user_id': user.id,
                    'amount': float(args['amount']),
                    'type': args.get('commission_type', 'sale'),
                    'date': args.get('date') or str(_date.today()),
                })
                return json.dumps({'success': True, 'commission_id': comm.id,
                    'ref': comm.name,
                    'mensaje': f"Comisión {comm.name} de ${float(args['amount']):,.2f} registrada para {user.name}"})

            elif tool_name == 'approve_commission':
                comm = env['estate.commission'].sudo().browse(int(args['commission_id']))
                if not comm.exists():
                    return json.dumps({'error': f"Comisión {args['commission_id']} no encontrada"})
                action = args.get('action', 'approve')
                new_state = 'approved' if action == 'approve' else 'paid'
                comm.write({'state': new_state})
                label = 'APROBADA' if new_state == 'approved' else 'PAGADA'
                return json.dumps({'success': True,
                    'mensaje': f"Comisión {comm.name} de ${comm.amount:,.2f} marcada como {label}"})

            # ── A2: PDF ────────────────────────────────────────────────────────
            elif tool_name == 'generate_pdf_report':
                rtype = args.get('report_type', '')
                rid = int(args.get('record_id', 0))
                report_map = {
                    'ficha_propiedad':       ('estate_management.action_report_ficha_tecnica', 'estate.property'),
                    'estado_cuenta_contrato': ('estate_reports.action_report_contract_statement', 'estate.contract'),
                    'cotizacion_lead':        ('estate_crm.action_report_crm_quotation', 'crm.lead'),
                }
                if rtype not in report_map:
                    return json.dumps({'error': f"report_type '{rtype}' no reconocido. Usa: {', '.join(report_map)}"})
                import base64
                xmlid, model = report_map[rtype]
                report_action = env.ref(xmlid, raise_if_not_found=False)
                if not report_action:
                    return json.dumps({'error': f"Reporte {xmlid} no encontrado"})
                record = env[model].sudo().browse(rid)
                if not record.exists():
                    return json.dumps({'error': f"Registro {rid} no encontrado en {model}"})
                pdf_content, _ = report_action._render_qweb_pdf([rid])
                attachment = env['ir.attachment'].sudo().create({
                    'name': f'{rtype}_{rid}.pdf',
                    'datas': base64.b64encode(pdf_content).decode(),
                    'res_model': model,
                    'res_id': rid,
                    'mimetype': 'application/pdf',
                })
                url = f'/web/content/{attachment.id}?download=true'
                return json.dumps({'success': True, 'url': url,
                    'mensaje': f"PDF generado → [Descargar aquí]({url})"})

            # ── A4: Archivar ───────────────────────────────────────────────────
            elif tool_name == 'archive_property':
                prop = env['estate.property'].sudo().browse(int(args['property_id']))
                if not prop.exists():
                    return json.dumps({'error': f"Propiedad {args['property_id']} no encontrada"})
                title = prop.title
                reason = args.get('reason', 'Archivada vía Agente IA')
                prop.message_post(body=f'Archivada: {reason}')
                prop.write({'active': False})
                return json.dumps({'success': True,
                    'mensaje': f"Propiedad '{title}' archivada. Motivo: {reason}"})

            elif tool_name == 'cancel_payment':
                payment = env['estate.payment'].sudo().browse(int(args['payment_id']))
                if not payment.exists():
                    return json.dumps({'error': f"Pago {args['payment_id']} no encontrado"})
                reason = args.get('reason', 'Cancelado vía Agente IA')
                payment.message_post(body=f'Cancelado: {reason}')
                payment.action_cancel()
                return json.dumps({'success': True,
                    'mensaje': f"Pago {payment.name} anulado. Motivo: {reason}"})

            # ── A5: Lote ───────────────────────────────────────────────────────
            elif tool_name == 'batch_update_properties':
                domain = []
                if args.get('city'):
                    domain.append(('city', 'ilike', args['city']))
                if args.get('state_filter'):
                    domain.append(('state', '=', args['state_filter']))
                props = env['estate.property'].sudo().search(domain, limit=50)
                if not props:
                    return json.dumps({'error': 'No se encontraron propiedades con esos filtros'})
                vals = {}
                if args.get('new_state') and args['new_state'] in ('available', 'reserved', 'sold', 'rented'):
                    vals['state'] = args['new_state']
                if args.get('new_price_pct'):
                    pct = float(args['new_price_pct']) / 100
                    for p in props:
                        p.write({'price': round(p.price * (1 + pct), 2)})
                    if vals:
                        props.write(vals)
                    note = args.get('notes', f'Precio ajustado {args["new_price_pct"]}% vía Agente IA')
                    for p in props:
                        p.message_post(body=f'Lote: {note}')
                    return json.dumps({'success': True,
                        'affected': len(props),
                        'mensaje': f"{len(props)} propiedades actualizadas — precio {args['new_price_pct']}%"})
                if vals:
                    props.write(vals)
                note = args.get('notes', 'Actualización masiva vía Agente IA')
                for p in props:
                    p.message_post(body=f'Lote: {note}')
                return json.dumps({'success': True, 'affected': len(props),
                    'mensaje': f"{len(props)} propiedades actualizadas: {list(vals.keys())}"})

            elif tool_name == 'batch_archive_leads':
                from datetime import date as _date, timedelta
                domain = [('type', '=', 'lead'), ('active', '=', True)]
                if args.get('temperature'):
                    domain.append(('lead_temperature', '=', args['temperature']))
                if args.get('days_inactive'):
                    cutoff = str(_date.today() - timedelta(days=int(args['days_inactive'])))
                    domain.append(('write_date', '<', cutoff))
                leads = env['crm.lead'].sudo().search(domain, limit=50)
                if not leads:
                    return json.dumps({'error': 'No se encontraron leads con esos criterios'})
                reason = args.get('reason', 'Archivado masivo vía Agente IA')
                for l in leads:
                    l.message_post(body=f'Archivado: {reason}')
                leads.write({'active': False})
                return json.dumps({'success': True, 'affected': len(leads),
                    'mensaje': f"{len(leads)} leads archivados. Motivo: {reason}"})

            # ── A6: Email ──────────────────────────────────────────────────────
            elif tool_name == 'send_email':
                email_to = args.get('email_to', '')
                partner = None
                if args.get('partner_name') and not email_to:
                    partner = env['res.partner'].sudo().search(
                        [('name', 'ilike', args['partner_name'])], limit=1)
                    if partner:
                        email_to = partner.email or ''
                if not email_to:
                    return json.dumps({'error': 'No se encontró email del destinatario'})
                mail = env['mail.mail'].sudo().create({
                    'subject': args['subject'],
                    'body_html': args['body'],
                    'email_to': email_to,
                    'author_id': env.user.partner_id.id,
                })
                mail.send()
                return json.dumps({'success': True,
                    'mensaje': f"Email enviado a {email_to} — Asunto: {args['subject']}"})

            # ── B1: Lead Scoring IA ────────────────────────────────────────────
            elif tool_name == 'analyze_lead_probability':
                lead = env['crm.lead'].sudo().browse(int(args['lead_id']))
                if not lead.exists():
                    return json.dumps({'error': f"Lead {args['lead_id']} no encontrado"})
                # Recopilar historico de leads cerrados
                won_leads = env['crm.lead'].sudo().search([
                    ('stage_id.is_won', '=', True), ('type', '=', 'opportunity')], limit=30)
                lost_leads = env['crm.lead'].sudo().search([
                    ('active', '=', False), ('probability', '<', 10)], limit=20)
                won_summary = [{'budget': l.client_budget, 'match': l.match_percentage,
                    'visits': l.completed_visits_count if hasattr(l, 'completed_visits_count') else 0,
                    'score': l.lead_score, 'temp': l.lead_temperature} for l in won_leads[:15]]
                lost_summary = [{'budget': l.client_budget, 'match': l.match_percentage,
                    'score': l.lead_score, 'temp': l.lead_temperature} for l in lost_leads[:10]]
                return json.dumps({
                    'lead_id': lead.id, 'nombre': lead.name,
                    'cliente': lead.partner_id.name if lead.partner_id else lead.contact_name,
                    'presupuesto': lead.client_budget, 'match': lead.match_percentage,
                    'temperatura': lead.lead_temperature, 'puntuacion': lead.lead_score,
                    'visitas': getattr(lead, 'completed_visits_count', 0),
                    'historico_ganados': won_summary, 'historico_perdidos': lost_summary,
                    'instruccion': 'Con estos datos, calcula la probabilidad de cierre 0-100%, clasifica A/B/C, da 3 factores clave y una acción recomendada.',
                }, ensure_ascii=False)

            # ── B2: Churn ──────────────────────────────────────────────────────
            elif tool_name == 'analyze_churn_risk':
                from datetime import date as _date, timedelta
                today = _date.today()
                days = int(args.get('days_to_expiry', 60))
                limit_date = today + timedelta(days=days)
                contracts = env['estate.contract'].sudo().search([
                    ('state', '=', 'active'), ('contract_type', '=', 'rent'),
                    ('date_end', '!=', False), ('date_end', '<=', str(limit_date)),
                ])
                result = []
                for c in contracts:
                    days_left = (c.date_end - today).days if c.date_end else 999
                    overdue_payments = env['estate.payment'].sudo().search_count([
                        ('contract_id', '=', c.id), ('state', '=', 'pending'),
                        ('date', '<', str(today)),
                    ])
                    maintenance_open = 0
                    if 'estate.tenant.request' in env:
                        maintenance_open = env['estate.tenant.request'].sudo().search_count([
                            ('property_id', '=', c.property_id.id), ('state', 'not in', ('done', 'cancel')),
                        ])
                    risk = 'ALTO' if (days_left < 30 or overdue_payments >= 2) else \
                           'MEDIO' if (days_left < 45 or overdue_payments == 1) else 'BAJO'
                    result.append({
                        'contrato': c.name, 'cliente': c.partner_id.name,
                        'propiedad': c.property_id.title,
                        'dias_para_vencer': days_left, 'pagos_vencidos': overdue_payments,
                        'mantenimientos_abiertos': maintenance_open, 'riesgo': risk,
                    })
                result.sort(key=lambda x: {'ALTO': 0, 'MEDIO': 1, 'BAJO': 2}[x['riesgo']])
                return json.dumps({'total_analizados': len(result), 'contratos': result,
                    'alto_riesgo': sum(1 for r in result if r['riesgo'] == 'ALTO')},
                    ensure_ascii=False)

            # ── B3: AVM IA ─────────────────────────────────────────────────────
            elif tool_name == 'recalculate_avm_ai':
                prop = env['estate.property'].sudo().browse(int(args['property_id']))
                if not prop.exists():
                    return json.dumps({'error': f"Propiedad {args['property_id']} no encontrada"})
                comp_domain = [('state', '=', 'sold'), ('property_type_id', '=', prop.property_type_id.id)]
                if prop.city:
                    comp_domain.append(('city', 'ilike', prop.city))
                comparables = env['estate.property'].sudo().search(comp_domain, limit=20)
                comp_data = [{'precio': c.price, 'area': c.area or 0,
                    'habitaciones': c.bedrooms or 0, 'ciudad': c.city,
                    'precio_m2': round(c.price / c.area, 2) if c.area else 0,
                    'dias_venta': c.days_on_market or 0} for c in comparables]
                return json.dumps({
                    'propiedad': {'id': prop.id, 'titulo': prop.title, 'precio_actual': prop.price,
                        'area': prop.area or 0, 'habitaciones': prop.bedrooms or 0,
                        'ciudad': prop.city, 'tipo': prop.property_type_id.name},
                    'comparables': comp_data,
                    'instruccion': (
                        'Con los comparables, calcula el valor justo de mercado de esta propiedad. '
                        'Devuelve: {"valor_estimado": X, "rango_min": Y, "rango_max": Z, '
                        '"confianza_pct": W, "estado_avm": "fair|high|low", "factores": [...], "recomendacion": "..."}'
                    ),
                }, ensure_ascii=False)

            # ── B4: Descripción + WP ───────────────────────────────────────────
            elif tool_name == 'generate_and_apply_description':
                prop = env['estate.property'].sudo().browse(int(args['property_id']))
                if not prop.exists():
                    return json.dumps({'error': f"Propiedad {args['property_id']} no encontrada"})
                style = args.get('style', 'emocional')
                style_map = {
                    'formal': 'profesional y técnico, orientado a inversores',
                    'emocional': 'emotivo y aspiracional, orientado a familias',
                    'directo': 'conciso y directo, con bullet points de beneficios',
                }
                style_desc = style_map.get(style, style_map['emocional'])
                prop_info = (
                    f"Título: {prop.title}\nTipo: {prop.property_type_id.name}\n"
                    f"Ciudad: {prop.city}\nÁrea: {prop.area}m²\n"
                    f"Habitaciones: {prop.bedrooms}\nBaños: {prop.bathrooms}\n"
                    f"Precio: ${prop.price:,.0f}\n"
                    f"Descripción actual: {prop.description or '(sin descripción)'}"
                )
                # If apply=True, we return instructions for the LLM to call update_property after generating
                result = {
                    'property_id': prop.id, 'titulo': prop.title,
                    'info': prop_info, 'estilo': style_desc,
                    'instruccion': (
                        f'Genera una descripción de marketing estilo {style_desc} para esta propiedad. '
                        'Luego, si apply=True, llama a update_property para guardarla.'
                    ),
                    'apply': args.get('apply', False),
                    'publish_wp': args.get('publish_wp', False),
                }
                if args.get('apply') and prop.description:
                    # Already has a description — signal to apply immediately after generation
                    result['nota'] = 'Después de generar, usa update_property para actualizar la descripción.'
                if args.get('publish_wp'):
                    # Check if WordPress integration available
                    has_wp = 'wp_published' in env['estate.property']._fields
                    result['wp_available'] = has_wp
                    if has_wp:
                        result['nota_wp'] = 'Después de actualizar la descripción, llama a update_property con state=available para publicar en WP.'
                return json.dumps(result, ensure_ascii=False)

            # ── PDF DE REPORTE ANALÍTICO ───────────────────────────────────────
            elif tool_name == 'generate_analytics_pdf':
                import base64, subprocess, tempfile, os
                rtype = args.get('report_type', '')
                title = args.get('title', '') or rtype.replace('_', ' ').title()
                limit = int(args.get('limit', 30))

                # Reuse get_report_data internally
                raw = self._execute_tool('get_report_data', {'report_type': rtype, 'limit': limit}, env=env)
                rdata = json.loads(raw)
                if 'error' in rdata:
                    return json.dumps({'error': rdata['error']})

                data_dict = rdata.get('data', {})
                detalle = rdata.get('detalle', [])
                report_title = rdata.get('report', title)
                from datetime import datetime as _dt
                fecha = _dt.now().strftime('%d/%m/%Y %H:%M')

                # Build HTML rows from detalle (list of dicts) or data_dict
                rows_html = ''
                if detalle:
                    headers = list(detalle[0].keys())
                    thead = ''.join(f'<th>{h.replace("_"," ").title()}</th>' for h in headers)
                    rows_html = f'<thead><tr>{thead}</tr></thead><tbody>'
                    for row in detalle:
                        cells = ''.join(f'<td>{str(row.get(h,"") or "")}</td>' for h in headers)
                        rows_html += f'<tr>{cells}</tr>'
                    rows_html += '</tbody>'
                elif data_dict:
                    rows_html = '<thead><tr><th>Categoría</th><th>Valor</th></tr></thead><tbody>'
                    for k, v in data_dict.items():
                        rows_html += f'<tr><td>{k}</td><td>{v}</td></tr>'
                    rows_html += '</tbody>'

                html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<title>{report_title}</title>
<style>
  body{{font-family:Arial,sans-serif;padding:32px;color:#222;font-size:12px}}
  h1{{color:#004274;font-size:18px;margin-bottom:4px}}
  .meta{{color:#888;font-size:10px;margin-bottom:20px}}
  table{{width:100%;border-collapse:collapse;margin-top:12px}}
  th{{background:#004274;color:white;padding:8px 10px;text-align:left;font-size:11px}}
  td{{padding:7px 10px;border-bottom:1px solid #e8ecf0;font-size:11px}}
  tr:nth-child(even){{background:#f5f7fa}}
  .footer{{margin-top:24px;font-size:9px;color:#aaa;border-top:1px solid #eee;padding-top:10px}}
</style></head><body>
<h1>{report_title}</h1>
<div class="meta">Generado: {fecha} — Sistema Inmobiliario InmoBot</div>
<table>{rows_html}</table>
<div class="footer">Reporte generado automáticamente por el Agente IA — InmoBot</div>
</body></html>"""

                try:
                    with tempfile.NamedTemporaryFile(suffix='.html', delete=False, mode='w', encoding='utf-8') as fh:
                        fh.write(html)
                        html_path = fh.name
                    pdf_path = html_path.replace('.html', '.pdf')
                    subprocess.run(
                        ['wkhtmltopdf', '--quiet', '--page-size', 'A4',
                         '--margin-top', '15mm', '--margin-bottom', '15mm',
                         '--margin-left', '15mm', '--margin-right', '15mm',
                         html_path, pdf_path],
                        check=True, timeout=30
                    )
                    with open(pdf_path, 'rb') as pf:
                        pdf_bytes = pf.read()
                    os.unlink(html_path)
                    os.unlink(pdf_path)

                    attachment = env['ir.attachment'].sudo().create({
                        'name': f'{rtype}_{fecha[:10].replace("/","-")}.pdf',
                        'datas': base64.b64encode(pdf_bytes).decode(),
                        'res_model': 'estate.property',
                        'res_id': 0,
                        'mimetype': 'application/pdf',
                    })
                    url = f'/web/content/{attachment.id}?download=true'
                    filas = len(detalle) or len(data_dict)
                    return json.dumps({
                        'success': True, 'url': url,
                        'mensaje': f'PDF "{report_title}" generado con {filas} registros → [**Descargar PDF aquí**]({url})',
                    })
                except subprocess.TimeoutExpired:
                    return json.dumps({'error': 'Tiempo de generación agotado. Intenta con menos filas.'})
                except Exception as e:
                    return json.dumps({'error': f'Error generando PDF: {str(e)}'})

            # ── PACK DE MARKETING ──────────────────────────────────────────────
            elif tool_name == 'generate_marketing_pack':
                prop = env['estate.property'].sudo().browse(int(args.get('property_id', 0)))
                if not prop.exists():
                    return json.dumps({'error': f"Propiedad {args.get('property_id')} no encontrada"})

                style = args.get('style', 'emocional')
                channels = args.get('channels') or [
                    'instagram', 'facebook', 'whatsapp',
                    'email_asunto', 'email_cuerpo',
                    'google_ads', 'puntos_clave', 'slogan',
                ]

                img_count = len(prop.image_ids) if hasattr(prop, 'image_ids') and prop.image_ids else 0
                visit_count = 0
                if 'property_id' in env['calendar.event']._fields:
                    visit_count = env['calendar.event'].sudo().search_count([('property_id', '=', prop.id)])

                avm_price = getattr(prop, 'avm_estimated_price', 0) or 0
                avm_status = getattr(prop, 'avm_status', '') or ''
                avm_note = f"Valor de mercado estimado: ${avm_price:,.0f} ({avm_status})" if avm_price else ''

                tags = []
                if hasattr(prop, 'tag_ids'):
                    tags = [t.name for t in prop.tag_ids]

                prop_info = {
                    'id': prop.id,
                    'ref': prop.name,
                    'titulo': prop.title or '',
                    'tipo': prop.property_type_id.name if prop.property_type_id else '',
                    'ciudad': prop.city or '',
                    'calle': prop.street or '',
                    'precio': prop.price,
                    'precio_fmt': f"${prop.price:,.0f}",
                    'area': prop.area,
                    'habitaciones': prop.bedrooms,
                    'banos': prop.bathrooms,
                    'parking': getattr(prop, 'parking_spaces', 0) or 0,
                    'piso': getattr(prop, 'floor', 0) or 0,
                    'anio_construccion': getattr(prop, 'year_built', 0) or 0,
                    'tipo_operacion': 'en venta' if prop.offer_type == 'sale' else 'en arriendo',
                    'descripcion_actual': (prop.description or '')[:800],
                    'avm_info': avm_note,
                    'imagenes_disponibles': img_count,
                    'visitas_realizadas': visit_count,
                    'asesor': prop.user_id.name if prop.user_id else '',
                    'asesor_extra_prompt': getattr(prop, 'ai_extra_prompt', '') or '',
                    'tags': tags,
                }

                style_map = {
                    'emocional': 'emotivo y aspiracional, con storytelling que conecta emocionalmente con familias y parejas, usa emojis moderados y lenguaje cercano',
                    'formal': 'profesional y técnico, enfocado en datos concretos de inversión, ROI y plusvalía, sin emojis, para inversores y empresas',
                    'directo': 'conciso y directo, bullet points claros, CTA fuerte, genera urgencia con frases como "disponibilidad limitada", usa emojis de checkmark',
                    'lujoso': 'premium y exclusivo, lenguaje sofisticado, destaca exclusividad y estilo de vida, para compradores de alto poder adquisitivo',
                }
                style_desc = style_map.get(style, style_map['emocional'])
                channels_str = ', '.join(channels)

                return json.dumps({
                    'success': True,
                    'property': prop_info,
                    'estilo': style_desc,
                    'canales': channels,
                    'instruccion': (
                        f'Con los datos de la propiedad, genera el pack de marketing estilo: "{style_desc}". '
                        f'Canales requeridos: {channels_str}. '
                        'USA EXACTAMENTE ESTE FORMATO para cada canal (las etiquetas son obligatorias): '
                        '[PACK:instagram]caption emocionante con emojis + 25 hashtags relevantes[/PACK] '
                        '[PACK:facebook]post narrativo completo con emojis y CTA al final[/PACK] '
                        '[PACK:whatsapp]mensaje corto y directo, máx 3 párrafos, incluye precio y datos clave[/PACK] '
                        '[PACK:email_asunto]Asunto del email, máx 60 caracteres, sin emojis[/PACK] '
                        '[PACK:email_cuerpo]Cuerpo del email completo, formal, con saludo y firma[/PACK] '
                        '[PACK:google_ads]Titular (≤30 chars)\n---\nDescripción (≤90 chars)[/PACK] '
                        '[PACK:puntos_clave]• 6 bullets con los atributos más vendibles de la propiedad[/PACK] '
                        '[PACK:slogan]Una frase memorable de máx 10 palabras[/PACK] '
                        'IMPORTANTE: Incluye datos reales (precio, área, habitaciones, ciudad). '
                        'Si hay prompt extra del asesor, incorpóralo naturalmente en todos los copies.'
                    ),
                    'save_description': args.get('save_description', False),
                }, ensure_ascii=False)

            # ── B5: Memoria ────────────────────────────────────────────────────
            elif tool_name == 'save_memory':
                if 'estate.ai.memory' not in env:
                    return json.dumps({'error': 'Módulo de memoria no instalado aún'})
                mem = env['estate.ai.memory'].sudo().create({
                    'user_id': env.uid,
                    'content': args['content'],
                    'memory_type': args.get('memory_type', 'fact'),
                })
                return json.dumps({'success': True, 'memory_id': mem.id,
                    'mensaje': f"Memorizado: {args['content'][:80]}"})

            elif tool_name == 'recall_memory':
                if 'estate.ai.memory' not in env:
                    return json.dumps({'memories': [], 'mensaje': 'Sin memorias guardadas aún'})
                domain = [('user_id', '=', env.uid)]
                query = args.get('query', '')
                if query:
                    domain.append(('content', 'ilike', query))
                memories = env['estate.ai.memory'].sudo().search(domain, order='create_date desc', limit=20)
                return json.dumps({'memories': [
                    {'id': m.id, 'tipo': m.memory_type, 'contenido': m.content,
                     'fecha': str(m.create_date)[:10]} for m in memories
                ]}, ensure_ascii=False)

            # ── C1: Búsqueda de Contactos ──────────────────────────────────────
            elif tool_name == 'search_contacts':
                from datetime import date as _date
                query = args.get('query', '')
                limit = int(args.get('limit', 10))
                domain = [('active', '=', True)]
                if query:
                    domain = ['|', '|', ('name', 'ilike', query),
                              ('email', 'ilike', query), ('phone', 'ilike', query)]
                partners = env['res.partner'].sudo().search(domain, limit=limit)
                result = []
                for p in partners:
                    # count leads and contracts
                    lead_count = env['crm.lead'].sudo().search_count([('partner_id', '=', p.id)])
                    contract_count = env['estate.contract'].sudo().search_count([
                        ('partner_id', '=', p.id), ('state', '=', 'active')])
                    if args.get('has_contracts') and contract_count == 0:
                        continue
                    result.append({
                        'id': p.id,
                        'nombre': p.name,
                        'email': p.email or '',
                        'telefono': p.phone or '',
                        'empresa': p.parent_id.name if p.parent_id else '',
                        'leads_activos': lead_count,
                        'contratos_activos': contract_count,
                    })
                return json.dumps(result[:limit], ensure_ascii=False)

            # ── C2: Comparar Propiedades ────────────────────────────────────────
            elif tool_name == 'compare_properties':
                ids = args.get('property_ids', [])
                if len(ids) < 2:
                    return json.dumps({'error': 'Se requieren al menos 2 IDs de propiedades para comparar'})
                props = env['estate.property'].sudo().browse(ids).filtered('id')
                result = []
                for p in props:
                    result.append({
                        'id': p.id,
                        'ref': p.name,
                        'titulo': p.title,
                        'ciudad': p.city,
                        'precio': p.price,
                        'area_m2': p.area,
                        'habitaciones': p.bedrooms,
                        'banos': p.bathrooms,
                        'tipo': p.property_type_id.name if p.property_type_id else '',
                        'estado': p.state,
                        'dias_mercado': p.days_on_market,
                        'precio_m2': round(p.price / p.area, 2) if p.area else None,
                        'avm_precio': p.avm_estimated_price,
                        'avm_status': p.avm_status,
                        'asesor': p.user_id.name if p.user_id else '',
                    })
                return json.dumps({'comparacion': result}, ensure_ascii=False)

            # ── C3: Análisis de Tendencias ──────────────────────────────────────
            elif tool_name == 'get_trend_analysis':
                from datetime import date as _date, timedelta
                metric = args.get('metric', 'all')
                period = args.get('period', 'month')
                today = _date.today()

                if period == 'month':
                    cur_start = today.replace(day=1)
                    prev_end = cur_start - timedelta(days=1)
                    prev_start = prev_end.replace(day=1)
                    cur_label = cur_start.strftime('%B %Y')
                    prev_label = prev_start.strftime('%B %Y')
                elif period == 'quarter':
                    q = (today.month - 1) // 3
                    cur_start = today.replace(month=q * 3 + 1, day=1)
                    prev_end = cur_start - timedelta(days=1)
                    prev_start = prev_end.replace(day=1).replace(month=((prev_end.month - 1) // 3) * 3 + 1)
                    cur_label = f'Q{q+1} {today.year}'
                    prev_label = f'Q{((prev_end.month - 1) // 3) + 1} {prev_end.year}'
                else:  # year
                    cur_start = today.replace(month=1, day=1)
                    prev_start = cur_start.replace(year=today.year - 1)
                    prev_end = cur_start - timedelta(days=1)
                    cur_label = str(today.year)
                    prev_label = str(today.year - 1)

                trends = {}

                if metric in ('sales', 'all'):
                    cur_sales = env['estate.property'].sudo().search_count([
                        ('state', '=', 'sold'), ('date_sold', '>=', str(cur_start))])
                    prev_sales = env['estate.property'].sudo().search_count([
                        ('state', '=', 'sold'),
                        ('date_sold', '>=', str(prev_start)),
                        ('date_sold', '<=', str(prev_end))])
                    delta = cur_sales - prev_sales
                    trends['ventas'] = {
                        cur_label: cur_sales, prev_label: prev_sales,
                        'variacion': f"{'+' if delta >= 0 else ''}{delta}",
                        'tendencia': 'subida' if delta > 0 else ('bajada' if delta < 0 else 'igual'),
                    }

                if metric in ('leads', 'all'):
                    cur_leads = env['crm.lead'].sudo().search_count([
                        ('create_date', '>=', str(cur_start))])
                    prev_leads = env['crm.lead'].sudo().search_count([
                        ('create_date', '>=', str(prev_start)),
                        ('create_date', '<=', str(prev_end))])
                    delta = cur_leads - prev_leads
                    trends['leads_nuevos'] = {
                        cur_label: cur_leads, prev_label: prev_leads,
                        'variacion': f"{'+' if delta >= 0 else ''}{delta}",
                        'tendencia': 'subida' if delta > 0 else ('bajada' if delta < 0 else 'igual'),
                    }

                if metric in ('revenue', 'all'):
                    cur_rev_props = env['estate.property'].sudo().search([
                        ('state', '=', 'sold'), ('date_sold', '>=', str(cur_start))])
                    prev_rev_props = env['estate.property'].sudo().search([
                        ('state', '=', 'sold'),
                        ('date_sold', '>=', str(prev_start)),
                        ('date_sold', '<=', str(prev_end))])
                    cur_rev = sum(cur_rev_props.mapped('price'))
                    prev_rev = sum(prev_rev_props.mapped('price'))
                    delta_pct = round((cur_rev - prev_rev) / prev_rev * 100, 1) if prev_rev else 0
                    trends['volumen_casas'] = {
                        cur_label: round(cur_rev, 2), prev_label: round(prev_rev, 2),
                        'variacion_pct': f"{'+' if delta_pct >= 0 else ''}{delta_pct}%",
                        'tendencia': 'subida' if delta_pct > 0 else ('bajada' if delta_pct < 0 else 'igual'),
                    }
                    cur_comm = sum(cur_rev_props.mapped('commission_amount'))
                    prev_comm = sum(prev_rev_props.mapped('commission_amount'))
                    delta_comm_pct = round((cur_comm - prev_comm) / prev_comm * 100, 1) if prev_comm else 0
                    trends['comisiones_agencia'] = {
                        cur_label: round(cur_comm, 2), prev_label: round(prev_comm, 2),
                        'variacion_pct': f"{'+' if delta_comm_pct >= 0 else ''}{delta_comm_pct}%",
                        'tendencia': 'subida' if delta_comm_pct > 0 else ('bajada' if delta_comm_pct < 0 else 'igual'),
                    }

                if metric in ('days_on_market', 'all'):
                    sold_now = env['estate.property'].sudo().search([
                        ('state', '=', 'sold'), ('date_sold', '>=', str(cur_start)),
                        ('days_on_market', '>', 0)])
                    sold_prev = env['estate.property'].sudo().search([
                        ('state', '=', 'sold'),
                        ('date_sold', '>=', str(prev_start)),
                        ('date_sold', '<=', str(prev_end)),
                        ('days_on_market', '>', 0)])
                    cur_dom = round(sum(sold_now.mapped('days_on_market')) / len(sold_now), 1) if sold_now else 0
                    prev_dom = round(sum(sold_prev.mapped('days_on_market')) / len(sold_prev), 1) if sold_prev else 0
                    delta = round(cur_dom - prev_dom, 1)
                    trends['dias_en_mercado_promedio'] = {
                        cur_label: cur_dom, prev_label: prev_dom,
                        'variacion': f"{'+' if delta >= 0 else ''}{delta}",
                        'tendencia': 'subida' if delta < 0 else ('bajada' if delta > 0 else 'igual'),
                    }

                return json.dumps({'periodo': period, 'tendencias': trends}, ensure_ascii=False)

            # ── C4: Próximas Visitas ────────────────────────────────────────────
            elif tool_name == 'get_upcoming_visits':
                from datetime import datetime as _dt, timedelta
                days = int(args.get('days_ahead', 7))
                now = _dt.now()
                future = now + timedelta(days=days)
                domain = [
                    ('start', '>=', str(now)),
                    ('start', '<=', str(future)),
                ]
                if args.get('advisor_name'):
                    domain.append(('user_id.name', 'ilike', args['advisor_name']))
                # Try estate calendar events first
                try:
                    events = env['calendar.event'].sudo().search(domain, order='start asc', limit=20)
                    result = []
                    for e in events:
                        prop = getattr(e, 'property_id', None)
                        result.append({
                            'id': e.id,
                            'titulo': e.name,
                            'inicio': str(e.start)[:16],
                            'fin': str(e.stop)[:16] if e.stop else '',
                            'propiedad': prop.title if prop else '',
                            'asesor': e.user_id.name if e.user_id else '',
                            'tipo': getattr(e, 'appointment_type', '') or '',
                            'estado_visita': getattr(e, 'visit_state', '') or '',
                        })
                    return json.dumps({'visitas': result, 'total': len(result)}, ensure_ascii=False)
                except Exception as ev_err:
                    return json.dumps({'error': str(ev_err)})

            # ── C5: Resumen de Cliente ──────────────────────────────────────────
            elif tool_name == 'get_client_summary':
                partner = None
                if args.get('partner_id'):
                    partner = env['res.partner'].sudo().browse(int(args['partner_id']))
                    if not partner.exists():
                        partner = None
                if not partner and args.get('partner_name'):
                    partner = env['res.partner'].sudo().search(
                        [('name', 'ilike', args['partner_name'])], limit=1)
                if not partner:
                    return json.dumps({'error': 'Cliente no encontrado'})

                leads = env['crm.lead'].sudo().search([('partner_id', '=', partner.id)], limit=10)
                contracts = env['estate.contract'].sudo().search([('partner_id', '=', partner.id)])
                payments_pending = env['estate.payment'].sudo().search([
                    ('partner_id', '=', partner.id), ('state', '=', 'pending')]) if hasattr(env['estate.payment']._fields, 'partner_id') else []
                try:
                    visits = env['calendar.event'].sudo().search([
                        ('partner_ids', 'in', [partner.id])], order='start desc', limit=5)
                    visit_list = [{'titulo': v.name, 'fecha': str(v.start)[:16]} for v in visits]
                except Exception:
                    visit_list = []

                summary = {
                    'id': partner.id,
                    'nombre': partner.name,
                    'email': partner.email or '',
                    'telefono': partner.phone or '',
                    'leads': [{'id': l.id, 'nombre': l.name, 'temperatura': l.lead_temperature,
                                'etapa': l.stage_id.name if l.stage_id else '',
                                'presupuesto': l.client_budget} for l in leads],
                    'contratos': [{'id': c.id, 'tipo': c.contract_type,
                                    'estado': c.state, 'monto': c.amount,
                                    'propiedad': c.property_id.title if c.property_id else ''
                                    } for c in contracts],
                    'ultimas_visitas': visit_list,
                    'total_leads': len(leads),
                    'total_contratos': len(contracts),
                }
                return json.dumps(summary, ensure_ascii=False)

            # ── C6: Cotización PDF ──────────────────────────────────────────────
            elif tool_name == 'generate_quote_pdf':
                lead_id = int(args.get('lead_id', 0))
                if not lead_id:
                    return json.dumps({'error': 'lead_id requerido'})
                lead = env['crm.lead'].sudo().browse(lead_id)
                if not lead.exists():
                    return json.dumps({'error': f'Lead #{lead_id} no encontrado'})
                # Use the property from arg or from lead
                prop_id = args.get('property_id') or (lead.target_property_id.id if lead.target_property_id else None)
                if not prop_id:
                    return json.dumps({'error': 'No hay propiedad asignada al lead. Asigna una primero con update_lead.'})
                base_url = env['ir.config_parameter'].sudo().get_param('web.base.url', 'http://localhost:8069')
                pdf_url = f"{base_url}/report/pdf/estate_crm.action_report_cotizacion_lead/{lead_id}"
                return json.dumps({
                    'success': True,
                    'pdf_url': pdf_url,
                    'mensaje': f"Cotización generada para Lead #{lead_id} — [{lead.name}]({pdf_url})",
                    'link': pdf_url,
                }, ensure_ascii=False)

            # ── PLAN DE CAMPAÑA DE MARKETING ──────────────────────────────────
            elif tool_name == 'plan_marketing_campaign':
                property_id = int(args.get('property_id', 0))
                prop = env['estate.property'].sudo().browse(property_id)
                if not prop.exists():
                    return json.dumps({'error': f'Propiedad {property_id} no encontrada'})

                # Conteo de imágenes adjuntas
                img_count = env['ir.attachment'].sudo().search_count([
                    ('res_model', '=', 'estate.property'),
                    ('res_id', '=', property_id),
                    ('mimetype', 'like', 'image%'),
                ])

                price = prop.price or 0
                is_rent = prop.offer_type == 'rent'
                beds = prop.bedrooms or 0
                prop_type = prop.property_type_id.name if prop.property_type_id else 'propiedad'
                city = prop.city or 'Ecuador'

                # Canal principal recomendado
                if price >= 200000:
                    primary_channel = 'Instagram'
                    channel_reason = 'Propiedades premium → audiencia visual de alto poder adquisitivo'
                elif is_rent:
                    primary_channel = 'Facebook Marketplace + WhatsApp'
                    channel_reason = 'Arrendamientos → búsqueda local activa e inmediata'
                elif price >= 80000:
                    primary_channel = 'Facebook + WhatsApp'
                    channel_reason = 'Rango medio → alcance masivo local con conversión directa'
                else:
                    primary_channel = 'WhatsApp Broadcast + Facebook'
                    channel_reason = 'Precio accesible → difusión directa y rápida'

                # Buyer persona
                if is_rent:
                    if beds <= 1:
                        persona = 'Profesional joven 25-35 años, soltero/pareja, trabaja en zona urbana'
                        persona_needs = 'Ubicación, transporte, servicios cercanos, precio'
                    else:
                        persona = 'Familia joven 28-42 años, 1-2 hijos, busca estabilidad'
                        persona_needs = 'Seguridad del sector, colegios cercanos, espacio'
                else:
                    if price < 80000:
                        persona = 'Primer comprador 25-38 años con acceso a crédito bancario'
                        persona_needs = 'Precio justo, facilidades de pago, plusvalía futura'
                    elif price < 200000:
                        persona = 'Familia establecida 35-52 años, mejora de residencia o segunda vivienda'
                        persona_needs = 'Calidad de construcción, sector, acabados, funcionalidad'
                    else:
                        persona = 'Inversionista o ejecutivo 40-58 años, compra al contado o crédito propio'
                        persona_needs = 'ROI, exclusividad, ubicación premium, acabados de lujo'

                # Presupuesto Facebook Ads mensual sugerido
                if price > 0:
                    fb_budget = max(50, min(600, round(price * 0.002 / 30) * 30))
                else:
                    fb_budget = 80

                # Urgencia según días en mercado
                dom = prop.days_on_market or 0
                if dom > 60:
                    urgency = f'ALTA — {dom} días sin vender. Considera reducir precio o intensificar campaña urgente.'
                elif dom > 30:
                    urgency = f'MEDIA — {dom} días en el mercado. Momento clave para reforzar la difusión.'
                else:
                    urgency = f'NORMAL — propiedad reciente ({dom} días). Establece presencia desde el inicio.'

                # Checklist de publicación
                checklist = []
                if img_count == 0:
                    checklist.append({'item': 'Fotos de la propiedad', 'status': 'FALTA', 'priority': 'crítico'})
                elif img_count < 5:
                    checklist.append({'item': f'Más fotos ({img_count} de mínimo 5 recomendadas)', 'status': 'MEJORAR', 'priority': 'alto'})
                else:
                    checklist.append({'item': f'Fotos ({img_count} disponibles)', 'status': 'OK', 'priority': 'ok'})

                desc_len = len(prop.description or '')
                if desc_len < 50:
                    checklist.append({'item': 'Descripción detallada de la propiedad', 'status': 'FALTA', 'priority': 'crítico'})
                elif desc_len < 200:
                    checklist.append({'item': 'Ampliar descripción (muy corta)', 'status': 'MEJORAR', 'priority': 'alto'})
                else:
                    checklist.append({'item': 'Descripción completa', 'status': 'OK', 'priority': 'ok'})

                wp_published = getattr(prop, 'wp_published', False)
                if not wp_published:
                    checklist.append({'item': 'Publicar en sitio web WordPress', 'status': 'PENDIENTE', 'priority': 'alto'})
                else:
                    checklist.append({'item': 'Publicado en WordPress', 'status': 'OK', 'priority': 'ok'})

                avm_status = getattr(prop, 'avm_status', '') or ''
                if avm_status in ('', 'pending'):
                    checklist.append({'item': 'Calcular valoración AVM (comparar precio vs mercado)', 'status': 'PENDIENTE', 'priority': 'medio'})
                else:
                    checklist.append({'item': f'Valoración AVM calculada ({avm_status})', 'status': 'OK', 'priority': 'ok'})

                lat = getattr(prop, 'latitude', None)
                lng = getattr(prop, 'longitude', None)
                if not (lat and lng):
                    checklist.append({'item': 'Registrar coordenadas GPS (mejora visibilidad en mapas)', 'status': 'PENDIENTE', 'priority': 'medio'})
                else:
                    checklist.append({'item': 'Coordenadas GPS registradas', 'status': 'OK', 'priority': 'ok'})

                # SEO keywords
                action_word = 'arriendo' if is_rent else 'venta'
                seo_keywords = [
                    f"{prop_type} en {action_word} {city}",
                    f"{prop_type} {beds} dormitorios {city}" if beds else f"{prop_type} {city}",
                    f"inmueble en {action_word} {city} Ecuador",
                    f"alquilar {prop_type} {city}" if is_rent else f"comprar {prop_type} {city}",
                    f"{prop_type} económico {city}" if price < 100000 else f"{prop_type} {city} precio",
                ]

                # Calendario de contenidos
                calendar = [
                    {'day': 'Lunes', 'content': 'Foto exterior + datos clave en Instagram y Facebook'},
                    {'day': 'Miércoles', 'content': 'Tour de interiores en Stories + destaca una característica única'},
                    {'day': 'Viernes', 'content': 'WhatsApp broadcast a lista de interesados activos'},
                    {'day': 'Sábado', 'content': 'Reel/video corto (mayor alcance orgánico en fin de semana)'},
                ]

                return json.dumps({
                    'property_name': prop.title or prop.name or f'Propiedad #{property_id}',
                    'property_ref': prop.name or '',
                    'property_type': prop_type,
                    'price_fmt': f'${price:,.0f}',
                    'offer_type': 'Arriendo' if is_rent else 'Venta',
                    'city': city,
                    'primary_channel': primary_channel,
                    'channel_reason': channel_reason,
                    'buyer_persona': persona,
                    'persona_needs': persona_needs,
                    'fb_ads_budget_monthly': fb_budget,
                    'urgency': urgency,
                    'checklist': checklist,
                    'seo_keywords': seo_keywords,
                    'content_calendar': calendar,
                    'img_count': img_count,
                }, ensure_ascii=False)

            # ── HERRAMIENTA UNIVERSAL: SQL de solo lectura ──────────────────
            elif tool_name == 'query_database':
                sql = (args.get('sql') or '').strip()
                if not sql:
                    return json.dumps({'error': 'Se requiere una consulta SQL'})
                # Security: only allow SELECT statements
                sql_upper = sql.upper().lstrip()
                forbidden = ['INSERT', 'UPDATE', 'DELETE', 'DROP', 'ALTER', 'CREATE',
                             'TRUNCATE', 'GRANT', 'REVOKE', 'EXECUTE', 'COPY']
                if not sql_upper.startswith('SELECT') and not sql_upper.startswith('WITH'):
                    return json.dumps({'error': 'Solo se permiten consultas SELECT (solo lectura)'})
                for word in forbidden:
                    # Check for forbidden keywords as whole words (not inside column names)
                    import re
                    if re.search(rf'\b{word}\b', sql_upper):
                        return json.dumps({'error': f'Operación {word} no permitida. Solo SELECT.'})
                # Force LIMIT if not present
                if 'LIMIT' not in sql_upper:
                    sql = sql.rstrip(';') + ' LIMIT 50'
                try:
                    env.cr.execute(sql)
                    columns = [desc[0] for desc in env.cr.description] if env.cr.description else []
                    rows = env.cr.dictfetchall()
                    return json.dumps({
                        'columns': columns,
                        'rows': rows,
                        'row_count': len(rows),
                        'explanation': args.get('explanation', ''),
                    }, ensure_ascii=False, default=str)
                except Exception as sql_err:
                    return json.dumps({'error': f'Error SQL: {str(sql_err)}'})

            # ── BASE DE CONOCIMIENTO (RAG) ─────────────────────────────────────
            elif tool_name == 'search_knowledge':
                query = (args.get('query') or '').strip()
                if not query:
                    return json.dumps({'error': 'Se requiere una consulta'})
                results = env['estate.ai.knowledge'].sudo().search_knowledge(query, top_k=4)
                if not results:
                    return json.dumps({
                        'found': False,
                        'message': ('No hay documentación indexada o no se encontró nada relevante. '
                                    'Un administrador puede indexarla en Agente IA → Base de Conocimiento.'),
                    }, ensure_ascii=False)
                return json.dumps({
                    'found': True,
                    'passages': results,
                    'instruction': ('Responde al usuario usando estos extractos de la documentación. '
                                    'Cita la fuente cuando sea útil. Si no contienen la respuesta, dilo.'),
                }, ensure_ascii=False, default=str)

            # ── INFORME EJECUTIVO COMPLETO ─────────────────────────────────────
            elif tool_name == 'generate_executive_report':
                from datetime import date, timedelta
                today = date.today()
                month = int(args.get('month', 0) or today.month)
                year = int(args.get('year', 0) or today.year)
                month_start = date(year, month, 1)
                month_end = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)

                props = env['estate.property'].sudo().search([])
                available = props.filtered(lambda p: p.state == 'available')
                sold_all = props.filtered(lambda p: p.state == 'sold')
                rented_all = props.filtered(lambda p: p.state == 'rented')

                closed_month = env['estate.property'].sudo().search([
                    ('state', 'in', ['sold', 'rented']),
                    ('date_sold', '>=', str(month_start)),
                    ('date_sold', '<', str(month_end)),
                ])
                month_revenue = sum(p.price for p in closed_month if p.price)
                month_commission = sum(p.commission_amount for p in closed_month if p.commission_amount)

                leads = env['crm.lead'].sudo().search([('active', '=', True)])
                temp_count = {}
                for l in leads:
                    t = getattr(l, 'lead_temperature', 'cold') or 'cold'
                    temp_count[t] = temp_count.get(t, 0) + 1

                stale_threshold = today - timedelta(days=60)
                stale = [p for p in available if p.date_listed and p.date_listed <= stale_threshold]

                advisor_sales = {}
                for p in closed_month:
                    name = (p.user_id.name if p.user_id else None) or 'Sin asesor'
                    if name not in advisor_sales:
                        advisor_sales[name] = {'count': 0, 'revenue': 0, 'commissions': 0}
                    advisor_sales[name]['count'] += 1
                    advisor_sales[name]['revenue'] += p.price or 0
                    advisor_sales[name]['commissions'] += p.commission_amount or 0
                top_advisors = sorted(advisor_sales.items(), key=lambda x: x[1]['count'], reverse=True)[:5]

                try:
                    overdue_payments = env['estate.payment'].sudo().search_count([
                        ('state', '=', 'pending'), ('date', '<', today)])
                except Exception:
                    overdue_payments = 0

                return json.dumps({
                    'periodo': f"{month:02d}/{year}",
                    'inventario': {
                        'total': len(props),
                        'disponibles': len(available),
                        'vendidas_total': len(sold_all),
                        'arrendadas_total': len(rented_all),
                    },
                    'mes': {
                        'operaciones_cerradas': len(closed_month),
                        'ingresos_volumen_casas': round(month_revenue, 2),
                        'comisiones_agencia': round(month_commission, 2),
                        'detalle': [
                            {'nombre': p.name, 'precio_inmueble': p.price, 'comision_agencia': p.commission_amount or 0, 'tipo': 'Venta' if p.state == 'sold' else 'Arriendo',
                             'asesor': p.user_id.name if p.user_id else 'N/A'}
                            for p in closed_month
                        ],
                    },
                    'leads': {
                        'total_activos': len(leads),
                        'por_temperatura': temp_count,
                    },
                    'alertas': {
                        'propiedades_sin_movimiento_60d': len(stale),
                        'top_5_mas_tiempo': [
                            {'ref': p.name, 'dias': p.days_on_market or 0, 'precio': p.price, 'ciudad': p.city or ''}
                            for p in sorted(stale, key=lambda x: x.days_on_market or 0, reverse=True)[:5]
                        ],
                        'pagos_vencidos': overdue_payments,
                    },
                    'ranking_asesores': [
                        {'asesor': k, 'operaciones': v['count'], 'honorarios_agencia': round(v['commissions'], 2), 'volumen_inmuebles': round(v['revenue'], 2)}
                        for k, v in top_advisors
                    ],
                    'nota': f"Datos al {today.strftime('%d/%m/%Y')}",
                }, ensure_ascii=False)

            return json.dumps({'error': f'Herramienta desconocida: {tool_name}'})

        except Exception as e:
            _logger.error("Error ejecutando herramienta %s: %s", tool_name, str(e))
            return json.dumps({'error': str(e)})

    # -----------------------------------------------------------------------
    # Report Data Tool
    # -----------------------------------------------------------------------
    def _execute_report_data(self, args, env):
        """Return aggregated data for a given report_type so the AI can render charts/tables."""
        report_type = args.get('report_type', '')
        limit = int(args.get('limit', 8))

        try:
            if report_type == 'properties_by_state':
                states = [
                    ('available', 'Disponibles'),
                    ('reserved', 'Reservadas'),
                    ('sold', 'Vendidas'),
                    ('rented', 'Alquiladas'),
                ]
                data = {label: env['estate.property'].sudo().search_count([('state', '=', key)])
                        for key, label in states}
                return json.dumps({'report': 'Propiedades por Estado', 'data': data,
                                   'chart_hint': 'circular'}, ensure_ascii=False)

            elif report_type in ('properties_by_type', 'sales_by_type'):
                is_sold = (report_type == 'sales_by_type')
                env.cr.execute("""
                    SELECT pt.name as tipo,
                           COUNT(*) as unidades,
                           COALESCE(SUM(ep.commission_amount), 0) as comisiones,
                           COALESCE(SUM(ep.price), 0) as volumen,
                           COALESCE(AVG(ep.commission_amount), 0) as comision_promedio,
                           COALESCE(AVG(ep.price), 0) as precio_promedio
                    FROM estate_property ep
                    JOIN estate_property_type pt ON ep.property_type_id = pt.id
                    """ + ("WHERE ep.state = 'sold'" if is_sold else "") + """
                    GROUP BY pt.name
                    ORDER BY comisiones DESC, unidades DESC
                    LIMIT %s
                """, (limit,))
                rows = env.cr.dictfetchall()
                if not rows and not is_sold:
                    types = env['estate.property.type'].sudo().search([], limit=limit)
                    data = {}
                    for t in types:
                        cnt = env['estate.property'].sudo().search_count([('property_type_id', '=', t.id)])
                        if cnt:
                            data[t.name] = cnt
                    return json.dumps({'report': 'Propiedades por Tipo', 'data': data,
                                       'chart_hint': 'barra'}, ensure_ascii=False)
                if not rows and is_sold:
                    return json.dumps({'report': 'Ventas por Tipo de Propiedad', 'data': {},
                                       'mensaje': 'Sin ventas registradas aún.'}, ensure_ascii=False)
                data = {r['tipo']: float(f"{float(r['comisiones']):.2f}") for r in rows}
                data_unidades = {r['tipo']: int(r['unidades']) for r in rows}
                data_volumen = {r['tipo']: float(f"{float(r['volumen']):.2f}") for r in rows}
                detalle_clean = [
                    {
                        'tipo_propiedad': r['tipo'],
                        'operaciones': int(r['unidades']),
                        'ventas_agencia_comisiones_totales': float(f"{float(r['comisiones']):.2f}"),
                        'ventas_agencia_comision_promedio': float(f"{float(r['comision_promedio']):.2f}"),
                        'volumen_inmuebles_precio_total': float(f"{float(r['volumen']):.2f}"),
                        'volumen_inmuebles_precio_promedio': float(f"{float(r['precio_promedio']):.2f}")
                    }
                    for r in rows
                ]
                titulo = 'Ventas y Honorarios por Tipo de Propiedad' if is_sold else 'Propiedades por Tipo (Comisiones vs Volumen)'
                return json.dumps({'report': titulo, 'data': data,
                                   'data_unidades': data_unidades, 'data_volumen': data_volumen,
                                   'chart_hint': 'barra', 'detalle': detalle_clean}, ensure_ascii=False, default=str)

            elif report_type == 'properties_by_prospects':
                # Propiedades ordenadas por nº de prospectos (leads que la tienen
                # como target_property_id). Más prospectos = más cerca de vender.
                env.cr.execute("""
                    SELECT ep.title AS propiedad, COUNT(cl.id) AS prospectos
                    FROM estate_property ep
                    LEFT JOIN crm_lead cl
                        ON cl.target_property_id = ep.id
                        AND cl.type = 'opportunity' AND cl.active = TRUE
                    WHERE ep.state IN ('available', 'reserved')
                    GROUP BY ep.id, ep.title
                    ORDER BY prospectos DESC, ep.title
                    LIMIT %s
                """, (limit,))
                rows = env.cr.dictfetchall()
                data = {r['propiedad'][:28]: int(r['prospectos']) for r in rows}
                return json.dumps({'report': 'Propiedades por Prospectos', 'data': data,
                                   'chart_hint': 'barra'}, ensure_ascii=False)

            elif report_type == 'sales_by_month':
                env.cr.execute("""
                    SELECT TO_CHAR(date_sold, 'Mon YYYY') as mes,
                           COUNT(*) as ventas,
                           COALESCE(SUM(price), 0) as ingresos,
                           COALESCE(SUM(commission_amount), 0) as comisiones
                    FROM estate_property
                    WHERE state = 'sold' AND date_sold IS NOT NULL
                    GROUP BY TO_CHAR(date_sold, 'Mon YYYY'), DATE_TRUNC('month', date_sold)
                    ORDER BY DATE_TRUNC('month', date_sold) DESC
                    LIMIT %s
                """, (limit,))
                rows = env.cr.dictfetchall()
                data = {r['mes']: float(f"{float(r['comisiones']):.2f}") for r in reversed(rows)}
                data_volumen = {r['mes']: float(f"{float(r['ingresos']):.2f}") for r in reversed(rows)}
                data_unidades = {r['mes']: int(r['ventas']) for r in reversed(rows)}
                detalle_clean = [
                    {
                        'mes': r['mes'],
                        'unidades_vendidas': int(r['ventas']),
                        'honorarios_agencia_comision': float(f"{float(r['comisiones']):.2f}"),
                        'valor_total_propiedades_volumen': float(f"{float(r['ingresos']):.2f}")
                    }
                    for r in rows
                ]
                return json.dumps({'report': 'Ventas por Mes (Honorarios vs Volumen)', 'data': data,
                                   'data_volumen': data_volumen, 'data_unidades': data_unidades,
                                   'chart_hint': 'linea',
                                   'detalle': detalle_clean}, ensure_ascii=False, default=str)

            elif report_type == 'visits_by_property':
                env.cr.execute("""
                    SELECT ep.title as propiedad, COUNT(ce.id) as visitas
                    FROM calendar_event ce
                    JOIN estate_property ep ON ce.property_id = ep.id
                    WHERE ce.property_id IS NOT NULL
                    GROUP BY ep.title
                    ORDER BY visitas DESC
                    LIMIT %s
                """, (limit,))
                rows = env.cr.dictfetchall()
                data = {r['propiedad']: int(r['visitas']) for r in rows}
                return json.dumps({'report': 'Visitas por Propiedad', 'data': data,
                                   'chart_hint': 'barra'}, ensure_ascii=False)

            elif report_type == 'commissions_by_advisor':
                from datetime import date
                start_month = date.today().replace(day=1)
                sold = env['estate.property'].sudo().search([
                    ('state', '=', 'sold'),
                    ('date_sold', '>=', start_month),
                    ('user_id', '!=', False),
                ])
                data = {}
                for p in sold:
                    name = p.user_id.name
                    data[name] = round(data.get(name, 0) + (p.commission_amount or 0), 2)
                data = {k: v for i, (k, v) in enumerate(
                    sorted(data.items(), key=lambda x: x[1], reverse=True)
                ) if i < limit}
                return json.dumps({'report': 'Comisiones por Asesor (Mes)', 'data': data,
                                   'chart_hint': 'barra'}, ensure_ascii=False)

            elif report_type == 'contracts_by_type':
                types = [('sale', 'Venta'), ('rent', 'Alquiler'), ('exclusivity', 'Exclusividad')]
                data = {label: env['estate.contract'].sudo().search_count(
                    [('contract_type', '=', key), ('state', '=', 'active')])
                        for key, label in types}
                return json.dumps({'report': 'Contratos Activos por Tipo', 'data': data,
                                   'chart_hint': 'circular'}, ensure_ascii=False)

            elif report_type == 'expenses_by_type':
                env.cr.execute("""
                    SELECT expense_type, COALESCE(SUM(amount), 0) as total
                    FROM estate_property_expense
                    WHERE state != 'cancelled'
                    GROUP BY expense_type
                    ORDER BY total DESC
                    LIMIT %s
                """, (limit,))
                rows = env.cr.dictfetchall()
                data = {r['expense_type']: float(f"{float(r['total']):.2f}") for r in rows}
                return json.dumps({'report': 'Gastos por Tipo', 'data': data,
                                   'chart_hint': 'circular'}, ensure_ascii=False)

            elif report_type == 'offers_by_state':
                states = [
                    ('draft', 'Borrador'), ('submitted', 'Presentada'),
                    ('countered', 'Contraoferta'), ('accepted', 'Aceptada'),
                    ('rejected', 'Rechazada'), ('expired', 'Expirada'),
                ]
                data = {label: env['estate.property.offer'].sudo().search_count([('state', '=', key)])
                        for key, label in states}
                data = {k: v for k, v in data.items() if v > 0}
                return json.dumps({'report': 'Ofertas por Estado', 'data': data,
                                   'chart_hint': 'barra'}, ensure_ascii=False)

            elif report_type == 'leads_by_temperature':
                temps = [('cold', 'Frío'), ('warm', 'Tibio'), ('hot', 'Caliente'), ('boiling', 'Hirviendo')]
                data = {label: env['crm.lead'].sudo().search_count(
                    [('lead_temperature', '=', key), ('type', '=', 'lead')])
                        for key, label in temps}
                return json.dumps({'report': 'Leads por Temperatura', 'data': data,
                                   'chart_hint': 'barra'}, ensure_ascii=False)

            elif report_type == 'payments_by_method':
                env.cr.execute("""
                    SELECT payment_method, COALESCE(SUM(amount), 0) as total
                    FROM estate_payment
                    WHERE state = 'paid'
                    GROUP BY payment_method
                    ORDER BY total DESC
                    LIMIT %s
                """, (limit,))
                rows = env.cr.dictfetchall()
                data = {r['payment_method']: float(f"{float(r['total']):.2f}") for r in rows}
                return json.dumps({'report': 'Pagos por Método', 'data': data,
                                   'chart_hint': 'circular'}, ensure_ascii=False)

            elif report_type == 'days_on_market_by_type':
                env.cr.execute("""
                    SELECT pt.name as tipo, ROUND(AVG(ep.days_on_market)::numeric, 1) as promedio
                    FROM estate_property ep
                    JOIN estate_property_type pt ON ep.property_type_id = pt.id
                    WHERE ep.state = 'sold' AND ep.days_on_market > 0
                    GROUP BY pt.name
                    ORDER BY promedio DESC
                    LIMIT %s
                """, (limit,))
                rows = env.cr.dictfetchall()
                data = {r['tipo']: float(r['promedio']) for r in rows}
                return json.dumps({'report': 'Días Promedio en Mercado por Tipo', 'data': data,
                                   'chart_hint': 'barra'}, ensure_ascii=False)

            elif report_type == 'time_to_sell_summary':
                env.cr.execute("""
                    SELECT COUNT(*) as ventas,
                           ROUND(AVG(days_on_market)::numeric, 1) as promedio,
                           MIN(days_on_market) as minimo,
                           MAX(days_on_market) as maximo,
                           ROUND(PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY days_on_market)::numeric, 1) as mediana,
                           COALESCE(AVG(commission_amount), 0) as comision_promedio,
                           COALESCE(AVG(price), 0) as precio_promedio
                    FROM estate_property
                    WHERE state = 'sold' AND date_listed IS NOT NULL AND days_on_market > 0
                """)
                summary_row = env.cr.dictfetchone()
                if not summary_row or not summary_row['ventas']:
                    return json.dumps({'report': 'Tiempo de Venta', 'data': {},
                                       'mensaje': 'Sin ventas con fecha de publicación registrada aún.'})
                env.cr.execute("""
                    SELECT pt.name as tipo,
                           ROUND(AVG(ep.days_on_market)::numeric, 1) as promedio_dias,
                           COALESCE(AVG(ep.commission_amount), 0) as comision_promedio,
                           COALESCE(AVG(ep.price), 0) as precio_promedio
                    FROM estate_property ep
                    JOIN estate_property_type pt ON ep.property_type_id = pt.id
                    WHERE ep.state = 'sold' AND ep.date_listed IS NOT NULL AND ep.days_on_market > 0
                    GROUP BY pt.name
                    ORDER BY promedio_dias DESC
                    LIMIT %s
                """, (limit,))
                by_type = env.cr.dictfetchall()
                data = {r['tipo']: float(r['promedio_dias']) for r in by_type}
                return json.dumps({
                    'report': 'Tiempo de Venta y Promedios de Operación por Tipo',
                    'data': data,
                    'chart_hint': 'barra',
                    'resumen': {
                        'operaciones_analizadas': int(summary_row['ventas']),
                        'comision_promedio_agencia': float(f"{float(summary_row['comision_promedio']):.2f}"),
                        'precio_promedio_inmuebles': float(f"{float(summary_row['precio_promedio']):.2f}"),
                        'promedio_dias_mercado': float(summary_row['promedio']),
                        'mediana_dias_mercado': float(summary_row['mediana']),
                        'minimo_dias': int(summary_row['minimo']),
                        'maximo_dias': int(summary_row['maximo']),
                    },
                    'detalle': [
                        {
                            'tipo_propiedad': r['tipo'],
                            'dias_promedio_mercado': float(r['promedio_dias']),
                            'comision_promedio_agencia': float(f"{float(r['comision_promedio']):.2f}"),
                            'precio_promedio_inmueble': float(f"{float(r['precio_promedio']):.2f}")
                        }
                        for r in by_type
                    ]
                }, ensure_ascii=False)

            elif report_type == 'sales_by_channel':
                env.cr.execute("""
                    SELECT
                        CASE sold_by
                            WHEN 'agency' THEN 'Agencia'
                            WHEN 'owner' THEN 'Propietario'
                            WHEN 'external' THEN 'Externo'
                            ELSE 'Sin especificar'
                        END as canal,
                        COUNT(*) as ventas,
                        COALESCE(SUM(price), 0) as ingresos,
                        COALESCE(SUM(commission_amount), 0) as comisiones
                    FROM estate_property
                    WHERE state = 'sold'
                    GROUP BY sold_by
                    ORDER BY ventas DESC
                """)
                rows = env.cr.dictfetchall()
                if not rows:
                    return json.dumps({'report': 'Ventas por Canal', 'data': {},
                                       'mensaje': 'Sin ventas registradas aún.'})
                data = {r['canal']: int(r['ventas']) for r in rows}
                return json.dumps({
                    'report': 'Ventas por Canal (Agencia vs Propietario)',
                    'data': data,
                    'chart_hint': 'circular',
                    'detalle': rows,
                }, ensure_ascii=False, default=str)

            elif report_type == 'ranking_advisors':
                from datetime import date as _d
                start = _d.today().replace(day=1)
                env.cr.execute("""
                    SELECT rp.name as asesor,
                           COUNT(ep.id) as ventas,
                           COALESCE(SUM(ep.price), 0) as ingresos,
                           COALESCE(SUM(ep.commission_amount), 0) as comisiones
                    FROM estate_property ep
                    JOIN res_users ru ON ep.user_id = ru.id
                    JOIN res_partner rp ON ru.partner_id = rp.id
                    WHERE ep.state = 'sold' AND ep.date_sold >= %s
                    GROUP BY rp.name
                    ORDER BY comisiones DESC, ventas DESC
                    LIMIT %s
                """, (start, limit))
                rows = env.cr.dictfetchall()
                data = {r['asesor']: float(f"{float(r['comisiones']):.2f}") for r in rows}
                data_volumen = {r['asesor']: float(f"{float(r['ingresos']):.2f}") for r in rows}
                data_unidades = {r['asesor']: int(r['ventas']) for r in rows}
                detalle_clean = [
                    {
                        'asesor': r['asesor'],
                        'operaciones_cerradas': int(r['ventas']),
                        'honorarios_agencia_comision': float(f"{float(r['comisiones']):.2f}"),
                        'valor_propiedades_volumen': float(f"{float(r['ingresos']):.2f}")
                    }
                    for r in rows
                ]
                return json.dumps({'report': 'Ranking Asesores (Mes - Honorarios vs Volumen)', 'data': data,
                                   'data_volumen': data_volumen, 'data_unidades': data_unidades,
                                   'chart_hint': 'barra', 'detalle': detalle_clean},
                                  ensure_ascii=False, default=str)

            elif report_type == 'kpi_general':
                from datetime import date as _d
                today = _d.today()
                start_m = today.replace(day=1)
                props = env['estate.property'].sudo()
                leads = env['crm.lead'].sudo()
                total_props    = props.search_count([])
                avail_props    = props.search_count([('state', '=', 'available')])
                sold_this_month= props.search_count([('state', '=', 'sold'), ('date_sold', '>=', start_m)])
                env.cr.execute("SELECT COALESCE(SUM(price),0) FROM estate_property WHERE state='sold' AND date_sold >= %s", (start_m,))
                ingresos_mes = float(env.cr.fetchone()[0] or 0)
                env.cr.execute("SELECT COALESCE(SUM(commission_amount),0) FROM estate_property WHERE state='sold' AND date_sold >= %s", (start_m,))
                comisiones_mes = float(env.cr.fetchone()[0] or 0)
                env.cr.execute("SELECT COALESCE(ROUND(AVG(days_on_market)::numeric,1),0) FROM estate_property WHERE state='sold' AND days_on_market > 0")
                avg_days = float(env.cr.fetchone()[0] or 0)
                active_leads   = leads.search_count([('type', '=', 'opportunity'), ('stage_id.is_won', '=', False)])
                hot_leads      = leads.search_count([('lead_temperature', 'in', ['hot', 'boiling']), ('type', '=', 'opportunity')])
                data = {
                    'Propiedades Totales': total_props,
                    'Disponibles': avail_props,
                    'Propiedades Vendidas (Mes)': sold_this_month,
                    'Honorarios de Agencia (Comisiones Mes $)': round(comisiones_mes, 2),
                    'Volumen Inmuebles Vendidos (Mes $)': round(ingresos_mes, 2),
                    'Días promedio en mercado': avg_days,
                    'Leads activos': active_leads,
                    'Leads calientes': hot_leads,
                }
                return json.dumps({'report': 'KPIs Generales (Honorarios vs Volumen)', 'data': data,
                                   'chart_hint': 'barra',
                                   'kpis': {
                                       'ventas_mes': sold_this_month,
                                       'ingresos_mes': round(ingresos_mes, 2),
                                       'comisiones_mes': round(comisiones_mes, 2),
                                       'dias_promedio': avg_days,
                                       'leads_activos': active_leads,
                                       'leads_calientes': hot_leads,
                                   }}, ensure_ascii=False, default=str)

            elif report_type == 'appraisals_by_state':
                states = [
                    ('scheduled', 'Programada'), ('in_progress', 'En Proceso'),
                    ('completed', 'Completada'), ('cancelled', 'Cancelada'),
                ]
                data = {label: env['estate.appraisal'].sudo().search_count([('state', '=', key)])
                        for key, label in states}
                data = {k: v for k, v in data.items() if v > 0}
                return json.dumps({'report': 'Tasaciones por Estado', 'data': data,
                                   'chart_hint': 'circular'}, ensure_ascii=False)

            elif report_type == 'maintenance_by_state':
                states = [
                    ('pending', 'Pendiente'), ('in_progress', 'En Proceso'),
                    ('resolved', 'Resuelto'), ('cancelled', 'Cancelado'),
                ]
                req_types = [('repair', 'Reparación'), ('cleaning', 'Limpieza'),
                             ('inspection', 'Inspección'), ('emergency', 'Emergencia'), ('other', 'Otro')]
                data_state = {label: env['estate.tenant.request'].sudo().search_count([('state', '=', key)])
                              for key, label in states}
                data = {k: v for k, v in data_state.items() if v > 0}
                return json.dumps({'report': 'Mantenimiento por Estado', 'data': data,
                                   'chart_hint': 'barra'}, ensure_ascii=False)

            elif report_type == 'leads_by_source':
                env.cr.execute("""
                    SELECT COALESCE(us.name, 'Sin fuente') as fuente, COUNT(cl.id) as leads
                    FROM crm_lead cl
                    LEFT JOIN utm_source us ON cl.source_id = us.id
                    WHERE cl.type = 'opportunity'
                    GROUP BY us.name
                    ORDER BY leads DESC
                    LIMIT %s
                """, (limit,))
                rows = env.cr.dictfetchall()
                data = {r['fuente']: int(r['leads']) for r in rows}
                return json.dumps({'report': 'Leads por Fuente de Captación', 'data': data,
                                   'chart_hint': 'circular'}, ensure_ascii=False)

            elif report_type == 'leads_by_stage':
                env.cr.execute("""
                    SELECT cs.name as etapa, COUNT(cl.id) as leads
                    FROM crm_lead cl
                    JOIN crm_stage cs ON cl.stage_id = cs.id
                    WHERE cl.type = 'opportunity' AND cs.is_won = false
                    GROUP BY cs.name, cs.sequence
                    ORDER BY cs.sequence
                    LIMIT %s
                """, (limit,))
                rows = env.cr.dictfetchall()
                data = {r['etapa']: int(r['leads']) for r in rows}
                return json.dumps({'report': 'Pipeline de Clientes por Etapa', 'data': data,
                                   'chart_hint': 'barra'}, ensure_ascii=False)

            elif report_type == 'deals_closed_by_month':
                env.cr.execute("""
                    SELECT TO_CHAR(cl.date_closed, 'Mon YYYY') as mes,
                           COUNT(*) as negocios,
                           COALESCE(SUM(cl.prorated_revenue), 0) as ingresos
                    FROM crm_lead cl
                    WHERE cl.type = 'opportunity'
                      AND cl.stage_id IN (SELECT id FROM crm_stage WHERE is_won = true)
                      AND cl.date_closed IS NOT NULL
                    GROUP BY TO_CHAR(cl.date_closed, 'Mon YYYY'), DATE_TRUNC('month', cl.date_closed)
                    ORDER BY DATE_TRUNC('month', cl.date_closed) DESC
                    LIMIT %s
                """, (limit,))
                rows = env.cr.dictfetchall()
                data = {r['mes']: int(r['negocios']) for r in reversed(rows)}
                return json.dumps({'report': 'Negocios Realizados por Mes', 'data': data,
                                   'chart_hint': 'linea', 'detalle': rows},
                                  ensure_ascii=False, default=str)

            elif report_type == 'income_by_month':
                env.cr.execute("""
                    SELECT TO_CHAR(payment_date, 'Mon YYYY') as mes,
                           COALESCE(SUM(amount), 0) as ingresos
                    FROM estate_payment
                    WHERE state = 'paid' AND payment_date IS NOT NULL
                    GROUP BY TO_CHAR(payment_date, 'Mon YYYY'), DATE_TRUNC('month', payment_date)
                    ORDER BY DATE_TRUNC('month', payment_date) DESC
                    LIMIT %s
                """, (limit,))
                rows = env.cr.dictfetchall()
                data = {r['mes']: float(f"{float(r['ingresos']):.2f}") for r in reversed(rows)}
                return json.dumps({'report': 'Ingresos por Mes (Pagos)', 'data': data,
                                   'chart_hint': 'linea'}, ensure_ascii=False, default=str)

            elif report_type == 'commissions_pending':
                env.cr.execute("""
                    SELECT rp.name as asesor,
                           COUNT(ec.id) as comisiones,
                           COALESCE(SUM(ec.amount), 0) as total
                    FROM estate_commission ec
                    JOIN res_users ru ON ec.user_id = ru.id
                    JOIN res_partner rp ON ru.partner_id = rp.id
                    WHERE ec.state IN ('draft', 'approved')
                    GROUP BY rp.name
                    ORDER BY total DESC
                    LIMIT %s
                """, (limit,))
                rows = env.cr.dictfetchall()
                data = {r['asesor']: float(f"{float(r['total']):.2f}") for r in rows}
                return json.dumps({'report': 'Comisiones Pendientes por Asesor', 'data': data,
                                   'chart_hint': 'barra', 'detalle': rows},
                                  ensure_ascii=False, default=str)

            elif report_type == 'visits_done_summary':
                env.cr.execute("""
                    SELECT TO_CHAR(ce.start, 'Mon YYYY') as mes, COUNT(*) as visitas
                    FROM calendar_event ce
                    WHERE ce.property_id IS NOT NULL
                      AND (ce.visit_state IS NULL OR ce.visit_state = 'done')
                    GROUP BY TO_CHAR(ce.start, 'Mon YYYY'), DATE_TRUNC('month', ce.start)
                    ORDER BY DATE_TRUNC('month', ce.start) DESC
                    LIMIT %s
                """, (limit,))
                rows = env.cr.dictfetchall()
                data = {r['mes']: int(r['visitas']) for r in reversed(rows)}
                return json.dumps({'report': 'Visitas Realizadas por Mes', 'data': data,
                                   'chart_hint': 'barra'}, ensure_ascii=False)

            elif report_type == 'social_facebook':
                try:
                    env.cr.execute("""
                        SELECT ep.title as propiedad,
                               COALESCE(ef.reach, 0) as alcance,
                               COALESCE(ef.impressions, 0) as impresiones,
                               COALESCE(ef.engagement_rate, 0) as engagement
                        FROM estate_facebook_stats ef
                        JOIN estate_property ep ON ef.property_id = ep.id
                        ORDER BY ef.reach DESC
                        LIMIT %s
                    """, (limit,))
                    rows = env.cr.dictfetchall()
                    if not rows:
                        return json.dumps({'report': 'Estadísticas Facebook', 'data': {},
                                           'mensaje': 'No hay estadísticas de Facebook registradas aún.'})
                    data = {r['propiedad']: int(r['alcance']) for r in rows}
                    return json.dumps({'report': 'Alcance Facebook por Propiedad', 'data': data,
                                       'chart_hint': 'barra', 'detalle': rows},
                                      ensure_ascii=False, default=str)
                except Exception:
                    return json.dumps({'error': 'Módulo de Facebook no disponible o sin datos.'})

            elif report_type == 'sales_avg_summary':
                env.cr.execute("""
                    SELECT
                        TO_CHAR(date_sold, 'Mon YYYY') as mes,
                        DATE_TRUNC('month', date_sold) as mes_ord,
                        COUNT(*) as unidades,
                        COALESCE(SUM(price), 0) as ingresos_total,
                        COALESCE(AVG(price), 0) as precio_promedio,
                        COALESCE(SUM(commission_amount), 0) as comisiones_total,
                        COALESCE(AVG(commission_amount), 0) as comision_promedio,
                        COALESCE(AVG(days_on_market), 0) as dias_promedio
                    FROM estate_property
                    WHERE state = 'sold' AND date_sold IS NOT NULL
                    GROUP BY TO_CHAR(date_sold, 'Mon YYYY'), DATE_TRUNC('month', date_sold)
                    ORDER BY DATE_TRUNC('month', date_sold) DESC
                    LIMIT %s
                """, (limit,))
                rows = env.cr.dictfetchall()
                if not rows:
                    return json.dumps({'report': 'Promedio de Ventas', 'data': {},
                                       'mensaje': 'Sin ventas registradas aún.'})
                # Summary metrics
                total_units = sum(r['unidades'] for r in rows)
                total_revenue = sum(float(r['ingresos_total']) for r in rows)
                total_commissions = sum(float(r['comisiones_total']) for r in rows)
                n_months = len(rows)
                avg_units_month = round(total_units / n_months, 1)
                avg_revenue_month = round(total_revenue / n_months, 2)
                avg_comm_month = round(total_commissions / n_months, 2)
                avg_price_unit = round(total_revenue / total_units, 2) if total_units else 0
                avg_commission_unit = round(total_commissions / total_units, 2) if total_units else 0
                # Chart: comisiones y volumen por mes
                data_comisiones = {r['mes']: float(f"{float(r['comisiones_total']):.2f}") for r in reversed(rows)}
                data_revenue = {r['mes']: float(f"{float(r['ingresos_total']):.2f}") for r in reversed(rows)}
                data_units = {r['mes']: int(r['unidades']) for r in reversed(rows)}
                detalle_clean = [
                    {
                        'mes': r['mes'],
                        'unidades_vendidas': int(r['unidades']),
                        'honorarios_agencia_total': float(f"{float(r['comisiones_total']):.2f}"),
                        'honorarios_agencia_promedio': float(f"{float(r['comision_promedio']):.2f}"),
                        'valor_inmuebles_total': float(f"{float(r['ingresos_total']):.2f}"),
                        'valor_inmuebles_promedio': float(f"{float(r['precio_promedio']):.2f}")
                    }
                    for r in rows
                ]
                return json.dumps({
                    'report': 'Promedio de Ventas (Honorarios vs Volumen de Casas)',
                    'data': data_comisiones,
                    'data_volumen': data_revenue,
                    'data_units': data_units,
                    'chart_hint': 'linea',
                    'resumen': {
                        'meses_analizados': n_months,
                        'unidades_promedio_mes': avg_units_month,
                        'honorarios_agencia_promedio_mes': avg_comm_month,
                        'honorarios_agencia_promedio_por_venta': avg_commission_unit,
                        'volumen_casas_promedio_mes': avg_revenue_month,
                        'precio_promedio_por_inmueble': avg_price_unit,
                        'honorarios_agencia_totales': round(total_commissions, 2),
                        'volumen_casas_totales': round(total_revenue, 2),
                    },
                    'detalle': detalle_clean,
                }, ensure_ascii=False, default=str)

            elif report_type == 'social_instagram':
                try:
                    env.cr.execute("""
                        SELECT ep.title as propiedad,
                               COALESCE(ei.reach, 0) as alcance,
                               COALESCE(ei.impressions, 0) as impresiones,
                               COALESCE(ei.engagement_rate, 0) as engagement
                        FROM estate_instagram_stats ei
                        JOIN estate_property ep ON ei.property_id = ep.id
                        ORDER BY ei.reach DESC
                        LIMIT %s
                    """, (limit,))
                    rows = env.cr.dictfetchall()
                    if not rows:
                        return json.dumps({'report': 'Estadísticas Instagram', 'data': {},
                                           'mensaje': 'No hay estadísticas de Instagram registradas aún.'})
                    data = {r['propiedad']: int(r['alcance']) for r in rows}
                    return json.dumps({'report': 'Alcance Instagram por Propiedad', 'data': data,
                                       'chart_hint': 'barra', 'detalle': rows},
                                      ensure_ascii=False, default=str)
                except Exception:
                    return json.dumps({'error': 'Módulo de Instagram no disponible o sin datos.'})

            # ── POSTS PERSONALES FB DE ASESORES ───────────────────────────────
            elif report_type == 'advisor_fb_posts':
                try:
                    # Publicaciones por asesor (total FB + IG)
                    env.cr.execute("""
                        SELECT rp.name as asesor,
                               COUNT(fp.id) as total,
                               SUM(CASE WHEN fp.platform = 'facebook' THEN 1 ELSE 0 END) as facebook,
                               SUM(CASE WHEN fp.platform = 'instagram' THEN 1 ELSE 0 END) as instagram,
                               MAX(fp.published_date) as ultima
                        FROM estate_advisor_fb_post fp
                        JOIN res_users ru ON fp.user_id = ru.id
                        JOIN res_partner rp ON ru.partner_id = rp.id
                        GROUP BY rp.name
                        ORDER BY total DESC
                        LIMIT %s
                    """, (limit,))
                    rows_asesor = env.cr.dictfetchall()

                    # Propiedades más publicadas
                    env.cr.execute("""
                        SELECT COALESCE(ep.title, ep.name) as propiedad,
                               COUNT(fp.id) as total,
                               SUM(CASE WHEN fp.platform = 'facebook' THEN 1 ELSE 0 END) as facebook,
                               SUM(CASE WHEN fp.platform = 'instagram' THEN 1 ELSE 0 END) as instagram
                        FROM estate_advisor_fb_post fp
                        JOIN estate_property ep ON fp.property_id = ep.id
                        GROUP BY ep.id, ep.title, ep.name
                        ORDER BY total DESC
                        LIMIT 10
                    """)
                    rows_prop = env.cr.dictfetchall()

                    # Totales por plataforma
                    env.cr.execute("""
                        SELECT platform,
                               COUNT(*) as total
                        FROM estate_advisor_fb_post
                        GROUP BY platform
                    """)
                    rows_platform = env.cr.dictfetchall()

                    if not rows_asesor:
                        return json.dumps({'report': 'Posts FB/IG de Asesores', 'data': {},
                                           'mensaje': 'Aún no hay publicaciones personales registradas.'})

                    data_asesores = {r['asesor']: int(r['total']) for r in rows_asesor}
                    data_propiedades = {r['propiedad']: int(r['total']) for r in rows_prop}
                    data_plataforma = {r['platform'].capitalize(): int(r['total']) for r in rows_platform}

                    return json.dumps({
                        'report': 'Publicaciones Personales FB/Instagram por Asesor',
                        'data': data_asesores,
                        'chart_hint': 'barra',
                        'detalle_asesores': rows_asesor,
                        'detalle_propiedades': rows_prop,
                        'detalle_plataforma': data_plataforma,
                        'extra_charts': [
                            {'titulo': 'Por Plataforma', 'data': data_plataforma, 'tipo': 'circular'},
                            {'titulo': 'Propiedades más publicadas', 'data': data_propiedades, 'tipo': 'barra'},
                        ],
                    }, ensure_ascii=False, default=str)
                except Exception as e:
                    return json.dumps({'error': f'Error obteniendo datos: {str(e)}'})

            # ── PRECIO VS AVM (SOBRE/SUBVALORADAS) ────────────────────────────
            elif report_type == 'price_vs_avm':
                try:
                    env.cr.execute("""
                        SELECT ep.title,
                               ep.price,
                               ep.avm_estimated_price,
                               ep.avm_status,
                               ROUND(((ep.price - ep.avm_estimated_price) / NULLIF(ep.avm_estimated_price,0)) * 100, 1) as diferencia_pct,
                               ep.city,
                               ep.state,
                               ep.id
                        FROM estate_property ep
                        WHERE ep.active = TRUE
                          AND ep.avm_estimated_price > 0
                          AND ep.state IN ('available','reserved')
                        ORDER BY ABS(ep.price - ep.avm_estimated_price) DESC
                        LIMIT %s
                    """, (limit,))
                    rows = env.cr.dictfetchall()
                    if not rows:
                        return json.dumps({'report': 'Precio vs AVM', 'data': {},
                                           'mensaje': 'No hay propiedades con AVM calculado.'})
                    overvalued = sum(1 for r in rows if float(r.get('diferencia_pct') or 0) > 5)
                    undervalued = sum(1 for r in rows if float(r.get('diferencia_pct') or 0) < -5)
                    fair = len(rows) - overvalued - undervalued
                    data = {'Sobrevaluadas (>5% sobre AVM)': overvalued,
                            'Precio justo (±5% del AVM)': fair,
                            'Subvaluadas (>5% bajo AVM)': undervalued}
                    return json.dumps({'report': 'Precio vs Valor de Mercado (AVM)',
                                       'data': data, 'chart_hint': 'circular', 'detalle': rows},
                                      ensure_ascii=False, default=str)
                except Exception as e:
                    return json.dumps({'error': f'Error al calcular precio vs AVM: {str(e)}'})

            # ── PROPIEDADES SIN VISITAS ────────────────────────────────────────
            elif report_type == 'properties_no_visits':
                try:
                    from datetime import date as _date, timedelta as _td
                    cutoff = _date.today() - _td(days=30)
                    env.cr.execute("""
                        SELECT ep.id, ep.title, ep.city, ep.price, ep.days_on_market,
                               ep.state, ept.name as tipo,
                               rp.name as asesor
                        FROM estate_property ep
                        LEFT JOIN estate_property_type ept ON ep.property_type_id = ept.id
                        LEFT JOIN res_users ru ON ep.user_id = ru.id
                        LEFT JOIN res_partner rp ON ru.partner_id = rp.id
                        WHERE ep.active = TRUE
                          AND ep.state = 'available'
                          AND ep.id NOT IN (
                              SELECT DISTINCT property_id FROM calendar_event
                              WHERE property_id IS NOT NULL
                                AND start >= %s
                          )
                        ORDER BY ep.days_on_market DESC
                        LIMIT %s
                    """, (cutoff, limit))
                    rows = env.cr.dictfetchall()
                    if not rows:
                        return json.dumps({'report': 'Propiedades sin visitas', 'data': {},
                                           'mensaje': 'Todas las propiedades disponibles tienen visitas recientes.'})
                    data = {r['title'] or f"Prop #{r['id']}": int(r.get('days_on_market') or 0)
                            for r in rows[:8]}
                    return json.dumps({'report': 'Propiedades disponibles sin visitas (últimos 30 días)',
                                       'data': data, 'chart_hint': 'barra', 'detalle': rows,
                                       'total': len(rows)},
                                      ensure_ascii=False, default=str)
                except Exception as e:
                    return json.dumps({'error': f'Error: {str(e)}'})

            # ── EMBUDO DE CONVERSIÓN ───────────────────────────────────────────
            elif report_type == 'conversion_funnel':
                try:
                    total_leads = env['crm.lead'].sudo().search_count([('type', '=', 'opportunity')])
                    total_visits = 0
                    if 'property_id' in env['calendar.event']._fields:
                        total_visits = env['calendar.event'].sudo().search_count(
                            [('property_id', '!=', False)])
                    total_offers = env['estate.property.offer'].sudo().search_count([])
                    total_contracts = env['estate.contract'].sudo().search_count(
                        [('state', 'in', ('active', 'expired'))])
                    total_closed = env['estate.property'].sudo().search_count(
                        [('state', 'in', ('sold', 'rented'))])

                    data = {
                        'Leads/Oportunidades': total_leads,
                        'Visitas realizadas': total_visits,
                        'Ofertas presentadas': total_offers,
                        'Contratos firmados': total_contracts,
                        'Propiedades cerradas': total_closed,
                    }
                    conv_rate = round((total_closed / total_leads * 100), 1) if total_leads > 0 else 0
                    return json.dumps({'report': 'Embudo de Conversión',
                                       'data': data, 'chart_hint': 'barra',
                                       'tasa_conversion': conv_rate,
                                       'mensaje': f'Tasa de conversión global: {conv_rate}%'},
                                      ensure_ascii=False)
                except Exception as e:
                    return json.dumps({'error': f'Error al calcular embudo: {str(e)}'})

            # ── ESTADO SINCRONIZACIÓN WORDPRESS ───────────────────────────────
            elif report_type == 'wp_sync_status':
                try:
                    has_wp = 'wp_published' in env['estate.property']._fields
                    if not has_wp:
                        return json.dumps({'error': 'Módulo WordPress no instalado.'})
                    total = env['estate.property'].sudo().search_count([('active', '=', True), ('state', '=', 'available')])
                    published = env['estate.property'].sudo().search_count([('active', '=', True), ('wp_published', '=', True)])
                    unlinked = env['estate.property'].sudo().search_count([('active', '=', True), ('wp_published', '=', False), ('state', '=', 'available')])
                    data = {
                        'Publicadas en WordPress': published,
                        'Sin publicar (disponibles)': unlinked,
                    }
                    return json.dumps({'report': 'Estado de Publicación en WordPress',
                                       'data': data, 'chart_hint': 'circular',
                                       'total_activas': total},
                                      ensure_ascii=False)
                except Exception as e:
                    return json.dumps({'error': f'Error: {str(e)}'})

            # ── RANKING DE CONTACTOS / CLIENTES ───────────────────────────────
            elif report_type == 'contact_ranking':
                try:
                    env.cr.execute("""
                        SELECT rp.name as cliente,
                               COUNT(DISTINCT cl.id) as leads_activos,
                               MAX(cl.client_budget) as presupuesto_max,
                               COUNT(DISTINCT ec.id) as contratos
                        FROM res_partner rp
                        LEFT JOIN crm_lead cl ON cl.partner_id = rp.id AND cl.active = TRUE
                        LEFT JOIN estate_contract ec ON ec.partner_id = rp.id AND ec.state = 'active'
                        WHERE rp.is_company = FALSE
                          AND (cl.id IS NOT NULL OR ec.id IS NOT NULL)
                        GROUP BY rp.name
                        ORDER BY leads_activos DESC, presupuesto_max DESC NULLS LAST
                        LIMIT %s
                    """, (limit,))
                    rows = env.cr.dictfetchall()
                    if not rows:
                        return json.dumps({'report': 'Ranking de Clientes', 'data': {},
                                           'mensaje': 'No hay clientes con leads o contratos activos.'})
                    data = {r['cliente']: int(r['leads_activos'] or 0) for r in rows[:8]}
                    return json.dumps({'report': 'Clientes más Activos (por leads)',
                                       'data': data, 'chart_hint': 'barra', 'detalle': rows},
                                      ensure_ascii=False, default=str)
                except Exception as e:
                    return json.dumps({'error': f'Error: {str(e)}'})

            return json.dumps({'error': f'report_type desconocido: {report_type}'})

        except Exception as e:
            _logger.error("Error en get_report_data(%s): %s", report_type, str(e))
            return json.dumps({'error': str(e)})

    # -----------------------------------------------------------------------
    # OpenAI with Tool Calling
    # -----------------------------------------------------------------------
    def _query_chatgpt_with_tools(self, api_key, model, temperature, max_tokens,
                                   system_prompt, message, history):
        """Query OpenAI ChatGPT with conversation history and tool calling."""
        if not OPENAI_AVAILABLE:
            return 'La librería openai no está instalada. Ejecute: pip install openai'

        client = openai.OpenAI(api_key=api_key)

        messages = [{"role": "system", "content": system_prompt}]
        messages.extend(history)
        messages.append({"role": "user", "content": message})

        max_iterations = 4
        data_tool_called = False
        hallucination_warned = False
        for _ in range(max_iterations):
            response = client.chat.completions.create(
                model=model,
                messages=messages,
                temperature=temperature,
                max_completion_tokens=max_tokens,
                tools=TOOLS_OPENAI,
                tool_choice="auto",
            )
            choice = response.choices[0]

            if choice.finish_reason == 'tool_calls':
                # Execute each tool call and append results
                assistant_msg = choice.message
                messages.append(assistant_msg)
                for tc in (assistant_msg.tool_calls or []):
                    if tc.function.name in self._DATA_GROUNDING_TOOLS:
                        data_tool_called = True
                    try:
                        args = json.loads(tc.function.arguments)
                    except (json.JSONDecodeError, TypeError):
                        args = {}
                    tool_result = self._execute_tool(tc.function.name, args)
                    messages.append({
                        "role": "tool",
                        "tool_call_id": tc.id,
                        "content": tool_result,
                    })
            else:
                content = choice.message.content or ''
                if '[GRAFICO:' in content and not data_tool_called and not hallucination_warned:
                    hallucination_warned = True
                    messages.append(choice.message)
                    messages.append({
                        "role": "user",
                        "content": (
                            'ADVERTENCIA DEL SISTEMA: incluiste un [GRAFICO:...] sin haber llamado antes '
                            'a ninguna herramienta de datos reales (get_report_data, query_database, etc). '
                            'Nunca inventes cifras. Llama ahora a la herramienta correcta para obtener los '
                            'datos reales, o si el sistema no tiene datos para esa consulta, dilo '
                            'explícitamente sin incluir ningún [GRAFICO:...].'
                        ),
                    })
                    continue
                return content

        return choice.message.content or 'Sin respuesta tras múltiples iteraciones.'

    # -----------------------------------------------------------------------
    # Gemini with Tool Calling (new SDK)
    # -----------------------------------------------------------------------
    def _query_gemini_with_tools(self, api_key, model, temperature, max_tokens,
                                  system_prompt, message, history):
        """Query Google Gemini with conversation history and tool calling."""
        if not GEMINI_AVAILABLE:
            return 'Las librerías de Google Gemini no están instaladas.'

        if NEW_GEMINI_SDK:
            import time as _time
            last_error = None
            # Retry hasta 3 veces en caso de 503 UNAVAILABLE
            for attempt in range(3):
                try:
                    client = new_genai.Client(
                        api_key=api_key,
                        http_options=new_genai.types.HttpOptions(api_version='v1beta'),
                    )

                    # Build function declarations for Gemini
                    func_declarations = []
                    for tool_def in TOOLS_OPENAI:
                        fn = tool_def['function']
                        params = fn.get('parameters', {})
                        properties = {}
                        for prop_name, prop_def in params.get('properties', {}).items():
                            gtype = new_genai.types.Type.STRING
                            if prop_def.get('type') == 'number':
                                gtype = new_genai.types.Type.NUMBER
                            elif prop_def.get('type') == 'integer':
                                gtype = new_genai.types.Type.INTEGER
                            properties[prop_name] = new_genai.types.Schema(
                                type=gtype,
                                description=prop_def.get('description', ''),
                            )
                        gemini_schema = new_genai.types.Schema(
                            type=new_genai.types.Type.OBJECT,
                            properties=properties,
                        ) if properties else None
                        func_declarations.append(
                            new_genai.types.FunctionDeclaration(
                                name=fn['name'],
                                description=fn.get('description', ''),
                                parameters=gemini_schema,
                            )
                        )

                    gemini_tool = new_genai.types.Tool(function_declarations=func_declarations)

                    # Build contents (history + current message)
                    contents = []
                    for h in history:
                        role = 'user' if h['role'] == 'user' else 'model'
                        contents.append(new_genai.types.Content(
                            role=role,
                            parts=[new_genai.types.Part.from_text(text=h['content'])],
                        ))
                    contents.append(new_genai.types.Content(
                        role='user',
                        parts=[new_genai.types.Part.from_text(text=message)],
                    ))

                    config = new_genai.types.GenerateContentConfig(
                        system_instruction=system_prompt,
                        temperature=temperature,
                        max_output_tokens=max_tokens,
                        tools=[gemini_tool],
                    )

                    max_iterations = 4
                    data_tool_called = False
                    hallucination_warned = False
                    for _ in range(max_iterations):
                        response = client.models.generate_content(
                            model=model or _DEFAULT_GEMINI_MODEL,
                            contents=contents,
                            config=config,
                        )
                        candidate = response.candidates[0] if response.candidates else None
                        if not candidate or not candidate.content:
                            break

                        # Check for function calls
                        function_calls = []
                        text_parts = []
                        for part in (candidate.content.parts or []):
                            if hasattr(part, 'function_call') and part.function_call:
                                function_calls.append(part.function_call)
                            elif hasattr(part, 'text') and part.text:
                                text_parts.append(part.text)

                        if function_calls:
                            # Append model response
                            contents.append(candidate.content)
                            # Execute tools and append results
                            function_responses = []
                            for fc in function_calls:
                                if fc.name in self._DATA_GROUNDING_TOOLS:
                                    data_tool_called = True
                                fc_args = dict(fc.args) if fc.args else {}
                                tool_result_str = self._execute_tool(fc.name, fc_args)
                                try:
                                    tool_result = json.loads(tool_result_str)
                                    if not isinstance(tool_result, dict):
                                        tool_result = {'result': tool_result}
                                except Exception:
                                    tool_result = {'result': tool_result_str}
                                function_responses.append(
                                    new_genai.types.Part.from_function_response(
                                        name=fc.name,
                                        response=tool_result,
                                    )
                                )
                            contents.append(new_genai.types.Content(
                                role='user',
                                parts=function_responses,
                            ))
                        else:
                            candidate_text = ''.join(text_parts) or response.text
                            if ('[GRAFICO:' in (candidate_text or '') and not data_tool_called
                                    and not hallucination_warned):
                                hallucination_warned = True
                                contents.append(candidate.content)
                                contents.append(new_genai.types.Content(
                                    role='user',
                                    parts=[new_genai.types.Part.from_text(text=(
                                        'ADVERTENCIA DEL SISTEMA: incluiste un [GRAFICO:...] sin haber '
                                        'llamado antes a ninguna herramienta de datos reales '
                                        '(get_report_data, query_database, etc). Nunca inventes cifras. '
                                        'Llama ahora a la herramienta correcta para obtener los datos '
                                        'reales, o si el sistema no tiene datos para esa consulta, dilo '
                                        'explícitamente sin incluir ningún [GRAFICO:...].'
                                    ))],
                                ))
                                continue
                            return candidate_text

                    return response.text or 'Sin respuesta tras múltiples iteraciones.'

                except Exception as e:
                    last_error = e
                    err_str = _redact(str(e), api_key)
                    etype, emsg, esecs = _parse_gemini_error(err_str)
                    if etype == '429':
                        # Cuota agotada — no reintentar, devolver mensaje claro
                        _logger.warning("Gemini 429 cuota agotada: %s", err_str[:200])
                        return emsg
                    if etype == '503':
                        wait = 5 * (attempt + 1)
                        _logger.warning("Gemini 503 intento %d/3 — esperando %ds", attempt + 1, wait)
                        _time.sleep(wait)
                        continue
                    # Otro error — no reintentar
                    _logger.error("Error Gemini: %s", err_str)
                    return f'Error con Gemini: {err_str}'

            # Agotados los reintentos 503
            return (
                'Gemini no disponible tras 3 intentos (alta demanda). '
                'Prueba con **gemini-2.5-flash** o baja los tokens máximos en Ajustes → Agente IA.'
            )

        return "Error: SDK google-genai no disponible. Ejecute: pip install google-genai"

    # -----------------------------------------------------------------------
    # System Context
    # -----------------------------------------------------------------------
    def _get_system_context(self):
        """Get comprehensive system data for AI context."""
        env = request.env
        props = env['estate.property'].sudo().search([])
        available = props.filtered(lambda p: p.state == 'available')
        sold = props.filtered(lambda p: p.state == 'sold')
        rented = props.filtered(lambda p: p.state == 'rented')
        total_value = sum(props.mapped('price'))
        clients = env['res.partner'].sudo().search([('active', '=', True)])
        leads = env['crm.lead'].sudo().search([])
        pipeline_stats = {}
        for lead in leads:
            stage = lead.stage_id.name or 'Nuevo'
            pipeline_stats[stage] = pipeline_stats.get(stage, 0) + 1
        invoices = env['account.move'].sudo().search([('move_type', '=', 'out_invoice')])
        total_invoiced = sum(invoices.mapped('amount_total'))
        total_commissions = sum(props.filtered(lambda p: p.state == 'sold').mapped('commission_amount'))
        try:
            attendances = env['hr.attendance'].sudo().search([('check_out', '=', False)])
            present_count = len(attendances)
        except Exception:
            present_count = 0

        context = f"""== RESUMEN EJECUTIVO INMOBILIARIO ==
INVENTARIO: {len(props)} propiedades (${total_value:,.2f}) | {len(available)} Disponibles, {len(sold)} Vendidas, {len(rented)} Alquiladas.
FINANZAS: Total Facturado: ${total_invoiced:,.2f} | Comisiones: ${total_commissions:,.2f} | Facturas: {len(invoices)}
CRM: {len(clients)} clientes | {len(leads)} leads | Etapas: {pipeline_stats}
PERSONAL: {present_count} agentes con check-in activo
TOP 10 DISPONIBLES:
"""
        for p in available[:10]:
            context += f"  - {p.name} | {p.title} | {p.city} | ${p.price:,.2f} | {p.property_type_id.name}\n"
        return context

    # -----------------------------------------------------------------------
    # Query Classification
    # -----------------------------------------------------------------------
    def _classify_query(self, message):
        msg = message.lower()
        # Report keywords have HIGHEST priority (even if message also mentions "propiedad")
        if any(w in msg for w in ['reporte', 'informe', 'estadístic', 'dashboard', 'resumen',
                                    'gráfico', 'grafico', 'comision', 'ingreso', 'por estado',
                                    'por tipo', 'por mes', 'por asesor', 'tendencia', 'cuántos hay',
                                    'cuantos hay', 'desglose']):
            return 'report'
        elif any(w in msg for w in ['recuerda', 'memoria', 'anota', 'olvida', 'preferencia']):
            return 'memory'
        elif any(w in msg for w in ['contrato', 'pago', 'cuota', 'vencid', 'arrendamiento', 'alquiler']):
            return 'contract'
        elif any(w in msg for w in ['lead', 'prospecto', 'cliente', 'crm', 'oportunidad', 'temperatura',
                                    'interaccion', 'matchmaker']):
            return 'client'
        elif any(w in msg for w in ['propiedad', 'casa', 'departamento', 'terreno', 'oficina', 'inmueble',
                                   'precio', 'área', 'habitacion', 'baño', 'duplica', 'archiva', 'elimina']):
            return 'property'
        return 'general'

    # Smart tool selection — send only relevant tools based on query type to save tokens
    # -----------------------------------------------------------------------
    # Herramientas que devuelven datos reales de la BD: si el modelo genera un
    # [GRAFICO:...] sin haber llamado a ninguna de estas antes, es una alucinación.
    _DATA_GROUNDING_TOOLS = {
        'get_report_data', 'query_database', 'get_dashboard_summary', 'get_market_stats',
        'search_properties', 'get_leads', 'get_payments_contracts', 'get_trend_analysis',
        'generate_executive_report',
    }

    _TOOLS_BY_CATEGORY = {
        'property': [
            'search_properties', 'get_property_detail', 'analyze_property_improvements',
            'create_property', 'update_property',
            'archive_property', 'delete_property', 'duplicate_property', 'reserve_property',
            'sell_property', 'schedule_visit', 'get_market_stats', 'batch_update_properties',
            'recalculate_avm_ai', 'generate_and_apply_description', 'compare_properties',
            'get_trend_analysis', 'get_report_data', 'generate_analytics_pdf',
            'generate_marketing_pack', 'plan_marketing_campaign', 'query_database',
        ],
        'client': [
            'get_leads', 'create_lead', 'update_lead', 'archive_lead', 'batch_archive_leads',
            'create_crm_activity', 'send_whatsapp_lead', 'schedule_visit', 'search_properties',
            'analyze_lead_probability', 'send_email', 'search_contacts', 'get_client_summary',
            'generate_quote_pdf', 'get_upcoming_visits', 'query_database',
        ],
        'contract': [
            'get_payments_contracts', 'create_contract', 'update_contract', 'create_payment',
            'approve_payment', 'cancel_payment', 'create_offer', 'create_commission',
            'approve_commission', 'analyze_churn_risk', 'generate_pdf_report', 'query_database',
        ],
        'report': [
            'get_report_data', 'get_dashboard_summary', 'get_market_stats', 'get_payments_contracts',
            'get_leads', 'search_properties', 'generate_pdf_report', 'get_trend_analysis',
            'get_upcoming_visits', 'generate_executive_report', 'generate_analytics_pdf',
            'generate_excel_report', 'query_database',
        ],
        'memory': [
            'save_memory', 'recall_memory', 'get_leads', 'search_properties', 'query_database',
        ],
        'general': [
            'search_properties', 'get_property_detail', 'analyze_property_improvements',
            'create_property', 'update_property',
            'archive_property', 'reserve_property', 'sell_property',
            'get_leads', 'create_lead', 'update_lead',
            'get_market_stats', 'get_dashboard_summary', 'get_trend_analysis',
            'schedule_visit', 'get_upcoming_visits',
            'get_payments_contracts', 'save_memory', 'recall_memory',
            'search_contacts', 'get_client_summary', 'compare_properties',
            'get_report_data', 'generate_analytics_pdf', 'generate_excel_report',
            'generate_and_apply_description', 'generate_marketing_pack', 'plan_marketing_campaign',
            'recalculate_avm_ai', 'generate_executive_report', 'query_database',
        ],
    }

    def _get_tools_for_query(self, query_type):
        """Return only the tools relevant to this query type, reducing token usage."""
        allowed = self._TOOLS_BY_CATEGORY.get(query_type, self._TOOLS_BY_CATEGORY['general'])
        tools = [t for t in TOOLS_OPENAI if t['function']['name'] in allowed]
        # search_knowledge (RAG) siempre disponible: cualquier consulta puede ser
        # "cómo se hace", "qué significa este error" o "qué hace tal módulo".
        if not any(t['function']['name'] == 'search_knowledge' for t in tools):
            kb = next((t for t in TOOLS_OPENAI if t['function']['name'] == 'search_knowledge'), None)
            if kb:
                tools.append(kb)
        return tools

    # -----------------------------------------------------------------------
    # Public API endpoints
    # -----------------------------------------------------------------------
    # -----------------------------------------------------------------------
    # Streaming Chat Endpoint (SSE) — faster perceived response
    # -----------------------------------------------------------------------
    @http.route('/estate_ai/chat/stream', type='http', auth='user', methods=['POST'], csrf=False)
    def chat_stream(self, **kwargs):
        """
        Streaming SSE endpoint for the AI chat.
        Sends status events during tool-calling, then streams the final text
        word-by-word so the user sees the response as it is built.
        """
        try:
            data = json.loads(request.httprequest.data or '{}')
        except Exception:
            data = kwargs
        message = (data.get('message') or '').strip()
        session_id = (data.get('session_id') or '').strip() or None
        if not message:
            def _empty():
                yield 'data: {"error":"Mensaje vacío"}\n\ndata: [DONE]\n\n'
            return request.make_response(_empty(), headers=[
                ('Content-Type', 'text/event-stream; charset=utf-8'),
                ('Cache-Control', 'no-cache'),
            ])

        # Pre-fetch ALL database-dependent data NOW, while the cursor is still open.
        # The generate() generator runs AFTER the request cursor is closed (Werkzeug SSE).
        user_id = request.env.user.id
        db_name = request.env.cr.dbname

        ICP = request.env['ir.config_parameter'].sudo()
        ai_active = ICP.get_param('estate_ai.active', 'True')
        # Credenciales del proveedor activo (claves por proveedor con respaldo a las heredadas).
        _chain = self._ai_provider_chain(ICP)
        active_provider = _chain[0][0] if _chain else ''
        api_key = _chain[0][1] if _chain else ''
        model = _chain[0][2] if _chain else ''
        temperature = float(ICP.get_param('estate_ai.temperature', '0.7'))
        # Default 800 tokens — sufficient for most answers, avoids 503 overload
        max_tokens = int(ICP.get_param('estate_ai.max_tokens', '800'))
        system_extra = ICP.get_param('estate_ai.system_prompt', '')

        context = self._get_system_context()

        # Si el frontend manda property_id, enriquecer el contexto con datos de esa propiedad
        page_property_id = int(data.get('property_id') or 0)
        current_model = data.get('current_model') or ''
        if page_property_id and current_model == 'estate.property':
            try:
                p = request.env['estate.property'].sudo().browse(page_property_id)
                if p.exists():
                    visits = request.env['calendar.event'].sudo().search_count(
                        [('property_id', '=', p.id)]) if 'property_id' in request.env['calendar.event']._fields else 0
                    matching_leads = request.env['crm.lead'].sudo().search_count(
                        [('target_property_id', '=', p.id)]) if 'target_property_id' in request.env['crm.lead']._fields else 0
                    avm_info = ''
                    if getattr(p, 'avm_estimated_price', 0):
                        avm_info = f" | AVM: ${p.avm_estimated_price:,.0f} ({getattr(p, 'avm_status', '')})"
                    context += f"""
== PROPIEDAD ACTIVA EN PANTALLA ==
ID: {p.id} | Ref: {p.name} | Título: {p.title or ''}
Tipo: {p.property_type_id.name if p.property_type_id else ''} | Operación: {'Venta' if p.offer_type == 'sale' else 'Arriendo'}
Estado: {dict(p._fields['state'].selection).get(p.state, p.state)} | Ciudad: {p.city or ''} | Sector: {p.street or ''}
Precio: ${p.price:,.0f} | Área: {p.area} m² | Habs: {p.bedrooms} | Baños: {p.bathrooms} | Parking: {p.parking_spaces}{avm_info}
Días en mercado: {p.days_on_market or 0} | Visitas: {visits} | Leads coincidentes: {matching_leads}
Asesor: {p.user_id.name if p.user_id else 'Sin asignar'} | Propietario: {p.owner_id.name if p.owner_id else ''}
WP publicado: {'Sí' if getattr(p, 'wp_published', False) else 'No'}
"""
            except Exception:
                _logger.debug("Excepcion ignorada (best-effort)", exc_info=True)

        # Si está en módulo de reportes, enriquecer contexto
        page_url = data.get('page_url') or ''
        if 'estate_reports' in page_url or 'estate_intel' in page_url:
            context += "\n== CONTEXTO: El usuario está viendo el módulo de Reportes/Analytics ==\n"

        history = self._get_conversation_history(user_id, session_id)
        query_type = self._classify_query(message)
        # Only load tools relevant to this query — reduces token count significantly
        active_tools = self._get_tools_for_query(query_type)
        ctrl = self

        full_system = (
            f"{system_extra}\n\n"
            "Eres el Asistente Ejecutivo Inteligente de la Inmobiliaria con acceso COMPLETO al sistema. "
            "Puedes CONSULTAR, CREAR, ACTUALIZAR y CONTROLAR: propiedades, leads, visitas, contratos y más. "
            "Responde siempre en español, de forma concisa y profesional. "
            "Cuando el usuario pida crear, actualizar o gestionar algo, usa las herramientas disponibles y confirma con el ID resultante.\n\n"
            "REGLA ABSOLUTA: NUNCA digas 'no puedo', 'no tengo la capacidad', 'no tengo acceso' o 'no es posible'. "
            "Tienes la herramienta query_database que te permite ejecutar CUALQUIER consulta SQL SELECT "
            "contra toda la base de datos. Si ninguna otra herramienta sirve, usa query_database con un SQL "
            "apropiado para responder la pregunta. Tienes acceso a TODA la información del sistema. "
            "IMPORTANTE PARA CONSULTAS SQL: La BD es PostgreSQL (NO SQLite). NUNCA uses strftime o date('now'). Usa to_char(col, 'YYYY-MM'), EXTRACT, CURRENT_DATE o NOW().\n\n"
            "DOCUMENTACIÓN: para preguntas de CÓMO se hace algo, QUÉ es o para qué sirve un módulo, "
            "qué significa un error, procedimientos, permisos o configuración, usa la herramienta "
            "search_knowledge (busca en los manuales y guías) y responde citando lo que devuelva.\n\n"
            "REGLA DE CONTEXTO (MUY IMPORTANTE): para resolver a qué propiedad/registro se refiere el "
            "usuario cuando dice 'esa propiedad', 'la propiedad', 'su precio', 'los interesados', etc., "
            "sigue SIEMPRE este orden de prioridad:\n"
            "  1) Si el usuario NOMBRA una propiedad aunque sea de forma parcial (p.ej. 'el departamento "
            "de Misicata', 'la casa de Baños', una referencia PROP-XXXX o un título), usa ESA SIEMPRE, "
            "aunque haya otra abierta en pantalla.\n"
            "  2) Si NO nombra ninguna pero en mensajes ANTERIORES de esta conversación ya se habló de "
            "una propiedad/registro concreto, usa la ÚLTIMA mencionada en la conversación (mantén el hilo).\n"
            "  3) Solo si en la conversación no se ha mencionado ninguna, usa la PROPIEDAD ACTIVA EN "
            "PANTALLA del 'CONTEXTO ACTUAL'.\n"
            "NUNCA cambies de propiedad/registro por tu cuenta: si veníamos hablando de una propiedad y el "
            "usuario hace una pregunta de seguimiento ('dame sus características', 'quién es el dueño', "
            "'genera su descripción', 'agenda una visita'), responde sobre ESA MISMA, no sobre otra ni "
            "sobre la que esté abierta en pantalla. Y NUNCA pidas el ID: usa el nombre o el contexto.\n\n"
            "REGLA COMERCIAL DE ANÁLISIS DE VENTAS (DOBLE ENFOQUE PROFESIONAL):\n"
            "Cuando el usuario pregunte por 'ventas', 'cierres', 'promedio de ventas', 'ingresos' o pida reportes o análisis de ventas, SIEMPRE analiza y presenta la información diferenciando DOS enfoques claros:\n"
            "  a) Honorarios y Comisiones de Agencia (commission_amount): Representa la ganancia real y el ingreso de la agencia inmobiliaria por las operaciones cerradas.\n"
            "  b) Volumen y Precio de Inmuebles (price): Representa el valor total y el precio promedio al que se vendieron las casas/propiedades, para evaluar qué tan costosos son los inmuebles movidos y el ticket promedio en el mercado.\n"
            "En tus explicaciones, resúmenes, gráficos y tablas compara ambas métricas por separado y NO las confundas.\n\n"
            "REGLA OBLIGATORIA PARA REPORTES Y GRÁFICOS:\n"
            "Cuando el usuario pida reporte, gráfico, estadística, resumen de datos, desglose, o use palabras como "
            "'muéstrame por', 'cuántos hay por', 'reporte de', 'gráfico de' → DEBES llamar a get_report_data. "
            "NUNCA respondas con solo texto cuando se pide un gráfico o reporte.\n"
            "NUNCA inventes cifras ni reutilices los números de los ejemplos de abajo (son solo formato, "
            "no datos reales de este sistema). El [GRAFICO:...] y la tabla deben usar EXACTAMENTE los "
            "valores que devolvió get_report_data/query_database. Si 'data' viene vacío o en 0, NO generes "
            "ningún [GRAFICO:...]: dilo explícitamente (ej. 'Aún no hay ventas registradas en el sistema').\n"
            "Con los datos recibidos SIEMPRE genera TODOS los gráficos posibles que apliquen.\n"
            "Tipos de gráfico disponibles (SIEMPRE incluye un título descriptivo con |):\n"
            "- [GRAFICO:barra|Título descriptivo,Label1:Valor1,Label2:Valor2,...] → barras horizontales\n"
            "- [GRAFICO:circular|Título descriptivo,Label1:Valor1,Label2:Valor2,...] → diagrama de torta\n"
            "- [GRAFICO:linea|Título descriptivo,Label1:Valor1,Label2:Valor2,...] → línea temporal\n"
            "- [GRAFICO:histograma|Título descriptivo,Label1:Valor1,...] → histograma de frecuencias\n"
            "- [GRAFICO:dispersion|Título descriptivo,Label1:Valor1,...] → dispersión/scatter\n"
            "- [GRAFICO:gantt|Título descriptivo,Label1:Valor1,...] → barras horizontales tipo Gantt\n"
            "- [GRAFICO:calor|Título descriptivo,Label1:Valor1,...] → mapa de calor (intensidad por color)\n"
            "Formato (solo de ejemplo, NUNCA copies estos valores): "
            "[GRAFICO:barra|<Título>,<Label1>:<Valor1>,<Label2>:<Valor2>]\n"
            "REGLA: Elige automáticamente el MEJOR tipo de gráfico según los datos:\n"
            "- Datos temporales (meses, años) → linea\n"
            "- Proporciones/distribuciones (estados, tipos) → circular\n"
            "- Comparaciones de cantidades/rankings → barra\n"
            "- Distribución de frecuencias → histograma\n"
            "- Actividad por períodos/intensidad → calor\n"
            "- Si hay duda, usa barra (es el más versátil)\n"
            "- Si los datos permiten más de una visualización útil, incluye MÚLTIPLES gráficos "
            "(ej: uno circular para % y uno de barra para cantidades absolutas).\n"
            "Después de los gráficos incluye una tabla Markdown con los mismos datos.\n"
            "report_types: properties_by_state, properties_by_type, sales_by_month, visits_by_property, "
            "commissions_by_advisor, contracts_by_type, expenses_by_type, offers_by_state, "
            "leads_by_temperature, payments_by_method, days_on_market_by_type, time_to_sell_summary, "
            "sales_by_channel.\n"
            "Usa time_to_sell_summary para 'cuánto se tarda en vender' o 'tiempo de venta' (resumen "
            "general: promedio, mediana, mínimo, máximo desde publicación hasta venta).\n"
            "Usa sales_by_channel para 'ventas por agencia', 'quién vendió' o 'cerrado por "
            "agencia/propietario' (según el campo sold_by de la propiedad).\n\n"
            f"CONTEXTO ACTUAL:\n{context}"
        )

        def sse(payload):
            return f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"

        def _tool_with_cursor(tool_name, args):
            """Execute a tool using a fresh DB cursor (safe to call inside generator)."""
            from odoo.modules.registry import Registry
            from odoo import api as odoo_api
            with Registry(db_name).cursor() as new_cr:
                new_env = odoo_api.Environment(new_cr, user_id, {})
                result = ctrl._execute_tool(tool_name, args, env=new_env)
                new_cr.commit()
                return result

        def _save_history(final_text):
            """Persist chat history with a fresh DB cursor."""
            from odoo.modules.registry import Registry
            from odoo import api as odoo_api
            with Registry(db_name).cursor() as new_cr:
                new_env = odoo_api.Environment(new_cr, user_id, {})
                vals = {
                    'user_id': user_id,
                    'query': message,
                    'response': final_text,
                    'query_type': query_type,
                    'processing_time': 0,
                }
                if session_id:
                    vals['session_id'] = session_id
                new_env['estate.ai.chat.history'].sudo().create(vals)
                new_cr.commit()

        def _stream_chatgpt():
            """Streaming (por palabras) para el proveedor ChatGPT/OpenAI.
            Antes esta ruta usaba SIEMPRE el cliente de Gemini sin importar el
            proveedor configurado, por lo que una API Key de OpenAI terminaba
            enviándose a Google y fallaba con 'API key not valid'."""
            if not OPENAI_AVAILABLE:
                yield sse({'error': 'Librería openai no instalada. Ejecute: pip install openai'})
                return
            tool_labels = {
                'search_properties': 'Buscando propiedades', 'get_leads': 'Consultando leads CRM',
                'get_market_stats': 'Calculando estadísticas', 'create_lead': 'Creando lead',
                'create_property': 'Registrando propiedad', 'update_lead': 'Actualizando lead',
                'update_property': 'Actualizando propiedad', 'schedule_visit': 'Agendando visita',
                'reserve_property': 'Reservando propiedad', 'sell_property': 'Cerrando venta',
                'get_report_data': 'Cargando datos', 'query_database': 'Consultando base de datos',
                'get_dashboard_summary': 'Generando resumen', 'get_payments_contracts': 'Consultando pagos',
                'generate_analytics_pdf': 'Generando PDF del reporte',
                'generate_excel_report': 'Generando Excel', 'save_memory': 'Guardando memoria',
                'recall_memory': 'Consultando memorias', 'search_knowledge': 'Buscando en manuales',
            }
            try:
                client = openai.OpenAI(api_key=api_key)
                messages = [{"role": "system", "content": full_system}]
                messages.extend(history)
                messages.append({"role": "user", "content": message})

                final_text = ''
                data_tool_called = False
                hallucination_warned = False
                for _ in range(4):
                    response = client.chat.completions.create(
                        model=model, messages=messages, temperature=temperature,
                        max_completion_tokens=max_tokens, tools=active_tools, tool_choice="auto",
                    )
                    choice = response.choices[0]
                    if choice.finish_reason == 'tool_calls':
                        assistant_msg = choice.message
                        messages.append(assistant_msg)
                        calls = assistant_msg.tool_calls or []
                        names = [tool_labels.get(tc.function.name, tc.function.name) for tc in calls]
                        if names:
                            yield sse({'status': ' · '.join(names) + '...'})
                        for tc in calls:
                            if tc.function.name in ctrl._DATA_GROUNDING_TOOLS:
                                data_tool_called = True
                            try:
                                args = json.loads(tc.function.arguments)
                            except (json.JSONDecodeError, TypeError):
                                args = {}
                            tool_result = _tool_with_cursor(tc.function.name, args)
                            messages.append({
                                "role": "tool", "tool_call_id": tc.id, "content": tool_result,
                            })
                        yield sse({'status': 'Redactando respuesta...'})
                    else:
                        content = choice.message.content or ''
                        if ('[GRAFICO:' in content and not data_tool_called
                                and not hallucination_warned):
                            hallucination_warned = True
                            messages.append(choice.message)
                            messages.append({
                                "role": "user",
                                "content": (
                                    'ADVERTENCIA DEL SISTEMA: incluiste un [GRAFICO:...] sin haber '
                                    'llamado antes a ninguna herramienta de datos reales '
                                    '(get_report_data, query_database, etc). Nunca inventes cifras. '
                                    'Llama ahora a la herramienta correcta para obtener los datos '
                                    'reales, o si el sistema no tiene datos para esa consulta, dilo '
                                    'explícitamente sin incluir ningún [GRAFICO:...].'
                                ),
                            })
                            continue
                        final_text = content
                        break

                # Si se agotaron las rondas de herramientas sin redactar una
                # respuesta final, forzar una respuesta de texto SIN herramientas
                # que resuma los datos ya obtenidos (misma red de seguridad que
                # la ruta de Gemini) en vez de rendirse con un mensaje genérico.
                if not final_text:
                    try:
                        yield sse({'status': 'Redactando respuesta...'})
                        messages.append({
                            "role": "user",
                            "content": (
                                "Responde AHORA al usuario en español, resumiendo de forma clara y "
                                "útil los datos ya obtenidos de las herramientas. No llames más "
                                "herramientas."
                            ),
                        })
                        final_response = client.chat.completions.create(
                            model=model, messages=messages, temperature=temperature,
                            max_completion_tokens=max_tokens, tool_choice="none",
                        )
                        final_text = (final_response.choices[0].message.content or '').strip()
                    except Exception as fe:
                        _logger.warning("Generación final sin herramientas falló: %s", str(fe))

                if not final_text:
                    final_text = (
                        'No pude generar una respuesta para esa consulta. Intenta reformularla '
                        'o sé más específico (por ejemplo, indica la propiedad, el cliente o el periodo).'
                    )

                chunk = ''
                for word in final_text.split(' '):
                    chunk += word + ' '
                    if len(chunk) >= 12:
                        yield sse({'text': chunk})
                        chunk = ''
                if chunk:
                    yield sse({'text': chunk})

                try:
                    _save_history(final_text)
                except Exception as he:
                    _logger.warning("No se pudo guardar historial IA: %s", str(he))

            except Exception as e:
                err_str = _redact(str(e), api_key)
                _logger.error("Error en streaming ChatGPT: %s", err_str)
                yield sse({'text': ctrl._friendly_ai_error('chatgpt', e)})

        def generate():
            # All values come from closure — no ORM calls here
            if ai_active != 'True':
                yield sse({'error': 'Agente IA desactivado.'})
                yield 'data: [DONE]\n\n'
                return

            if not api_key:
                yield sse({'error': 'No hay API Key. Vaya a Configuración > Agente IA.'})
                yield 'data: [DONE]\n\n'
                return

            if active_provider == 'chatgpt':
                yield from _stream_chatgpt()
                yield 'data: [DONE]\n\n'
                return

            if not NEW_GEMINI_SDK:
                yield sse({'error': 'Librería google-genai no instalada.'})
                yield 'data: [DONE]\n\n'
                return

            import time as _time
            tool_labels = {
                'search_properties': 'Buscando propiedades',
                'get_property_detail': 'Consultando propiedad',
                'analyze_property_improvements': 'Analizando mejoras de la propiedad',
                'get_leads': 'Consultando leads CRM',
                'get_market_stats': 'Calculando estadísticas',
                'create_crm_activity': 'Creando actividad',
                'create_lead': 'Creando lead',
                'create_property': 'Registrando propiedad',
                'update_lead': 'Actualizando lead',
                'update_property': 'Actualizando propiedad',
                'delete_property': 'Eliminando propiedad',
                'duplicate_property': 'Duplicando propiedad',
                'schedule_visit': 'Agendando visita',
                'reserve_property': 'Reservando propiedad',
                'sell_property': 'Cerrando venta',
                'send_whatsapp_lead': 'Generando enlace WhatsApp',
                'archive_lead': 'Archivando lead',
                'archive_property': 'Archivando propiedad',
                'get_payments_contracts': 'Consultando pagos',
                'get_dashboard_summary': 'Generando resumen',
                'create_contract': 'Creando contrato',
                'create_payment': 'Registrando pago',
                'create_offer': 'Creando oferta',
                'approve_payment': 'Aprobando pago',
                'generate_pdf_report': 'Generando PDF',
                'save_memory': 'Guardando memoria',
                'recall_memory': 'Consultando memorias',
                'analyze_lead_probability': 'Analizando lead',
                'analyze_churn_risk': 'Analizando riesgo',
                'recalculate_avm_ai': 'Calculando valoración',
                'generate_and_apply_description': 'Generando descripción',
                'generate_marketing_pack': 'Creando pack de marketing',
                'generate_analytics_pdf': 'Generando PDF del reporte',
                'plan_marketing_campaign': 'Analizando estrategia de campaña',
                'generate_executive_report': 'Generando informe ejecutivo del mes',
                'send_email': 'Enviando email',
                'get_report_data': 'Cargando datos',
                'query_database': 'Consultando base de datos',
            }

            last_err = None
            for _attempt in range(3):
                try:
                    client = new_genai.Client(
                        api_key=api_key,
                        http_options=new_genai.types.HttpOptions(api_version='v1beta'),
                    )

                    # Build function declarations — only relevant tools (saves tokens)
                    func_decls = []
                    for td in active_tools:
                        fn = td['function']
                        props = {}
                        for pname, pdef in fn.get('parameters', {}).get('properties', {}).items():
                            gtype = new_genai.types.Type.STRING
                            if pdef.get('type') == 'number':
                                gtype = new_genai.types.Type.NUMBER
                            elif pdef.get('type') == 'integer':
                                gtype = new_genai.types.Type.INTEGER
                            props[pname] = new_genai.types.Schema(
                                type=gtype,
                                description=pdef.get('description', ''),
                            )
                        func_decls.append(new_genai.types.FunctionDeclaration(
                            name=fn['name'],
                            description=fn.get('description', ''),
                            parameters=new_genai.types.Schema(
                                type=new_genai.types.Type.OBJECT,
                                properties=props,
                            ) if props else None,
                        ))

                    gemini_tool = new_genai.types.Tool(function_declarations=func_decls)
                    cfg = new_genai.types.GenerateContentConfig(
                        system_instruction=full_system,
                        temperature=temperature,
                        max_output_tokens=max_tokens,
                        tools=[gemini_tool],
                    )

                    # Build contents from pre-fetched history
                    contents = []
                    for h in history:
                        role = 'user' if h['role'] == 'user' else 'model'
                        contents.append(new_genai.types.Content(
                            role=role,
                            parts=[new_genai.types.Part.from_text(text=h['content'])],
                        ))
                    contents.append(new_genai.types.Content(
                        role='user',
                        parts=[new_genai.types.Part.from_text(text=message)],
                    ))

                    # Tool-calling loop (max 4 rounds)
                    final_text = ''
                    data_tool_called = False
                    hallucination_warned = False
                    for _ in range(4):
                        response = client.models.generate_content(
                            model=model, contents=contents, config=cfg)
                        candidate = response.candidates[0] if response.candidates else None
                        if not candidate or not candidate.content:
                            break

                        fn_calls, text_parts = [], []
                        for part in (candidate.content.parts or []):
                            if hasattr(part, 'function_call') and part.function_call:
                                fn_calls.append(part.function_call)
                            elif hasattr(part, 'text') and part.text:
                                text_parts.append(part.text)

                        if not fn_calls:
                            candidate_text = ''.join(text_parts) or (response.text or '')
                            # El modelo generó un [GRAFICO:...] sin haber consultado datos
                            # reales primero: nunca confiar en cifras inventadas, se le pide
                            # corregir antes de aceptar la respuesta (una sola vez).
                            if ('[GRAFICO:' in candidate_text and not data_tool_called
                                    and not hallucination_warned):
                                hallucination_warned = True
                                contents.append(candidate.content)
                                contents.append(new_genai.types.Content(
                                    role='user',
                                    parts=[new_genai.types.Part.from_text(text=(
                                        'ADVERTENCIA DEL SISTEMA: incluiste un [GRAFICO:...] sin haber '
                                        'llamado antes a ninguna herramienta de datos reales '
                                        '(get_report_data, query_database, etc). Nunca inventes cifras. '
                                        'Llama ahora a la herramienta correcta para obtener los datos '
                                        'reales, o si el sistema no tiene datos para esa consulta, dilo '
                                        'explícitamente sin incluir ningún [GRAFICO:...].'
                                    ))],
                                ))
                                continue
                            final_text = candidate_text
                            break

                        # Notify which tools are running
                        names = [tool_labels.get(fc.name, fc.name) for fc in fn_calls]
                        yield sse({'status': ' · '.join(names) + '...'})

                        contents.append(candidate.content)
                        fn_responses = []
                        for fc in fn_calls:
                            if fc.name in self._DATA_GROUNDING_TOOLS:
                                data_tool_called = True
                            fc_args = dict(fc.args) if fc.args else {}
                            result_str = _tool_with_cursor(fc.name, fc_args)
                            try:
                                result = json.loads(result_str)
                                if not isinstance(result, dict):
                                    result = {'result': result}
                            except Exception:
                                result = {'result': result_str}
                            fn_responses.append(new_genai.types.Part.from_function_response(
                                name=fc.name, response=result))
                        contents.append(new_genai.types.Content(role='user', parts=fn_responses))
                        yield sse({'status': 'Redactando respuesta...'})

                    # Si el modelo usó herramientas pero no redactó respuesta
                    # (o agotó las rondas), forzar una respuesta final SIN
                    # herramientas que resuma los datos ya obtenidos.
                    if not final_text:
                        try:
                            yield sse({'status': 'Redactando respuesta...'})
                            cfg_final = new_genai.types.GenerateContentConfig(
                                system_instruction=full_system + (
                                    "\n\nResponde AHORA al usuario en español, resumiendo de forma "
                                    "clara y útil los datos ya obtenidos de las herramientas. "
                                    "No llames más herramientas."),
                                temperature=temperature,
                                max_output_tokens=max_tokens,
                            )
                            final_resp = client.models.generate_content(
                                model=model, contents=contents, config=cfg_final)
                            final_text = (getattr(final_resp, 'text', '') or '').strip()
                        except Exception as fe:
                            _logger.warning("Generación final sin herramientas falló: %s", str(fe))

                    # Stream final text word by word
                    if final_text:
                        chunk = ''
                        for word in final_text.split(' '):
                            chunk += word + ' '
                            if len(chunk) >= 12:
                                yield sse({'text': chunk})
                                chunk = ''
                        if chunk:
                            yield sse({'text': chunk})
                    else:
                        yield sse({'text': (
                            'No pude generar una respuesta para esa consulta. '
                            'Intenta reformularla o sé más específico (por ejemplo, indica '
                            'la propiedad, el cliente o el periodo).')})

                    # Persist history
                    try:
                        _save_history(final_text)
                    except Exception as he:
                        _logger.warning("No se pudo guardar historial IA: %s", str(he))

                    break  # success — exit retry loop

                except Exception as e:
                    last_err = e
                    err_str = _redact(str(e), api_key)
                    etype, emsg, esecs = _parse_gemini_error(err_str)
                    if etype == '429':
                        _logger.warning("Gemini 429 cuota agotada (streaming)")
                        yield sse({'text': emsg})
                        break  # No reintentar — cuota es diaria
                    if etype == '503':
                        wait = 5 * (_attempt + 1)
                        _logger.warning("Gemini 503 intento %d/3 — esperando %ds", _attempt + 1, wait)
                        yield sse({'status': f'⏳ Servidor ocupado, reintentando en {wait}s...'})
                        _time.sleep(wait)
                        continue
                    # Otro error no reintentable
                    _logger.error("Error en streaming IA: %s", err_str)
                    yield sse({'text': f'Error: {err_str}'})
                    break
            else:
                # 3 intentos 503 fallidos
                yield sse({'text': (
                    'Gemini no disponible (alta demanda). '
                    'Ve a **Ajustes → Agente IA** y cambia el modelo a `gemini-2.5-flash` '
                    'o baja los tokens máximos a 500.'
                )})

            yield 'data: [DONE]\n\n'

        return request.make_response(generate(), headers=[
            ('Content-Type', 'text/event-stream; charset=utf-8'),
            ('Cache-Control', 'no-cache, no-transform'),
            ('X-Accel-Buffering', 'no'),
        ])

    @http.route('/estate_ai/history', type='jsonrpc', auth='user', methods=['POST'])
    def get_history(self, limit=20, **kwargs):
        """Get chat history for current user."""
        history = request.env['estate.ai.chat.history'].search(
            [('user_id', '=', request.env.user.id)],
            limit=limit, order='create_date desc')
        return [{'query': h.query, 'date': h.create_date.strftime('%d/%m/%Y %H:%M')}
                for h in reversed(history)]

    @http.route('/estate_ai/clear', type='jsonrpc', auth='user', methods=['POST'])
    def clear_history(self, session_id=None, **kwargs):
        """Clear chat history for current user. If session_id given, clears only that session."""
        domain = [('user_id', '=', request.env.user.id)]
        if session_id:
            domain.append(('session_id', '=', session_id))
        request.env['estate.ai.chat.history'].search(domain).unlink()
        return True

    @http.route('/estate_ai/sessions', type='jsonrpc', auth='user', methods=['POST'])
    def get_sessions(self, **kwargs):
        """Return list of distinct chat sessions for the current user."""
        env = request.env
        user_id = env.user.id
        # Get all history records ordered by date, group by session_id in Python
        records = env['estate.ai.chat.history'].sudo().search(
            [('user_id', '=', user_id)],
            order='create_date asc',
            limit=200,
        )
        sessions = {}
        for r in records:
            sid = r.session_id or 'default'
            if sid not in sessions:
                sessions[sid] = {
                    'session_id': sid,
                    'title': r.query[:60] if r.query else 'Conversación',
                    'date': r.create_date.strftime('%d/%m %H:%M'),
                    'count': 0,
                }
            sessions[sid]['count'] += 1
            sessions[sid]['date'] = r.create_date.strftime('%d/%m %H:%M')
        # Most recent first
        result = list(reversed(list(sessions.values())))
        return result[:50]

    @http.route('/estate_ai/session_messages', type='jsonrpc', auth='user', methods=['POST'])
    def get_session_messages(self, session_id=None, **kwargs):
        """Return all messages for a specific session."""
        domain = [('user_id', '=', request.env.user.id)]
        if session_id and session_id != 'default':
            domain.append(('session_id', '=', session_id))
        else:
            domain.append(('session_id', 'in', [False, '', 'default']))
        records = request.env['estate.ai.chat.history'].sudo().search(
            domain, order='create_date asc', limit=100)
        result = []
        for r in records:
            result.append({'type': 'user', 'text': r.query, 'date': r.create_date.strftime('%H:%M')})
            result.append({'type': 'bot', 'text': r.response or '', 'date': r.create_date.strftime('%H:%M')})
        return result

    @http.route('/estate_ai/suggestions', type='jsonrpc', auth='user', methods=['POST'])
    def get_suggestions(self, context=None, **kwargs):
        """Get dynamic suggested queries based on current system state."""
        from datetime import date, timedelta
        env = request.env
        suggestions = []

        try:
            # Context-aware suggestions based on system state
            today = date.today()

            # Overdue payments alert
            overdue = env['estate.payment'].sudo().search_count([
                ('state', '=', 'pending'), ('date', '<', today)])
            if overdue:
                suggestions.append(f'Muéstrame los {overdue} pagos vencidos y qué hacer')

            # Hot leads without activity
            stale_hot = env['crm.lead'].sudo().search_count([
                ('lead_temperature', 'in', ['hot', 'boiling']),
                ('write_date', '<=', str(fields.Datetime.now() - timedelta(days=5))),
                ('type', '=', 'opportunity'),
            ])
            if stale_hot:
                suggestions.append(f'Hay {stale_hot} leads calientes sin actividad — ¿qué hago?')

            # Properties available > 60 days
            stale_props = env['estate.property'].sudo().search_count([
                ('state', '=', 'available'),
                ('date_listed', '<=', today - timedelta(days=60)),
            ])
            if stale_props:
                suggestions.append(f'Analiza las {stale_props} propiedades sin vender en más de 60 días')

            # Visits today
            today_visits = env['calendar.event'].sudo().search_count([
                ('start', '>=', str(today)),
                ('start', '<', str(today + timedelta(days=1))),
            ])
            if today_visits:
                suggestions.append(f'¿Cuáles son mis {today_visits} visitas de hoy?')

        except Exception:
            _logger.debug("Excepcion ignorada (best-effort)", exc_info=True)

        # Always include these core suggestions
        suggestions += [
            'Dame el briefing del día',
            'Compara tendencias de este mes vs el mes pasado',
            'Busca clientes interesados en casa en Cuenca',
            '¿Cuáles son los leads más calientes del CRM?',
            'Genera un reporte de comisiones por asesor',
            'Analiza el riesgo de churn de mis contratos',
        ]

        return suggestions[:8]

    @http.route('/estate_ai/alert_chips', type='jsonrpc', auth='user', methods=['POST'])
    def get_alert_chips(self, **kwargs):
        """Returns contextual alert chips with real counts from the DB."""
        from datetime import date, timedelta
        env = request.env
        chips = []
        try:
            today = date.today()
            stale_count = env['estate.property'].sudo().search_count([
                ('state', '=', 'available'),
                ('date_listed', '<=', today - timedelta(days=60)),
            ])
            if stale_count:
                chips.append(f'Propiedades sin vender +60 días ({stale_count})')

            cold_leads = 0
            if 'lead_temperature' in env['crm.lead']._fields:
                cold_leads = env['crm.lead'].sudo().search_count([
                    ('active', '=', True),
                    ('type', '=', 'opportunity'),
                    ('lead_temperature', 'in', ['cold', 'warm']),
                    ('write_date', '<=', fields.Datetime.now() - timedelta(days=30)),
                ])
            if cold_leads:
                chips.append(f'Leads fríos sin actividad ({cold_leads})')

            overdue = env['estate.payment'].sudo().search_count([
                ('state', '=', 'pending'), ('date', '<', today)])
            if overdue:
                chips.append(f'Pagos vencidos ({overdue})')
        except Exception:
            _logger.debug("Excepcion ignorada (best-effort)", exc_info=True)

        chips += [
            'Informe ejecutivo completo del mes',
            'KPIs generales del mes',
            'Ranking de asesores',
        ]
        return chips[:6]

    @http.route('/estate_ai/feedback', type='jsonrpc', auth='user', methods=['POST'])
    def record_feedback(self, vote=None, **kwargs):
        """B5: guarda el voto positivo/negativo del usuario sobre una respuesta de la IA."""
        try:
            fid = request.env['estate.ai.feedback'].sudo().record_vote(
                vote,
                session_id=kwargs.get('session_id'),
                question=kwargs.get('question'),
                answer=kwargs.get('answer'),
                page_context=kwargs.get('page_context'),
            )
            return {'ok': bool(fid), 'id': fid}
        except Exception:
            return {'ok': False}

    @http.route('/estate_ai/briefing', type='jsonrpc', auth='user', methods=['POST'])
    def get_briefing(self, **kwargs):
        """
        Briefing matutino: resumen ejecutivo + visitas del día + tendencias + alertas.
        Devuelve el texto del briefing directamente (sin pasar por el modelo de IA).
        """
        from datetime import date, timedelta
        from datetime import datetime as _dt
        env = request.env
        today = date.today()
        now = _dt.now()

        lines = [f"## Briefing del {today.strftime('%A %d de %B de %Y')}\n"]

        try:
            # 1. Inventario
            available = env['estate.property'].sudo().search_count([('state', '=', 'available')])
            reserved = env['estate.property'].sudo().search_count([('state', '=', 'reserved')])
            lines.append(f"**Inventario:** {available} disponibles · {reserved} reservadas\n")

            # 2. Visitas de hoy
            visits_today = env['calendar.event'].sudo().search([
                ('start', '>=', str(today)),
                ('start', '<', str(today + timedelta(days=1))),
            ], order='start asc', limit=5)
            if visits_today:
                lines.append(f"**Visitas hoy ({len(visits_today)}):**")
                for v in visits_today:
                    prop = getattr(v, 'property_id', None)
                    prop_name = prop.title if prop else ''
                    lines.append(f"  - {str(v.start)[11:16]} | {v.name} {('— ' + prop_name) if prop_name else ''}")
                lines.append('')
            else:
                lines.append("**Visitas hoy:** Ninguna programada\n")

            # 3. Alertas críticas
            alerts = []
            overdue = env['estate.payment'].sudo().search_count([
                ('state', '=', 'pending'), ('date', '<', today)])
            if overdue:
                alerts.append(f"{overdue} pagos vencidos")

            hot_stale = env['crm.lead'].sudo().search_count([
                ('lead_temperature', 'in', ['hot', 'boiling']),
                ('write_date', '<=', str(fields.Datetime.now() - timedelta(days=7))),
                ('type', '=', 'opportunity'),
            ])
            if hot_stale:
                alerts.append(f"{hot_stale} leads calientes sin actividad en 7+ días")

            expiring = env['estate.contract'].sudo().search_count([
                ('state', '=', 'active'),
                ('date_end', '>=', str(today)),
                ('date_end', '<=', str(today + timedelta(days=30))),
            ])
            if expiring:
                alerts.append(f"{expiring} contratos vencen en 30 días")

            if alerts:
                lines.append("**Alertas:**")
                for a in alerts:
                    lines.append(f"  - {a}")
                lines.append('')
            else:
                lines.append("**Alertas:** Sin alertas críticas\n")

            # 4. Tendencia del mes
            month_start = today.replace(day=1)
            sales_month = env['estate.property'].sudo().search_count([
                ('state', '=', 'sold'), ('date_sold', '>=', str(month_start))])
            leads_month = env['crm.lead'].sudo().search_count([
                ('create_date', '>=', str(month_start))])
            lines.append(f"**Este mes:** {sales_month} ventas · {leads_month} nuevos leads")

        except Exception as e:
            lines.append(f"(Error generando briefing: {e})")

        briefing_text = "\n".join(lines)

        # Log to history
        env['estate.ai.chat.history'].sudo().create({
            'user_id': env.user.id,
            'query': '/briefing',
            'response': briefing_text,
            'query_type': 'report',
        })

        return {'response': briefing_text}

    @http.route('/estate_ai/memories', type='jsonrpc', auth='user', methods=['POST'])
    def get_memories(self, **kwargs):
        """Memorias activas del usuario (para mostrarlas de forma visible en el chat)."""
        try:
            Memory = request.env['estate.ai.memory'].sudo()
            return Memory.get_active_memories_for_user(request.env.user.id, limit=8)
        except Exception as e:
            _logger.warning("get_memories falló: %s", e)
            return []

    # -----------------------------------------------------------------------
    # OCR Endpoint (C1) — extract data from uploaded documents via Gemini Vision
    # -----------------------------------------------------------------------
    @http.route('/estate_ai/ocr', type='http', auth='user', methods=['POST'], csrf=False)
    def ocr_document(self, **kwargs):
        """
        Upload a file (image or PDF) and extract structured data using Gemini Vision.
        Returns JSON with extracted fields.
        """
        import base64
        import mimetypes

        try:
            uploaded_file = kwargs.get('file') or request.httprequest.files.get('file')
            extract_type = kwargs.get('extract_type', 'auto')  # auto, property, contract, identity

            if not uploaded_file:
                return request.make_response(
                    json.dumps({'error': 'No se recibió ningún archivo.'}),
                    headers=[('Content-Type', 'application/json')]
                )

            file_bytes = uploaded_file.read()
            filename = getattr(uploaded_file, 'filename', 'document')
            mime_type = mimetypes.guess_type(filename)[0] or 'application/octet-stream'
            file_b64 = base64.b64encode(file_bytes).decode('utf-8')

            ICP = request.env['ir.config_parameter'].sudo()
            provider = ICP.get_param('estate_ai.provider', 'gemini')
            api_key = ICP.get_param('estate_ai.api_key', '')

            if not api_key:
                return request.make_response(
                    json.dumps({'error': 'No hay API Key configurada.'}),
                    headers=[('Content-Type', 'application/json')]
                )

            # Build extraction prompt based on type
            prompts = {
                'property': (
                    "Extrae los datos de esta propiedad inmobiliaria en JSON: "
                    "{titulo, direccion, ciudad, precio, area_m2, habitaciones, banos, descripcion}"
                ),
                'contract': (
                    "Extrae los datos de este contrato en JSON: "
                    "{tipo_contrato, nombre_propietario, nombre_inquilino, fecha_inicio, "
                    "fecha_fin, monto_mensual, direccion_propiedad}"
                ),
                'identity': (
                    "Extrae los datos de este documento de identidad en JSON: "
                    "{nombre_completo, numero_cedula, fecha_nacimiento, direccion}"
                ),
                'auto': (
                    "Analiza este documento y extrae TODOS los datos relevantes en formato JSON. "
                    "Identifica el tipo de documento y devuelve los campos más importantes."
                ),
            }
            ocr_prompt = prompts.get(extract_type, prompts['auto'])

            extracted = {}
            if GEMINI_AVAILABLE and (provider == 'gemini' or not OPENAI_AVAILABLE):
                client = new_genai.Client(api_key=api_key)
                response = client.models.generate_content(
                    model=_DEFAULT_GEMINI_MODEL,
                    contents=[
                        {
                            'parts': [
                                {'inline_data': {'mime_type': mime_type, 'data': file_b64}},
                                {'text': ocr_prompt},
                            ]
                        }
                    ]
                )
                raw_text = response.text or ''
                # Try to parse JSON from response
                import re
                json_match = re.search(r'\{[\s\S]*\}', raw_text)
                if json_match:
                    try:
                        extracted = json.loads(json_match.group())
                    except json.JSONDecodeError:
                        extracted = {'raw_text': raw_text}
                else:
                    extracted = {'raw_text': raw_text}

            elif OPENAI_AVAILABLE:
                import openai as _openai
                client = _openai.OpenAI(api_key=api_key)
                response = client.chat.completions.create(
                    model='gpt-4o',
                    messages=[{
                        'role': 'user',
                        'content': [
                            {'type': 'image_url',
                             'image_url': {'url': f'data:{mime_type};base64,{file_b64}'}},
                            {'type': 'text', 'text': ocr_prompt},
                        ]
                    }],
                    max_tokens=1000,
                )
                raw_text = response.choices[0].message.content or ''
                import re
                json_match = re.search(r'\{[\s\S]*\}', raw_text)
                if json_match:
                    try:
                        extracted = json.loads(json_match.group())
                    except json.JSONDecodeError:
                        extracted = {'raw_text': raw_text}
                else:
                    extracted = {'raw_text': raw_text}
            else:
                return request.make_response(
                    json.dumps({'error': 'No hay proveedor de IA disponible para OCR.'}),
                    headers=[('Content-Type', 'application/json')]
                )

            result = {
                'success': True,
                'filename': filename,
                'extract_type': extract_type,
                'extracted': extracted,
            }
            return request.make_response(
                json.dumps(result, ensure_ascii=False),
                headers=[('Content-Type', 'application/json; charset=utf-8')]
            )

        except Exception as e:
            err_safe = _redact(str(e), api_key if 'api_key' in dir() else '')
            _logger.error("OCR error: %s", err_safe)
            return request.make_response(
                json.dumps({'error': err_safe}),
                headers=[('Content-Type', 'application/json')]
            )

    # ── Property Chatter Summary ──────────────────────────────────────────────

    @http.route('/estate/property/chatter_summary', type='jsonrpc', auth='user', methods=['POST'])
    def property_chatter_summary(self, property_id, **kw):
        """Return all data needed by the PropertySummary chatter panel."""
        prop = request.env['estate.property'].browse(property_id).exists()
        if not prop:
            return {}

        fields_to_read = [
            'name', 'state', 'price', 'city', 'offer_type',
            'meeting_count', 'property_invoice_count', 'commission_count',
            'offer_count', 'expense_count', 'days_on_market',
            'price_history_count',
            'ai_property_summary', 'ai_summary_date',
        ]
        data = prop.read(fields_to_read)[0]

        # Leads interesados (estate_crm) — directos + presupuesto compatible
        lead_count = 0
        if 'target_property_id' in request.env['crm.lead']._fields:
            direct = request.env['crm.lead'].sudo().search_count([
                ('target_property_id', '=', property_id),
                ('active', 'in', [True, False]),
            ])
            # Also count leads with matching budget range and same offer_type
            budget_domain = [
                ('target_property_id', '=', False),
                ('active', '=', True),
            ]
            price = prop.price
            if price:
                budget_domain += [
                    '|',
                    ('client_budget', '=', False),
                    '&',
                    ('client_budget', '>=', price * 0.75),
                    ('client_budget', '<=', price * 1.25),
                ]
            if hasattr(prop, 'offer_type') and prop.offer_type:
                budget_domain.append(('type', '=', 'opportunity'))
            budget_leads = request.env['crm.lead'].sudo().search_count(budget_domain)
            lead_count = direct + budget_leads
        data['lead_count'] = lead_count

        # Documentos (estate_document)
        doc_count = 0
        if 'estate.document' in request.env:
            doc_count = request.env['estate.document'].sudo().search_count([
                ('property_id', '=', property_id)
            ])
        data['document_count'] = doc_count

        # Format datetime for display
        if data.get('ai_summary_date'):
            data['ai_summary_date'] = data['ai_summary_date'].strftime('%d %b, %H:%M')

        # ai_property_summary is Html field - ensure it's a plain string
        if data.get('ai_property_summary'):
            data['ai_property_summary'] = str(data['ai_property_summary'])

        return data

    @http.route('/estate/property/regenerate_summary', type='jsonrpc', auth='user', methods=['POST'])
    def property_regenerate_summary(self, property_id, **kw):
        """Trigger AI summary regeneration for the given property."""
        prop = request.env['estate.property'].browse(property_id).exists()
        if not prop:
            return {'error': 'Property not found'}
        try:
            prop.action_generate_property_summary()
            return {'ok': True}
        except Exception as e:
            _logger.error("Error regenerating property summary: %s", str(e))
            return {'error': str(e)}

    @http.route('/estate/property/chat_stream', type='http', auth='user', methods=['POST'], csrf=False)
    def property_chat_stream(self, **kwargs):
        """Streaming SSE endpoint: chat contextual sobre una propiedad específica."""
        import re as _re
        try:
            data = json.loads(request.httprequest.data or '{}')
        except Exception:
            data = {}
        property_id = int(data.get('property_id') or 0)
        question = (data.get('question') or '').strip()
        history_raw = data.get('history') or []

        def _err_stream(msg):
            def _g():
                yield f'data: {json.dumps({"text": msg})}\n\ndata: [DONE]\n\n'
            return request.make_response(_g(), headers=[
                ('Content-Type', 'text/event-stream; charset=utf-8'),
                ('Cache-Control', 'no-cache'),
                ('X-Accel-Buffering', 'no'),
            ])

        if not property_id or not question:
            return _err_stream('Faltan parámetros.')

        prop = request.env['estate.property'].browse(property_id).exists()
        if not prop:
            return _err_stream('Propiedad no encontrada.')

        ICP = request.env['ir.config_parameter'].sudo()
        api_key = ICP.get_param('estate_ai.api_key', '')
        if not api_key or not GEMINI_AVAILABLE:
            return _err_stream('IA no configurada. Configure la API Key en Ajustes.')

        # ── Construir contexto completo de la propiedad ──────────────────────
        p = prop
        tipo = p.property_type_id.name if p.property_type_id else 'Inmueble'
        operacion = 'Venta' if p.offer_type == 'sale' else 'Arriendo'
        estado = dict(p._fields['state'].selection).get(p.state, p.state)
        precio = f"${p.price:,.2f}" if p.price else 'No definido'

        # Historial de precios
        ph_lines = []
        for ph in p.price_history_ids[:10]:
            motivo = ph.change_reason or ''
            ph_lines.append(
                f"  {ph.date.strftime('%d/%m/%Y')}: ${ph.old_price:,.0f}→${ph.new_price:,.0f} "
                f"({ph.change_pct:+.1f}%) {motivo}"
            )

        # Leads interesados
        lead_lines = []
        if 'target_property_id' in request.env['crm.lead']._fields:
            leads = request.env['crm.lead'].search(
                [('target_property_id', '=', p.id)], limit=20
            )
            temp_map = {'cold': 'Frío', 'warm': 'Tibio', 'hot': 'Caliente', 'boiling': 'Hirviendo'}
            for l in leads:
                nombre = (l.partner_id.name if l.partner_id else None) or l.partner_name or l.name or 'Sin nombre'
                budget = f"${l.client_budget:,.0f}" if l.client_budget else 'N/A'
                temp = temp_map.get(l.lead_temperature, l.lead_temperature or '—')
                stage = l.stage_id.name if l.stage_id else 'Sin etapa'
                lead_lines.append(f"  {nombre} | Presupuesto: {budget} | Temperatura: {temp} | Etapa: {stage}")

        # Visitas / citas
        visit_lines = []
        if 'property_id' in request.env['calendar.event']._fields:
            visits = request.env['calendar.event'].search(
                [('property_id', '=', p.id)], order='start desc', limit=10
            )
            rating_map = {'1': 'Muy malo', '2': 'Malo', '3': 'Regular', '4': 'Bueno', '5': 'Excelente'}
            for v in visits:
                attendees = ', '.join(a.partner_id.name for a in v.attendee_ids if a.partner_id) or 'Sin asistentes'
                rating = rating_map.get(str(getattr(v, 'visit_rating', '') or ''), 'Sin calificar')
                visit_lines.append(f"  {v.start.strftime('%d/%m/%Y %H:%M')}: {attendees} — {rating}")

        # Mensajes del chatter
        msgs = request.env['mail.message'].search([
            ('res_id', '=', p.id), ('model', '=', 'estate.property'),
            ('message_type', 'in', ['comment', 'email']),
        ], order='date desc', limit=8)
        msg_lines = []
        for m in msgs:
            body = _re.sub('<[^<]+?>', '', m.body or '').strip()[:200]
            if body:
                msg_lines.append(f"  [{m.date.strftime('%d/%m/%Y')}] {m.author_id.name if m.author_id else 'Sistema'}: {body}")

        context_text = f"""PROPIEDAD: {p.name} | Código: {getattr(p,'ref','') or p.id}
Tipo: {tipo} | Operación: {operacion} | Estado: {estado}
Ubicación: {p.street or ''}, {p.city or ''} {p.state_id.name if p.state_id else ''}
Área: {p.area or 0}m² | Habitaciones: {p.bedrooms or 0} | Baños: {p.bathrooms or 0} | Parqueaderos: {p.parking_spaces or 0}
Precio: {precio} | Días en mercado: {p.days_on_market or 0}
AVM: {'$'+str(f"{p.avm_estimated_price:,.2f}") if p.avm_estimated_price else 'No calculado'}
Propietario: {p.owner_id.name if p.owner_id else 'Sin asignar'}
Asesor: {p.user_id.name if p.user_id else 'Sin asignar'}

HISTORIAL DE PRECIOS ({len(ph_lines)} cambio(s)):
{chr(10).join(ph_lines) if ph_lines else '  Sin cambios registrados'}

LEADS INTERESADOS ({len(lead_lines)} lead(s)):
{chr(10).join(lead_lines) if lead_lines else '  Sin leads registrados'}

VISITAS REALIZADAS ({len(visit_lines)} visita(s)):
{chr(10).join(visit_lines) if visit_lines else '  Sin visitas registradas'}

NOTAS Y MENSAJES RECIENTES:
{chr(10).join(msg_lines) if msg_lines else '  Sin mensajes'}"""

        system_prompt = (
            "Eres un asistente inmobiliario experto. Tienes acceso completo a los datos de la siguiente propiedad. "
            "Responde SIEMPRE en español. Sé directo, preciso y útil. "
            "Cuando menciones personas usa sus nombres exactos. "
            "Si el dato no está disponible, dilo claramente y sugiere cómo obtenerlo.\n\n"
            f"DATOS DE LA PROPIEDAD:\n{context_text}"
        )

        # Historial de conversación previa
        contents = []
        for msg in history_raw[-8:]:  # últimos 8 mensajes
            role = 'user' if msg.get('role') == 'user' else 'model'
            contents.append({'role': role, 'parts': [{'text': msg.get('text', '')}]})
        contents.append({'role': 'user', 'parts': [{'text': question}]})

        # Pre-capturar variables antes de que se cierre el cursor
        _api_key = api_key
        _system = system_prompt
        _contents = contents

        def generate():
            try:
                client = new_genai.Client(
                    api_key=_api_key,
                    http_options=new_genai.types.HttpOptions(api_version='v1beta'),
                )
                response = client.models.generate_content_stream(
                    model='gemini-2.5-flash',
                    contents=_contents,
                    config=new_genai.types.GenerateContentConfig(
                        system_instruction=_system,
                        temperature=0.4,
                        max_output_tokens=1500,
                    ),
                )
                for chunk in response:
                    text = getattr(chunk, 'text', None)
                    if text:
                        yield f'data: {json.dumps({"text": text}, ensure_ascii=False)}\n\n'
                yield 'data: [DONE]\n\n'
            except Exception as exc:
                _logger.error("Property chat stream error: %s", exc)
                yield f'data: {json.dumps({"text": f"Error: {exc}"})}\n\ndata: [DONE]\n\n'

        return request.make_response(generate(), headers=[
            ('Content-Type', 'text/event-stream; charset=utf-8'),
            ('Cache-Control', 'no-cache'),
            ('X-Accel-Buffering', 'no'),
        ])
